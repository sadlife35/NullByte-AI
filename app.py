import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
from fpdf import FPDF
import io
from faker import Faker
import random
import openpyxl
import re
from scipy import stats
import string # Import the string module
from datetime import datetime
from datetime import timedelta
import zipfile # For downloading multiple tables as ZIP
import time # For simulating delays
 
from datetime import timezone # Import timezone explicitly
# Application Details
APP_NAME = "NullByte AI"
APP_TAGLINE = "Synthetic Data Generator"

# Configure Streamlit page
st.set_page_config(
    page_title=f"{APP_NAME} - {APP_TAGLINE}",
    page_icon="🔬",  # A science/tech icon to represent data generation
    layout="wide"
)

# PII Fields to Detect
PII_FIELDS = ['Name', 'Email', 'Phone', 'Address', 'SSN']
DPDP_PII_FIELDS = ['Phone', 'Aadhaar', 'PAN', 'Passport', 'Voter ID']  # DPDP-specific identifiers
AUTO_MASK_KEY = "auto_mask_pii"  # Key for session state - This key is being replaced by DEFAULT_PII_STRATEGY_KEY

# PII Handling Strategies
PII_HANDLING_STRATEGIES = {
    "realistic_fake": "Realistic Fake (Default)",
    "masked": "Masked (e.g., XXXXX1234)",
    "redacted": "Redacted (e.g., [REDACTED])",
    "scramble_column": "Scramble Column (Shuffle existing fakes)" # To be implemented later for schema generation
}
DEFAULT_PII_STRATEGY_KEY = "default_pii_strategy"

# --- NEW: Indian Language Support ---
INDIAN_LOCALES = {
    "en_IN": "English (India)",
    "hi_IN": "हिन्दी (Hindi)",
    "ta_IN": "தமிழ் (Tamil)",
    # "te_IN": "తెలుగు (Telugu)", # Temporarily removed due to Faker AttributeError
    # "bn_IN": "বাংলা (Bengali)", # Temporarily removed due to Faker AttributeError
    # "mr_IN": "मराठी (Marathi)",   # Temporarily removed due to Faker AttributeError
    # "gu_IN": "ગુજરાતી (Gujarati)", # Temporarily removed due to Faker AttributeError
    # "kn_IN": "ಕನ್ನಡ (Kannada)",   # Temporarily removed due to Faker AttributeError
    # "ml_IN": "മലയാളം (Malayalam)", # Temporarily removed due to Faker AttributeError
}
DEFAULT_LOCALE_KEY = "selected_locale"

# --- Session State Initialization ---
if 'table_schemas' not in st.session_state: # Changed from 'schema'
    st.session_state.table_schemas = {} # Dict: {table_name: [field_defs]}
if 'active_table_name' not in st.session_state:
    st.session_state.active_table_name = None
if 'relationships' not in st.session_state:
    st.session_state.relationships = [] # List of relationship dicts
if 'generated_data_frames' not in st.session_state: # New: To store {table_name: DataFrame}
    st.session_state.generated_data_frames = {}
if 'active_display_table_name' not in st.session_state: # New: For displaying selected table
    st.session_state.active_display_table_name = None
if 'edge_cases' not in st.session_state:
    st.session_state.edge_cases = []
if DEFAULT_PII_STRATEGY_KEY not in st.session_state:
    st.session_state[DEFAULT_PII_STRATEGY_KEY] = "realistic_fake"
if 'initial_schema_populated' not in st.session_state: # For template loading logic
    st.session_state.initial_schema_populated = False
if 'prompt_generated_df' not in st.session_state:
    st.session_state.prompt_generated_df = None
if 'uploaded_df_for_schema' not in st.session_state:
    st.session_state.uploaded_df_for_schema = None
if 'num_rows_for_file_upload_tab3' not in st.session_state:
    st.session_state.num_rows_for_file_upload_tab3 = 10 # Default if no file uploaded yet
if 'uploaded_file_name_tab3' not in st.session_state:
    st.session_state.uploaded_file_name_tab3 = None
if 'file_action_tab3' not in st.session_state: # To store user choice in Tab 3
    st.session_state.file_action_tab3 = None
if 'playground_selected_template_name' not in st.session_state:
    st.session_state.playground_selected_template_name = "None (Custom Schema)"
if 'playground_schema_fields' not in st.session_state: # This will store the fields of the selected template
    st.session_state.playground_schema_fields = []
if 'playground_edge_cases' not in st.session_state:
    st.session_state.playground_edge_cases = []
if 'playground_num_rows' not in st.session_state:
    st.session_state.playground_num_rows = 20 # Default for quick scenario testing
if 'playground_generated_df' not in st.session_state:
    st.session_state.playground_generated_df = None
if 'playground_table_name_for_conditions' not in st.session_state: # Fixed name for conditions
    st.session_state.playground_table_name_for_conditions = "PlaygroundScenarioTable"
if DEFAULT_LOCALE_KEY not in st.session_state:
    st.session_state[DEFAULT_LOCALE_KEY] = "en_IN" # Default to English (India)
if 'inferred_schema_from_prompt' not in st.session_state: # For sending schema from prompt to editor
    st.session_state.inferred_schema_from_prompt = None
if 'num_rows_from_prompt' not in st.session_state: # For sending schema from prompt to editor
    st.session_state.num_rows_from_prompt = 10
if 'load_prompt_schema_to_editor' not in st.session_state: # Flag to trigger loading in editor
    st.session_state.load_prompt_schema_to_editor = False
if 'last_processed_prompt_tab1' not in st.session_state: # For Tab 1 regeneration logic
    st.session_state.last_processed_prompt_tab1 = ""
if 'deep_model_type' not in st.session_state:
    st.session_state.deep_model_type = "GAN (Tabular)"
if 'deep_model_data_uploaded' not in st.session_state:
    st.session_state.deep_model_data_uploaded = False # Flag for GAN/VAE
if 'deep_model_prompt' not in st.session_state:
    st.session_state.deep_model_prompt = ""
if 'dp_epsilon' not in st.session_state:
    st.session_state.dp_epsilon = 1.0 # Default Epsilon for DP
if 'nlp_model_prompt' not in st.session_state:
    st.session_state.nlp_model_prompt = "Generate a short story about a futuristic city."
if 'image_model_prompt' not in st.session_state:
    st.session_state.image_model_prompt = "A futuristic cityscape at sunset, digital art"
if 'nlp_model_task' not in st.session_state:
    st.session_state.nlp_model_task = "Creative Writing"
if 'dp_mechanism_numeric' not in st.session_state:
    st.session_state.dp_mechanism_numeric = "Laplace Mechanism"
if 'dp_mechanism_categorical' not in st.session_state:
    st.session_state.dp_mechanism_categorical = "Randomized Response"
if 'benchmark_fidelity_score' not in st.session_state:
    st.session_state.benchmark_fidelity_score = None
if 'benchmark_utility_score' not in st.session_state:
    st.session_state.benchmark_utility_score = None
if 'benchmark_privacy_score' not in st.session_state: # Lower is better for risk, or higher for protection
    st.session_state.benchmark_privacy_score = None
if 'advanced_lab_selection' not in st.session_state:
    st.session_state.advanced_lab_selection = "🤖 AI-Powered Generation" # Default selection, ensure it's one of the options
if 'federated_participants_list' not in st.session_state:
    st.session_state.federated_participants_list = [] # List of dicts: {"name": "P1", "trained": False}
if 'federated_new_participant_name' not in st.session_state:
    st.session_state.federated_new_participant_name = ""
if 'federated_aggregation_done' not in st.session_state:
    st.session_state.federated_aggregation_done = False
if 'federated_global_schema_template_output' not in st.session_state:
    st.session_state.federated_global_schema_template_output = "None (Custom Schema)"
if 'federated_num_rows_output' not in st.session_state:
    st.session_state.federated_num_rows_output = 100
if 'federated_generated_df_output' not in st.session_state:
    st.session_state.federated_generated_df_output = None
if 'marketplace_templates' not in st.session_state:
    st.session_state.marketplace_templates = [ # Pre-populate with some examples
        {"name": "Realistic E-commerce Transactions", "description": "Detailed e-commerce data with customer behavior.", "author": "CommunityUser1", "rating": 4.5, "downloads": 120, "discussions": 15, "trust_badge": "Verified"},
        {"name": "Indian Patient Health Records (Anonymized Demo)", "description": "A rich dataset for healthcare analytics in an Indian context, PII faked. Suitable for demonstrating healthcare data synthesis.", "author": "HealthcareAI_Org", "rating": 4.8, "downloads": 250, "discussions": 30, "trust_badge": "Trusted Partner"},
        {"name": "Indian Fintech UPI Transactions (Simplified)", "description": "Simulated UPI transaction data for fintech analysis, including merchant details and transaction status. PII is faked.", "author": "FinSecureDev", "rating": 4.6, "downloads": 180, "discussions": 22, "trust_badge": "Verified"},
        {"name": "Simple IoT Sensor Data Stream", "description": "Basic sensor readings for IoT projects.", "author": "MakerPro", "rating": 4.0, "downloads": 80, "discussions": 5, "trust_badge": "Community Contributed"},
    ]
if 'marketplace_new_template_name' not in st.session_state:
    st.session_state.marketplace_new_template_name = ""
if 'marketplace_new_template_description' not in st.session_state:
    st.session_state.marketplace_new_template_description = ""
if 'marketplace_new_template_author' not in st.session_state: # In a real system, this would be the logged-in user
    st.session_state.marketplace_new_template_author = "YourName"
if 'cloud_deploy_db_type' not in st.session_state:
    st.session_state.cloud_deploy_db_type = "Supabase (PostgreSQL)"
if 'cloud_deploy_connection_string' not in st.session_state: # Conceptual
    st.session_state.cloud_deploy_connection_string = "your_db_connection_string_here"
if 'cloud_deploy_table_name' not in st.session_state:
    st.session_state.cloud_deploy_table_name = "synthetic_data_table"
if 'cloud_deploy_data_source' not in st.session_state: # To indicate which generated data to deploy
    st.session_state.cloud_deploy_data_source = None
if 'num_rows_smart_schema_editor' not in st.session_state: # For the number input in Smart Schema Editor
    st.session_state.num_rows_smart_schema_editor = 10


if 'data_generation_focus' not in st.session_state: # New: For Indian vs Global data
    st.session_state.data_generation_focus = "indian" # Default to Indian context

# --- Constants for Dependency Logic ---
STATE_CITY_MAP = {
    "California": ["Los Angeles", "San Francisco", "San Diego", "Sacramento"],
    "New York": ["New York City", "Buffalo", "Rochester", "Albany"],
    "Texas": ["Houston", "Dallas", "Austin", "San Antonio"],
    "Maharashtra": ["Mumbai", "Pune", "Nagpur", "Nashik"],
    "Delhi": ["New Delhi", "Gurugram", "Noida"], # Note: Delhi is a Union Territory, New Delhi is its capital
    "Karnataka": ["Bengaluru", "Mysuru", "Hubballi"],
}
COUNTRY_CURRENCY_MAP = {
    "India": "INR", "United States": "USD", "United Kingdom": "GBP",
    "Germany": "EUR", "Japan": "JPY", "China": "CNY",
}

# ---- Header ----
st.title(f"{APP_NAME}: {APP_TAGLINE}")
st.markdown(f"*Generate AI-ready, bias-checked, privacy-compliant synthetic datasets in minutes.*")

# Add a brief app description
st.markdown("""
<style>
.highlight-box {
    background-color: #0d1117; /* GitHub dark theme background */
    border-left: 5px solid #4e8098;
    padding: 15px;
    border-radius: 5px;
    margin-bottom: 20px;
}
.dpdp-warning {
    background-color: #fff3cd;
    border-left: 5px solid #ffc107;
    padding: 5px;
    border-radius: 3px;
    margin: 5px 0;
    font-size: 0.9em;
}
.indian-id {
    font-family: monospace;
    background-color: #f8f9fa;
    padding: 2px 4px;
    border-radius: 3px;
}
</style>
<div class="highlight-box">
**About NullByte AI**<br>
NullByte AI helps data scientists, researchers, solo entreprenuers and businesses create privacy-preserving synthetic datasets
with advanced bias detection and ethical compliance checks.
</div>
""", unsafe_allow_html=True)

st.caption("""
    ⚠️ **Disclaimer:** NullByte AI generates synthetic data. While it includes features for PII detection, masking, and bias checks,
    users are solely responsible for ensuring that their use of the generated data complies with all applicable laws and regulations,
    including data privacy laws like the DPDP Act (India), GDPR, CCPA, etc. The ethical AI scores and reports provided are
    for guidance and illustrative purposes only and do not constitute legal advice. Always perform thorough validation for your specific use case.
""", unsafe_allow_html=True)

# --- Sidebar Guide ---
def display_sidebar_guide():
    st.sidebar.title("📘 NullByte AI Guide")
    st.sidebar.markdown("Welcome to NullByte AI! Here's a quick guide to get you started:")

    st.sidebar.subheader("1. Text-based Generation")
    st.sidebar.markdown(
        "- **Describe your data:** Use the main text input (e.g., '100 rows of customer data with name, email, and purchase_amount').\n"
        "- **Focus & Language:** Select 'Indian Context' or 'Global/Random' and your preferred language for generated names, addresses, etc.\n"
        "- **Reproducibility:** Check 'Use fixed random seed' for consistent datasets.\n"
        "- **View & Download:** The generated data will appear below. You can edit column names, values, and download it.\n"
        "- **Refine in Editor:** Click 'Send Inferred Schema to Smart Editor' to customize the auto-detected schema."
    )

    st.sidebar.subheader("2. Smart Schema Editor")
    st.sidebar.markdown(
        "- **Manage Tables:** Add, delete, or select tables to edit.\n"
        "- **Load Templates:** Choose a pre-defined schema from the 'Load Schema Template' dropdown or from the 'Community Gallery' tab.\n"
        "- **Define Fields:** For the active table, add fields, specify their 'Name', 'Type', and 'Constraint' (e.g., range for numbers, categories for text).\n"
        "- **PII Handling:** Set a default PII strategy and override it per sensitive field (e.g., 'Masked', 'Redacted').\n"
        "- **Relationships:** Define one-to-many relationships between tables using Primary Keys (PK) and Foreign Keys (FK).\n"
        "- **Edge Cases:** Inject specific data patterns into a percentage of rows based on conditions.\n"
        "- **Generate:** Set the number of root rows and click 'Generate Data from All Schemas'."
    )

    st.sidebar.subheader("3. File-based Generation")
    st.sidebar.markdown(
        "- **Upload:** Upload your CSV or XLSX file.\n"
        "- **Choose Action:** \n"
        "  - **Generate Synthetic Data:** Creates new synthetic rows based on your file's structure and appends them. PII-like columns are faked.\n"
        "  - **Check Compliance:** Analyzes your original uploaded file for bias and PII.\n"
        "- **Drift Analysis:** If you generate synthetic data, a drift analysis compares the synthetic part to your original data."
    )

    st.sidebar.subheader("4. Scenario Playground")
    st.sidebar.markdown(
        "- **Test Ideas:** Quickly design and test specific data scenarios or edge cases.\n"
        "- **Select Template:** Choose a base schema.\n"
        "- **Define Edge Cases:** Specify conditions to inject into a percentage of the data.\n"
        "- **Generate:** Create a small dataset to see the results."
    )

    st.sidebar.subheader("5. Community Gallery")
    st.sidebar.markdown(
        "- **Explore Schemas:** Browse various predefined schema templates.\n"
        "- **Load to Editor:** Click 'Load ... into Smart Schema Editor' to use a template as a starting point in Tab 2."
    )
    
    st.sidebar.subheader("🔬 Advanced Lab")
    st.sidebar.markdown(
        "- **Experimental Features:** Explore conceptual tools like AI-Powered Generation (GANs, VAEs, GPT), Differential Privacy, Quality Benchmarking, Federated Generation, and more.\n"
        "- **Note:** These are currently placeholders for future development and illustrate advanced capabilities."
    )

    st.sidebar.markdown("---")
    st.sidebar.info("💡 **Tip:** Use the 'Ethical AI Dashboard' in relevant tabs to check for bias and PII risks.")

# Call the function to display the guide in the sidebar
display_sidebar_guide()


# Text Input for Dataset Description
prompt = st.text_input("Describe the dataset you want (e.g., '10 rows of employee data with name, age, salary, email')")

# --- NEW: Data Generation Focus ---
st.session_state.data_generation_focus = st.radio(
    "Select Data Generation Focus:",
    options=["indian", "global"],
    format_func=lambda x: "Indian Context (e.g., names, addresses, +91 phones)" if x == "indian" else "Global/Random (e.g., US-style names, addresses, generic phones)",
    index=["indian", "global"].index(st.session_state.data_generation_focus),
    horizontal=True,
    key="data_focus_selector"
)
# --- NEW: Language Selection ---
selected_locale_from_ui = st.selectbox(
    "Select Language for Generated Data (Names, Addresses, etc.):",
    options=list(INDIAN_LOCALES.keys()),
    format_func=lambda x: INDIAN_LOCALES[x],
    index=list(INDIAN_LOCALES.keys()).index(st.session_state[DEFAULT_LOCALE_KEY]),
    key="language_selector_main",
    help="If 'Indian Context' focus is selected, this chooses the specific Indian locale. If 'Global/Random' focus, this selection has less impact as Faker will use a global default (e.g., en_US)."
)

# Update session state if locale changed
if selected_locale_from_ui != st.session_state[DEFAULT_LOCALE_KEY]:
    st.session_state[DEFAULT_LOCALE_KEY] = selected_locale_from_ui
    # No explicit rerun needed here, Streamlit handles it.
    # Faker will be re-initialized below based on both data_generation_focus and selected_locale_from_ui.
    # Streamlit will rerun, and Faker will be re-initialized with the new locale.

# Reproducibility Option
use_fixed_seed = st.checkbox("Use fixed random seed for reproducible dataset")

# --- Seed Management & Faker Initialization ---
if use_fixed_seed:
    random.seed(42)
    Faker.seed(42)
    np.random.seed(42) # Add this for NumPy
else:
    random.seed(None)
    Faker.seed(None)
    np.random.seed(None) # Add this for NumPy

# Initialize Faker based on the selected focus and locale
try:
    if st.session_state.data_generation_focus == "indian":
        current_locale_for_faker = st.session_state[DEFAULT_LOCALE_KEY]
        fake = Faker(current_locale_for_faker)
        # st.caption(f"Faker initialized with: {current_locale_for_faker} (Indian Focus)") # For debugging
    else: # global focus
        current_locale_for_faker = 'en_US' # Default global locale
        fake = Faker(current_locale_for_faker)
        # st.caption(f"Faker initialized with: {current_locale_for_faker} (Global Focus)") # For debugging
except AttributeError as e:
    st.error(f"Error initializing Faker with locale '{current_locale_for_faker}': {e}. "
             f"This might indicate that the locale is not fully supported by your Faker installation. "
             f"Falling back to 'en_IN'. Please check Faker documentation or update the library if you need this locale.")
    st.session_state[DEFAULT_LOCALE_KEY] = "en_IN" # Fallback Indian locale
    fake = Faker("en_IN") # Re-initialize with fallback

# Field type definitions (FIELD_TYPES remains the same as it's for UI display of types)
FIELD_TYPES = {
    "string": "Text String",
    "email": "Email Address",
    "int": "Integer Number",
    "float": "Decimal Number",
    "date": "Date",
    "category": "Category/Enum",
    "phone": "Phone Number",
    "address": "Address",
    "name": "Full Name",
    "aadhaar": "Aadhaar Number",
    "pan": "PAN Number",
    "passport": "Passport Number",
    "voterid": "Voter ID",
    "ifsc": "IFSC Code",
    "upi": "UPI ID",
    "animal_name": "Animal Name (Pet Name)" # New type for animal names
}

def is_dpdp_pii(field_name):
    """Check if a field name matches DPDP-sensitive PII."""
    field_lower = field_name.lower()
    return (
        'phone' in field_lower or
        'aadhaar' in field_lower or
        'pan' in field_lower or
        'passport' in field_lower or
        'voter' in field_lower or
        'ifsc' in field_lower or
        'upi' in field_lower
    )

def mask_pii(value, field_type):
    """Mask PII data according to field type."""
    if pd.isna(value) or not str(value).strip(): # Check for empty or NaN
        return value

    if field_type == "phone":
        if len(str(value)) >= 4:
            return f"XXXXXX{str(value)[-4:]}"
        return "XXXXXX"
    elif field_type in ["aadhaar", "pan", "passport", "voterid", "ifsc", "upi"]:
        if len(str(value)) >= 4:
            return f"XXXXXX{str(value)[-4:]}"
        return "XXXXXX"
    elif field_type == "email":
        parts = str(value).split('@')
        if len(parts) == 2:
            return f"{parts[0][0]}***@{parts[1]}"
        return "***@***"
    return value

def validate_constraint(field_type, constraint):
    """Validate if the constraint is valid for the given field type."""
    if field_type in ["int", "float"]:
        # Check if constraint is in format "min-max"
        pattern = r"^\s*(-?\d+(?:\.\d+)?)\s*-\s*(-?\d+(?:\.\d+)?)\s*$"
        match = re.match(pattern, constraint)
        if match:
            min_val, max_val = match.groups()
            try:
                min_val = float(min_val)
                max_val = float(max_val)
                return True # Allow min == max for single value constraint
            except ValueError:
                return False
        return False
    elif field_type == "date":
        # Check if constraint is in format "YYYY-MM-DD - YYYY-MM-DD"
        pattern = r"^\s*(\d{4}-\d{2}-\d{2})\s*-\s*(\d{4}-\d{2}-\d{2})\s*$"
        match = re.match(pattern, constraint)
        if match:
            start_date, end_date = match.groups()
            try:
                start = datetime.strptime(start_date, "%Y-%m-%d")
                end = datetime.strptime(end_date, "%Y-%m-%d")
                return True
            except ValueError:
                return False
        return False
    elif field_type == "category":
        # Check if constraint is a comma-separated list of values
        values = [v.strip() for v in constraint.split(",") if v.strip()]
        return len(values) > 0
    elif field_type == "string":
        # For string type, constraint is optional
        return True
    elif field_type in ["email", "phone", "address", "name", "aadhaar", "pan", "passport", "voterid", "ifsc", "upi"]:
        # These types don't need constraints for generation, but validate if provided
        if constraint: # If a constraint is provided, check if it's a simple value for == edge case
             return True # We'll handle constraint validation for these types in the generator if needed
        return True
    return False

def generate_aadhaar():
    """Generate a fake but realistic Aadhaar number (12 digits, optionally with spaces)"""
    aadhaar = ''.join([str(random.randint(0, 9)) for _ in range(12)])
    # Randomly add spaces for formatting (40% chance)
    if random.random() < 0.4:
        aadhaar = f"{aadhaar[:4]} {aadhaar[4:8]} {aadhaar[8:12]}"
    return aadhaar

def generate_pan():
    """Generate a fake but realistic PAN number (AAAAA0000A format)"""
    first_five = ''.join([random.choice('ABCDEFGHIJKLMNOPQRSTUVWXYZ') for _ in range(5)])
    four_digits = ''.join([str(random.randint(0, 9)) for _ in range(4)])
    last_char = random.choice('ABCDEFGHIJKLMNOPQRSTUVWXYZ')
    return f"{first_five}{four_digits}{last_char}"

def generate_passport():
    """Generate a fake but realistic Indian passport number"""
    first_char = random.choice('ABCDEFGHIJKLMNOPQRSTUVWXYZ')
    seven_digits = ''.join([str(random.randint(0, 9)) for _ in range(7)])
    return f"{first_char}{seven_digits}"

def generate_voter_id():
    """Generate a fake but realistic Voter ID (2 letters followed by 7 digits)"""
    first_two = ''.join([random.choice('ABCDEFGHIJKLMNOPQRSTUVWXYZ') for _ in range(2)])
    seven_digits = ''.join([str(random.randint(0, 9)) for _ in range(7)])
    return f"{first_two}{seven_digits}"

def generate_ifsc():
    """Generate a fake but realistic IFSC code (4 letters + 0 + 6 digits)"""
    bank_code = ''.join([random.choice('ABCDEFGHIJKLMNOPQRSTUVWXYZ') for _ in range(4)])
    branch_code = ''.join([str(random.randint(0, 9)) for _ in range(6)])
    return f"{bank_code}0{branch_code}"

def generate_upi():
    """Generate a fake but realistic UPI ID"""
    username = fake.user_name()
    domain = random.choice(['@oksbi', '@paytm', '@ybl', '@axl', '@ibl'])
    return f"{username}{domain}".lower()

# --- Value Generation Helper Functions ---
def _generate_string_value(constraint, field_name, pii_strategy, edge_condition=None, full_field_schema=None):
    """
    Generates a string value based on the field name and any hints in the schema.
    Handles business-specific patterns like API paths, version numbers, etc.
    """
    field_lower = field_name.lower()
    
    # Handle edge conditions first
    if edge_condition and edge_condition.get('operator') == '==':
        return str(edge_condition['value'])
    
    # Handle business-specific patterns
    if full_field_schema and full_field_schema.get('_hint_api_path'):
        # Generate realistic API endpoint paths
        api_versions = ['v1', 'v2', 'v3']
        api_resources = ['users', 'products', 'orders', 'customers', 'payments', 'inventory', 'analytics']
        api_actions = ['get', 'create', 'update', 'delete', 'list', 'search', 'filter']
        version = random.choice(api_versions)
        resource = random.choice(api_resources)
        action = random.choice(api_actions)
        return f"/{version}/{resource}/{action}"
    
    elif full_field_schema and full_field_schema.get('_hint_version'):
        # Generate semantic version numbers
        major = random.randint(1, 5)
        minor = random.randint(0, 9)
        patch = random.randint(0, 9)
        return f"{major}.{minor}.{patch}"
    
    elif "connection_string" in field_lower:
        # Generate database connection strings
        db_types = ['postgresql', 'mysql', 'mongodb', 'redis']
        db_type = random.choice(db_types)
        user = fake.user_name()
        password = fake.password()
        host = fake.hostname()
        port = random.randint(1024, 65535)
        db_name = fake.word()
        return f"{db_type}://{user}:{password}@{host}:{port}/{db_name}"
    
    elif "sku" in field_lower:
        # Generate product SKUs
        prefix = "".join(random.choices(string.ascii_uppercase, k=3))
        middle = "".join(random.choices(string.ascii_uppercase + string.digits, k=3))
        suffix = "".join(random.choices(string.ascii_uppercase, k=2))
        return f"{prefix}-{middle}-{suffix}"
    
    elif "campaign_id" in field_lower:
        # Generate marketing campaign IDs
        year = datetime.now().year
        quarter = f"Q{(datetime.now().month-1)//3 + 1}"
        campaign_type = random.choice(['EMAIL', 'SOCIAL', 'PPC', 'SEO'])
        number = random.randint(1000, 9999)
        return f"{campaign_type}-{year}-{quarter}-{number}"
    
    elif "job_title" in field_lower:
        # Generate realistic job titles
        levels = ['Junior', 'Senior', 'Lead', 'Principal', 'Staff', 'Chief']
        roles = ['Software Engineer', 'Data Scientist', 'Product Manager', 'DevOps Engineer', 
                'UX Designer', 'Business Analyst', 'Project Manager', 'Solutions Architect']
        specialties = ['', 'Backend', 'Frontend', 'Full Stack', 'Cloud', 'Security', 'Mobile']
        
        level = random.choice(levels)
        role = random.choice(roles)
        specialty = random.choice(specialties)
        
        if specialty:
            return f"{level} {specialty} {role}"
        return f"{level} {role}"
    
    # Handle Faker-specific patterns
    elif full_field_schema and full_field_schema.get('is_faker_url'):
        return fake.url()
    elif full_field_schema and full_field_schema.get('is_faker_ipv4'):
        return fake.ipv4()
    elif full_field_schema and full_field_schema.get('is_faker_color_name'):
        return fake.color_name()
    elif full_field_schema and full_field_schema.get('is_faker_job'):
        return fake.job()
    elif full_field_schema and full_field_schema.get('is_faker_company'):
        return fake.company()
    
    # Handle inferred alphanumeric IDs
    elif full_field_schema and full_field_schema.get('_is_inferred_alphanum_id'):
        prefix = full_field_schema.get('_inferred_prefix', 'ID-')
        length = full_field_schema.get('length', 8)
        chars = string.ascii_uppercase + string.digits
        return prefix + ''.join(random.choices(chars, k=length))
    
    # Default string generation
    if constraint:
        # If there's a constraint, try to use it
        try:
            return str(constraint)
        except:
            pass
    
    # Generate a random string based on the field name
    if any(kw in field_lower for kw in ['name', 'title', 'description']):
        return fake.sentence(nb_words=3)
    elif any(kw in field_lower for kw in ['code', 'id', 'key']):
        return fake.uuid4()
    else:
        return fake.word()

def _generate_int_value(constraint, field_name, pii_strategy, edge_condition=None):
    min_default, max_default = 1, 1000
    min_default_initial, max_default_initial = min_default, max_default # Store initial defaults

    if constraint:
        pattern = r"^\s*(-?\d+(?:\.\d+)?)\s*-\s*(-?\d+(?:\.\d+)?)\s*$"
        match = re.match(pattern, constraint)
        if match:
            min_val_str, max_val_str = match.groups()
            try:
                parsed_min = int(float(min_val_str)) # Use float conversion for flexibility then int
                parsed_max = int(float(max_val_str))
                if parsed_min <= parsed_max:
                    min_default = parsed_min
                    max_default = parsed_max
                else:
                    st.warning(f"Invalid constraint range for INT field '{field_name}': min ({parsed_min}) > max ({parsed_max}). Using default range {min_default_initial}-{max_default_initial}.")
            except ValueError:
                st.warning(f"Error converting constraint parts '{min_val_str}', '{max_val_str}' to int for '{field_name}'. Using default range.")
        else:
            st.warning(f"Malformed INT constraint '{constraint}' for field '{field_name}'. Using default range {min_default_initial}-{max_default_initial}.")
    
    # min_default, max_default now hold either parsed values or initial defaults for the constraint
    
    min_target, max_target = min_default, max_default

    if edge_condition:
        op = edge_condition['operator']
        try:
            val = int(edge_condition['value'])
            if op == '>': min_target = val + 1
            elif op == '<': max_target = val - 1
            elif op == '==': min_target, max_target = val, val
            elif op == '>=': min_target = val
            elif op == '<=': max_target = val
            # Note: != is complex for direct generation, would require generate-then-check-retry.
        except ValueError:
             st.warning(f"Edge case value '{edge_condition['value']}' for '{field_name}' is not a valid integer. Ignoring edge case for this field.")


    if "age" in field_name.lower() or "salary" in field_name.lower(): # Ensure non-negative for these
        min_target = max(0, min_target)

    if min_target > max_target: # If edge case created invalid range, try to make a small valid one
        if edge_condition and edge_condition.get('operator') == '==': return int(edge_condition['value']) # Exact value if possible
        st.warning(f"Edge case for '{field_name}' created an invalid range ({min_target}-{max_target}). Attempting to adjust or use original constraint.")
        # Attempt to create a small valid range around the edge value if possible, else revert
        if edge_condition and edge_condition.get('operator') in ['>', '>=']: max_target = min_target + abs(min_target // 20 or 10) + 10
        elif edge_condition and edge_condition.get('operator') in ['<', '<=']: min_target = max_target - abs(max_target // 20 or 10) - 10
        if min_target > max_target: # Still invalid, revert to original default or constraint
            min_target, max_target = min_default, max_default
            if "age" in field_name.lower() or "salary" in field_name.lower(): min_target = max(0, min_target)
            if min_target > max_target: return random.randint(0,100) # Absolute fallback

    generated_value = random.randint(min_target, max_target)

    # --- Conceptual Differential Privacy (Laplace Mechanism) ---
    if (st.session_state.get('advanced_lab_selection') == "🛡️ Differential Privacy" and
        st.session_state.get('dp_mechanism_numeric') == "Laplace Mechanism" and
        st.session_state.get('dp_epsilon', 0) > 0):

        epsilon = st.session_state.dp_epsilon
        # Sensitivity: For direct perturbation of a value within a range [min, max],
        # sensitivity can be considered max - min.
        # If min_target == max_target, sensitivity is 0, no noise added.
        sensitivity = float(max_target - min_target)

        if sensitivity > 0: # Only add noise if there's a range
            noise = np.random.laplace(0, sensitivity / epsilon)
            noisy_value = generated_value + noise
            noisy_value_int = int(round(noisy_value))
            # Clip to original target range (Note: simple clipping can affect formal DP guarantees)
            generated_value = max(min_target, min(noisy_value_int, max_target))
            # st.sidebar.caption(f"DP Applied to {field_name}: val={generated_value}, noise={noise:.2f}, sens={sensitivity}, eps={epsilon}") # For debugging

    return generated_value


def _generate_float_value(constraint, field_name, pii_strategy, edge_condition=None):
    min_default, max_default = 1.0, 1000.0
    min_default_initial, max_default_initial = min_default, max_default # Store initial defaults

    if constraint:
        pattern = r"^\s*(-?\d+(?:\.\d+)?)\s*-\s*(-?\d+(?:\.\d+)?)\s*$"
        match = re.match(pattern, constraint)
        if match:
            min_val_str, max_val_str = match.groups()
            try:
                parsed_min = float(min_val_str)
                parsed_max = float(max_val_str)
                if parsed_min <= parsed_max:
                    min_default = parsed_min
                    max_default = parsed_max
                else:
                    st.warning(f"Invalid constraint range for FLOAT field '{field_name}': min ({parsed_min}) > max ({parsed_max}). Using default range {min_default_initial}-{max_default_initial}.")
            except ValueError:
                st.warning(f"Error converting constraint parts '{min_val_str}', '{max_val_str}' to float for '{field_name}'. Using default range.")
        else:
            st.warning(f"Malformed FLOAT constraint '{constraint}' for field '{field_name}'. Using default range {min_default_initial}-{max_default_initial}.")

    # min_default, max_default now hold either parsed values or initial defaults for the constraint
    
    min_target, max_target = min_default, max_default

    if edge_condition:
        op = edge_condition['operator']
        try:
            val = float(edge_condition['value'])
            if op == '>': min_target = val + 0.01 # Small epsilon for float
            elif op == '<': max_target = val - 0.01
            elif op == '==': min_target, max_target = val, val
            elif op == '>=': min_target = val
            elif op == '<=': max_target = val
        except ValueError:
             st.warning(f"Edge case value '{edge_condition['value']}' for '{field_name}' is not a valid float. Ignoring edge case for this field.")


    if "age" in field_name.lower() or "salary" in field_name.lower() or "price" in field_name.lower():
        min_target = max(0.0, min_target)

    if min_target > max_target:
        if edge_condition and edge_condition.get('operator') == '==': return float(edge_condition['value']) # Exact value if possible
        st.warning(f"Edge case for '{field_name}' created an invalid float range ({min_target}-{max_target}). Attempting to adjust or use original constraint.")
        if edge_condition and edge_condition.get('operator') in ['>', '>=']: max_target = min_target + abs(min_target / 20.0 or 10.0) + 10.0
        elif edge_condition and edge_condition.get('operator') in ['<', '<=']: min_target = max_target - abs(max_target / 20.0 or 10.0) - 10.0
        if min_target > max_target:
            min_target, max_target = min_default, max_default
            if "age" in field_name.lower() or "salary" in field_name.lower() or "price" in field_name.lower(): min_target = max(0.0, min_target)
            if min_target > max_target: return round(random.uniform(0.0,100.0), 2)

    generated_value = round(random.uniform(min_target, max_target), 2)

    # --- Conceptual Differential Privacy (Laplace Mechanism) ---
    if (st.session_state.get('advanced_lab_selection') == "🛡️ Differential Privacy" and
        st.session_state.get('dp_mechanism_numeric') == "Laplace Mechanism" and
        st.session_state.get('dp_epsilon', 0) > 0):

        epsilon = st.session_state.dp_epsilon
        # Sensitivity: For direct perturbation of a value within a range [min, max],
        # sensitivity can be considered max - min.
        sensitivity = float(max_target - min_target)

        if sensitivity > 0: # Only add noise if there's a range
            noise = np.random.laplace(0, sensitivity / epsilon)
            noisy_value = generated_value + noise
            # Clip to original target range (Note: simple clipping can affect formal DP guarantees)
            # Ensure the clipped value still respects the float nature (e.g. precision)
            generated_value = round(max(min_target, min(noisy_value, max_target)), 2)
            # st.sidebar.caption(f"DP Applied to {field_name}: val={generated_value}, noise={noise:.2f}, sens={sensitivity}, eps={epsilon}") # For debugging

    return generated_value


def _generate_date_value(constraint, field_name, pii_strategy, edge_condition=None):
    start_default = datetime.now() - timedelta(days=365)
    end_default = datetime.now()

    if constraint and validate_constraint("date", constraint):
        start_date_str, end_date_str = map(str.strip, constraint.split('-'))
        try:
            start_default = datetime.strptime(start_date_str, "%Y-%m-%d")
            end_default = datetime.strptime(end_date_str, "%Y-%m-%d")
        except ValueError:
            st.warning(f"Invalid date constraint format for '{field_name}'. Using default date range.")


    start_target, end_target = start_default, end_default

    if edge_condition:
        op = edge_condition['operator']
        try:
            val_date = datetime.strptime(str(edge_condition['value']), "%Y-%m-%d")
            if op == '>': start_target = val_date + timedelta(days=1)
            elif op == '<': end_target = val_date - timedelta(days=1)
            elif op == '==': start_target, end_target = val_date, val_date
            elif op == '>=': start_target = val_date
            elif op == '<=': end_target = val_date
        except ValueError:
             st.warning(f"Edge case value '{edge_condition['value']}' for '{field_name}' is not a valid date (YYYY-MM-DD). Ignoring edge case for this field.")


    if start_target > end_target:
        if edge_condition and edge_condition.get('operator') == '==': return val_date.strftime("%Y-%m-%d") # Exact value if possible
        st.warning(f"Edge case for '{field_name}' created an invalid date range. Using original constraint.")
        start_target, end_target = start_default, end_default
        if start_target > end_target: return fake.date_this_year().strftime("%Y-%m-%d") # Absolute fallback

    days_between = (end_target - start_target).days
    if days_between < 0: days_between = 0 # Should not happen if logic above is correct
    random_days = random.randint(0, days_between)
    return (start_target + timedelta(days=random_days)).strftime("%Y-%m-%d")

def _generate_category_value(constraint, field_name, pii_strategy, edge_condition=None):
    if edge_condition and edge_condition.get('operator') == '==':
        # Ensure the edge case value is one of the possible categories if constraint exists
        if constraint:
            values = [v.strip() for v in constraint.split(",") if v.strip()]
            if str(edge_condition['value']) in values:
                return str(edge_condition['value'])
            else: # Edge case value not in allowed categories, warn and pick from original
                st.warning(f"Edge case value '{edge_condition['value']}' for '{field_name}' not in allowed categories. Picking from original constraint.")
        else: # No original constraint, so edge case value is fine
            return str(edge_condition['value'])

    if constraint:
        values = [v.strip() for v in constraint.split(",") if v.strip()]
        if values: return random.choice(values)
    return random.choice(["Option A", "Option B", "Option C"])

def _apply_pii_strategy_to_value(value, field_type, pii_strategy):
    """Applies masking or redaction to a generated value based on the chosen strategy."""
    if pii_strategy == "masked":
        return mask_pii(value, field_type)
    elif pii_strategy == "redacted":
        return "[REDACTED]"
    # "realistic_fake" is the default if no other strategy applies or if value is already faked
    # "scramble_column" is handled *after* generation
    return value

def _generate_email_value(constraint, field_name, pii_strategy, edge_condition=None):
    # Edge cases for email (like '==') are less common for generation, but could be supported
    if edge_condition and edge_condition.get('operator') == '==':
         val = str(edge_condition['value'])
    else:
        val = fake.email()
    return _apply_pii_strategy_to_value(val, "email", pii_strategy)

def _generate_phone_value(constraint, field_name, pii_strategy, edge_condition=None):
    if edge_condition and edge_condition.get('operator') == '==':
         val = str(edge_condition['value'])
    else:
        if st.session_state.data_generation_focus == "indian":
            val = f"+91 {random.randint(7000, 9999)}{random.randint(100000, 999999)}"
        else: # global focus
            val = fake.phone_number() # Uses the globally configured 'fake' object (e.g., en_US)
    return _apply_pii_strategy_to_value(val, "phone", pii_strategy)

def _generate_address_value(constraint, field_name, pii_strategy, edge_condition=None):
    if edge_condition and edge_condition.get('operator') == '==':
         val = str(edge_condition['value'])
    else:
        val = fake.address().replace('\n', ', ')
    # Address is typically faked, masking/redaction can be applied if needed via strategy
    return _apply_pii_strategy_to_value(val, "address", pii_strategy)

def _generate_name_value(constraint, field_name, pii_strategy, edge_condition=None):
    # This function now needs the full field_schema to access prefix/suffix
    # For simplicity in this diff, we'll assume field_schema is passed if called from generate_value
    # The call signature in VALUE_GENERATOR_FUNCTIONS might need adjustment if we pass full schema there.
    # For now, we'll assume field_name can be used to look up its schema if needed, or this logic moves.
    if edge_condition and edge_condition.get('operator') == '==':
         val = str(edge_condition['value'])
    else:
        val = fake.name()
    return _apply_pii_strategy_to_value(val, "name", pii_strategy)

def _generate_aadhaar_value(constraint, field_name, pii_strategy, edge_condition=None):
    if edge_condition and edge_condition.get('operator') == '==':
         val = str(edge_condition['value'])
    else:
        val = generate_aadhaar()
    return _apply_pii_strategy_to_value(val, "aadhaar", pii_strategy)

def _generate_pan_value(constraint, field_name, pii_strategy, edge_condition=None):
    if edge_condition and edge_condition.get('operator') == '==':
         val = str(edge_condition['value'])
    else:
        val = generate_pan()
    return _apply_pii_strategy_to_value(val, "pan", pii_strategy)

def _generate_passport_value(constraint, field_name, pii_strategy, edge_condition=None):
    if edge_condition and edge_condition.get('operator') == '==':
         val = str(edge_condition['value'])
    else:
        val = generate_passport()
    return _apply_pii_strategy_to_value(val, "passport", pii_strategy)

def _generate_voterid_value(constraint, field_name, pii_strategy, edge_condition=None):
    if edge_condition and edge_condition.get('operator') == '==':
         val = str(edge_condition['value'])
    else:
        val = generate_voter_id()
    return _apply_pii_strategy_to_value(val, "voterid", pii_strategy)

def _generate_ifsc_value(constraint, field_name, pii_strategy, edge_condition=None):
    if edge_condition and edge_condition.get('operator') == '==':
         val = str(edge_condition['value'])
    else:
        val = generate_ifsc()
    return _apply_pii_strategy_to_value(val, "ifsc", pii_strategy)

def _generate_upi_value(constraint, field_name, pii_strategy, edge_condition=None):
    if edge_condition and edge_condition.get('operator') == '==':
         val = str(edge_condition['value'])
    else:
        val = generate_upi()
    return _apply_pii_strategy_to_value(val, "upi", pii_strategy)

def _generate_animal_name_value(constraint, field_name, pii_strategy, edge_condition=None):
    """Generates a pet-like name."""
    if edge_condition and edge_condition.get('operator') == '==':
        return str(edge_condition['value'])
    # Using fake.first_name() as a proxy for pet names.
    # Could be expanded with a dedicated list of actual animal names.
    return fake.first_name()
    # Pet names are generally not PII, so no _apply_pii_strategy_to_value needed here.

# --- Constants for Domain-Specific Data Generation ---
DIAGNOSES_LIST = ["Hypertension", "Diabetes", "Asthma", "Arthritis", "Migraine",
                  "Bronchitis", "Anemia", "Pneumonia", "Sinusitis", "Depression", "Anxiety",
                  "Allergy", "Osteoporosis", "Glaucoma", "Eczema", "GERD", "COVID-19", "Influenza",
                  "Chronic Kidney Disease", "Hyperlipidemia", "Hypothyroidism", "Sleep Apnea", "Obesity"]
INSURANCE_PROVIDERS_LIST = ["United Health", "Aetna", "Blue Cross", "Cigna", "Humana",
                            "Kaiser Permanente", "Anthem", "Centene", "Molina Healthcare", "WellCare"]
BLOOD_TYPES_LIST = ["A+", "A-", "B+", "B-", "AB+", "AB-", "O+", "O-"]
MEDICATIONS_LIST = ["Paracetamol", "Ibuprofen", "Amoxicillin", "Lisinopril", "Metformin",
                    "Atorvastatin", "Omeprazole", "Losartan", "Albuterol", "Gabapentin",
                    "Sertraline", "Amlodipine", "Simvastatin", "Hydrochlorothiazide", "Levothyroxine",
                    "Azithromycin", "Ciprofloxacin", "Prednisone", "Insulin Glargine", "Warfarin", "Clopidogrel"]
TRANSACTION_TYPES_LIST = ["Deposit", "Withdrawal", "Transfer", "Payment", "Refund", "Fee", "Interest"]
PRODUCT_NAMES_LIST = ["Laptop", "Smartphone", "Headphones", "Smartwatch", "Tablet",
                      "Camera", "Speaker", "Monitor", "Keyboard", "Mouse",
                      "Desk Chair", "Coffee Maker", "Blender", "Backpack", "Water Bottle",
                      "Running Shoes", "Yoga Mat", "Bluetooth Speaker", "External Hard Drive", "Webcam"]
PRODUCT_NAMES_LIST.extend(["Drone", "VR Headset", "Air Fryer", "Electric Scooter", "Fitness Tracker",
                           "Projector", "Robot Vacuum", "Security Camera System", "Portable SSD", "Gaming Console",
                           "Wireless Charger", "E-reader", "Digital Photo Frame", "Smart Thermostat", "Air Purifier",
                           "Electric Toothbrush", "Hair Dryer", "Instant Pot", "Microwave Oven", "Toaster", "Action Camera"])
PAYMENT_METHODS_LIST = ["Credit Card", "Debit Card", "UPI", "Net Banking", "Cash on Delivery", "Wallet", "EMI"]
PRODUCT_CATEGORIES_LIST = ["Electronics", "Clothing", "Home & Kitchen", "Books", "Beauty"]
SHIPPING_STATUSES_LIST = ["Pending", "Shipped", "Delivered", "Cancelled", "Returned"]
GRADES_LIST = ["A", "B", "C", "D", "F"]
SUBJECTS_LIST = ["Math", "Science", "English", "History", "Geography"]
DEPARTMENTS_LIST = ["IT", "HR", "Finance", "Marketing", "Operations"]
DEPARTMENTS_LIST.extend(["Legal", "Customer Support", "Research & Development", "Product Management", "Quality Assurance",
                         "Sales", "Supply Chain", "Administration", "Public Relations", "Design"])
POSITIONS_LIST = ["Manager", "Developer", "Analyst", "Designer", "Accountant",
                  "Consultant", "Specialist", "Coordinator", "Engineer", "Administrator",
                  "Director", "Vice President", "Executive", "Intern", "Team Lead", "Architect", "Scientist"]
DEGREES_LIST = ["B.Tech", "MBA", "B.Com", "B.A", "M.Sc"]
SKILLS_LIST = ["Python", "Java", "SQL", "Excel", "Communication"]
SKILLS_LIST.extend(["Project Management", "Data Analysis", "Machine Learning", "Cloud Computing", "Cybersecurity", "Agile Methodologies", "UI/UX Design",
                    "JavaScript", "React", "Node.js", "Angular", "Vue.js", "DevOps", "Kubernetes", "Docker", "Terraform",
                    "C++", "C#", ".NET", "PHP", "Ruby on Rails", "Go", "Swift", "Kotlin", "Mobile Development (iOS/Android)",
                    "Technical Writing", "Problem Solving", "Critical Thinking", "Leadership", "Teamwork", "Creativity",
                    "Digital Marketing", "SEO/SEM", "Content Creation", "Salesforce", "SAP", "Oracle", "Tableau", "Power BI"])
PROPERTY_TYPES_LIST = ["Apartment", "Villa", "Plot", "Office", "Shop"]
REALESTATE_STATUSES_LIST = ["Available", "Sold", "Rented", "Under Construction"]
AMENITIES_LIST = ["Swimming Pool", "Gym", "Park", "Security", "Play Area"]
FOOD_ITEMS_LIST = ["Pizza", "Burger", "Pasta", "Salad", "Fries", "Soda", "Sushi", "Tacos", "Biryani", "Noodles", "Sandwich", "Ice Cream",
                   "Salad Bowl", "Smoothie", "Coffee", "Tea", "Milkshake", "Wrap", "Steak", "Seafood Platter", "Dim Sum", "Ramen", "Curry",
                   "Dosa", "Idli", "Vada", "Samosa", "Spring Roll", "Momo", "Fried Chicken", "Hot Dog", "Pancakes", "Waffles",
                   "Omelette", "Cereal", "Yogurt", "Fruit Salad", "Juice", "Cake", "Pastry", "Cookie", "Donut", "Muffin"]
RESTAURANT_TYPES_LIST = ["Cafe", "Fine Dining", "Fast Food", "Pizzeria", "Bakery", "Cloud Kitchen"] # For restaurant name generation
DELIVERY_STATUSES_LIST = ["Order Placed", "Preparing", "Out for Delivery", "Delivered", "Cancelled by User", "Cancelled by Restaurant", "Delayed"]
ACADEMIC_JOURNALS_LIST = ["Nature", "Science", "Cell", "The Lancet", "JAMA", "IEEE Transactions", "Physical Review Letters",
                          "New England Journal of Medicine (NEJM)", "PNAS", "BMJ", "Nature Communications", "Journal of the American Chemical Society (JACS)",
                          "Angewandte Chemie", "Advanced Materials", "Nature Medicine", "Science Advances", "Cell Host & Microbe", "Immunity"]
COMMON_HASHTAGS_LIST = ["#instagood", "#photooftheday", "#love", "#travel", "#tech", "#science", "#innovation", "#news", "#health",
                        "#business", "#startup", "#motivation", "#art", "#foodie", "#fitness"]
SENSOR_TYPES_LIST = ["Temperature", "Humidity", "Pressure", "Light", "Motion", "GPS",
                     "Accelerometer", "Gyroscope", "Proximity", "Sound Level", "Air Quality", "CO2 Sensor",
                     "Water Flow", "Soil Moisture", "Radiation", "Magnetic Field", "Vibration", "Ultrasonic"]
SPECIES_LIST = ["Dog", "Cat", "Bird", "Fish", "Lion", "Tiger", "Elephant", "Bear", "Rabbit", "Horse", "Cow", "Sheep",
                "Deer", "Fox", "Wolf", "Monkey", "Snake", "Lizard", "Frog", "Turtle", "Shark", "Whale", "Dolphin",
                "Eagle", "Owl", "Penguin", "Crocodile", "Alligator", "Spider", "Ant", "Bee", "Butterfly"]
COMMON_DOG_BREEDS_LIST = ["Labrador Retriever", "German Shepherd", "Golden Retriever", "Bulldog", "Poodle", "Beagle"]
COMMON_CAT_BREEDS_LIST = ["Siamese", "Persian", "Maine Coon", "Ragdoll", "Bengal", "Sphynx"]
LOGISTICS_CARRIER_SUFFIXES_LIST = ["Logistics", "Shipping Lines", "Freight", "Express", "Carriers", "Transport", "Movers"]
LOGISTICS_SHIPMENT_STATUSES_LIST = ["Processing", "In Transit", "Out for Delivery", "Delivered",
                                    "Delayed", "Held at Customs", "Returned to Sender", "Exception",
                                    "Pending Pickup", "At Origin Facility", "At Destination Facility"]
HABITAT_LIST = ["Forest", "Desert", "Ocean", "Grassland", "Mountain", "Domestic", "Urban", "Arctic",
                "Jungle", "Savanna", "Wetland", "River", "Lake", "Cave", "Coral Reef", "Tundra", "Taiga", "Swamp"]
PUBLICATION_TYPES_LIST = ["Journal Article", "Conference Paper", "Book Chapter", "Preprint", "Thesis",
                          "Review Article", "Case Study", "Technical Report", "Poster Presentation", "Editorial",
                          "Letter to Editor", "Book Review", "Dataset Paper", "Software Paper", "Patent"]

# --- NEW LISTS FOR EXPANDED DOMAINS ---
LEAD_SOURCES_LIST = ["Website", "Referral", "Cold Call", "Advertisement", "Trade Show", "Social Media", "Email Campaign"]
DEAL_STAGES_LIST = ["Prospecting", "Qualification", "Needs Analysis", "Proposal", "Negotiation", "Closed Won", "Closed Lost"]
CAMPAIGN_TYPES_LIST = ["Email Marketing", "Social Media Ads", "Search Engine Marketing (SEM)", "Content Marketing", "Influencer Marketing", "Affiliate Marketing"]
WAREHOUSE_LOCATIONS_LIST = ["North Wing", "South Wing", "East Dock", "West Dock", "Central Hub", "Overflow Section"]
EMPLOYEE_STATUS_LIST = ["Active", "On Leave", "Terminated", "Resigned", "Retired", "Contractor"]
PROJECT_STATUS_LIST = ["Not Started", "In Progress", "Completed", "On Hold", "Cancelled"]
TASK_PRIORITY_LIST = ["High", "Medium", "Low", "Critical"]
VEHICLE_TYPES_LIST = ["Sedan", "SUV", "Truck", "Hatchback", "Coupe", "Minivan", "Motorcycle"]
RETURN_REASONS_LIST = ["Wrong Item", "Damaged Item", "Changed Mind", "Doesn't Fit", "Not as Described"]
POLICY_TYPES_LIST = ["Life Insurance", "Health Insurance", "Auto Insurance", "Home Insurance", "Travel Insurance"]
CLAIM_STATUSES_LIST = ["Submitted", "Processing", "Approved", "Rejected", "Paid", "Pending Information"]
TRANSPORT_MODES_LIST = ["Road", "Rail", "Air", "Sea", "Pipeline"]
COURSE_LEVELS_LIST = ["Beginner", "Intermediate", "Advanced", "Expert"]
PROPERTY_LISTING_STATUSES_LIST = ["For Sale", "For Rent", "Sold", "Rented", "Pending", "Off Market"]
ISSUE_TYPES_LIST = ["Bug", "Feature Request", "Task", "Improvement", "Question"]
ISSUE_STATUSES_LIST = ["Open", "In Progress", "Resolved", "Closed", "Reopened", "Pending QA"]
LEGAL_CASE_TYPES_LIST = ["Civil", "Criminal", "Family Law", "Corporate Law", "Real Estate Law"]
MEDIA_GENRES_LIST = ["Action", "Comedy", "Drama", "Sci-Fi", "Horror", "Romance", "Thriller", "Documentary", "Animation"]
ENERGY_TARIFF_TYPES_LIST = ["Fixed Rate", "Variable Rate", "Time-of-Use", "Prepaid"]
AGRICULTURE_CROP_TYPES_LIST = ["Wheat", "Rice", "Corn", "Soybean", "Cotton", "Sugarcane", "Fruits", "Vegetables"]
URL_SCHEMES_LIST = ["http", "https"] # For URL generation
FILE_EXTENSIONS_LIST = ["txt", "pdf", "jpg", "png", "docx", "xlsx", "csv", "py", "js", "html"]
BOOLEAN_REPRESENTATIONS_LIST = ["True,False", "Yes,No", "1,0", "Active,Inactive"] # For boolean-like categories


# Travel & Tourism Domain Lists
TRAVEL_DESTINATIONS_LIST = ["Paris", "Rome", "London", "New York", "Tokyo", "Dubai", "Bali", "Barcelona", "Amsterdam", "Sydney", "Bangkok", "Singapore", "Venice", "Prague", "Vienna", "Berlin", "San Francisco", "Los Angeles", "Miami", "Chicago", "Toronto", "Vancouver", "Mexico City", "Rio de Janeiro", "Buenos Aires", "Cairo", "Cape Town", "Mumbai", "Delhi", "Beijing", "Shanghai", "Seoul", "Moscow", "Istanbul"]
AIRLINE_NAMES_LIST = ["Emirates", "Qatar Airways", "Singapore Airlines", "ANA All Nippon Airways", "Qantas Airways", "Lufthansa", "British Airways", "Delta Air Lines", "American Airlines", "United Airlines", "Air France", "KLM", "Turkish Airlines", "Cathay Pacific", "IndiGo", "Southwest Airlines"]
HOTEL_AMENITIES_LIST = ["WiFi", "Swimming Pool", "Gym", "Spa", "Restaurant", "Bar", "Room Service", "Parking", "Air Conditioning", "Breakfast Included", "Airport Shuttle", "Pet Friendly", "Business Center", "Concierge"]
ROOM_TYPES_LIST = ["Standard Room", "Deluxe Room", "Suite", "Family Room", "Single Room", "Double Room", "King Room", "Queen Room", "Studio", "Apartment", "Villa", "Bungalow"]
TRAVEL_BOOKING_STATUSES_LIST = ["Confirmed", "Pending", "Cancelled", "Modified", "Completed", "No-Show", "Waitlisted"]
TRAVEL_ACTIVITY_TYPES_LIST = ["Sightseeing Tour", "Museum Visit", "Adventure Sport", "Cooking Class", "Wine Tasting", "Cultural Show", "Shopping Trip", "Beach Relaxation", "Hiking", "Cruise"]



GENERAL_CATEGORIES_LIST = ["Category A", "Category B", "Category C", "Category D"]
GENERAL_STATUSES_LIST = ["Active", "Inactive", "Pending", "Completed"]

# --- Field Generator Dispatch Dictionary ---
# This will be replaced by CANONICAL_FIELD_TO_SCHEMA_DETAILS_MAP and generate_value
# FIELD_GENERATORS = { ... } # Removed for brevity, will be deprecated

# --- NEW: Map of Canonical Field Names to their Schema Details (Type, Constraint) ---
# This map will be used by the refactored generate_synthetic_data function
CANONICAL_FIELD_TO_SCHEMA_DETAILS_MAP = {
    # General & Common
    "id": {"type": "string", "constraint": "", "is_generic_numeric_id": True, "display_name": "ID"},
    "name": {"type": "name", "constraint": "", "display_name": "Name"},
    "age": {"type": "int", "constraint": "18-80", "display_name": "Age"},
    "gender": {"type": "category", "constraint": "Male,Female,Other", "display_name": "Gender"},
    "email": {"type": "email", "constraint": "", "display_name": "Email"},
    "phone": {"type": "phone", "constraint": "", "display_name": "Phone"},
    "address": {"type": "address", "constraint": "", "display_name": "Address"},
    "date": {"type": "date", "constraint": "", "display_name": "Date"}, # Generic date
    "value": {"type": "float", "constraint": "10-1000", "display_name": "Value"},
    "category": {"type": "category", "constraint": ",".join(GENERAL_CATEGORIES_LIST), "display_name": "Category"},
    "description": {"type": "string", "constraint": "", "display_name": "Description"}, # Uses fake.sentence()
    "status": {"type": "category", "constraint": ",".join(GENERAL_STATUSES_LIST), "display_name": "Status"},
    "score": {"type": "float", "constraint": "0-100", "display_name": "Score"},
    "city": {"type": "string", "constraint": "", "is_faker_city": True, "display_name": "City"},
    "state": {"type": "string", "constraint": "", "is_faker_state": True, "display_name": "State"},
    "country": {"type": "string", "constraint": "", "is_faker_country": True, "display_name": "Country"},
    "pincode": {"type": "string", "constraint": "", "is_faker_postcode": True, "display_name": "Pincode"}, # also zipcode
    "zipcode": {"type": "string", "constraint": "", "is_faker_postcode": True, "display_name": "Zipcode"},
    "currency": {"type": "string", "constraint": "", "is_faker_currency_code": True, "display_name": "Currency"},
    "job": {"type": "string", "constraint": "", "is_faker_job": True, "display_name": "Job Title"},
    "company": {"type": "string", "constraint": "", "is_faker_company": True, "display_name": "Company"},
    "aadhaar": {"type": "aadhaar", "constraint": "", "display_name": "Aadhaar"},
    "pan": {"type": "pan", "constraint": "", "display_name": "PAN"},
    "passport": {"type": "passport", "constraint": "", "display_name": "Passport"},
    "url": {"type": "string", "constraint": "", "is_faker_url": True, "display_name": "URL"},
    "ip_address": {"type": "string", "constraint": "", "is_faker_ipv4": True, "display_name": "IP Address"},
    "mac_address": {"type": "string", "constraint": "", "is_faker_mac_address": True, "display_name": "MAC Address"},
    "version_number": {"type": "string", "constraint": "", "is_version_number_pattern": True, "display_name": "Version"},
    "blood_pressure_reading": {"type": "string", "constraint": "", "is_blood_pressure_pattern": True, "display_name": "Blood Pressure"},
    "boolean_flag": {"type": "category", "constraint": random.choice(BOOLEAN_REPRESENTATIONS_LIST), "display_name": "Flag"},
    "timestamp_detailed": {"type": "date", "constraint": "datetime_utc", "display_name": "Timestamp (UTC)"}, # Special constraint for datetime with UTC
    "file_name": {"type": "string", "constraint": "", "is_faker_file_name": True, "display_name": "File Name"},
    "mime_type": {"type": "string", "constraint": "", "is_faker_mime_type": True, "display_name": "MIME Type"},
    "voterid": {"type": "voterid", "constraint": "", "display_name": "Voter ID"},
    "ifsc": {"type": "ifsc", "constraint": "", "display_name": "IFSC Code"}, # also ifsc_code
    "upi": {"type": "upi", "constraint": "", "display_name": "UPI ID"}, # also upi_id

    # Hospital Domain
    "patient_id": {"type": "string", "constraint": "", "is_generic_numeric_id": True, "display_name": "Patient ID"},
    "doctor_name": {"type": "name", "constraint": "", "prefix": "Dr.", "display_name": "Doctor Name"}, # Note: space after Dr. handled in generation
    "hospital_name": {"type": "string", "constraint": "", "suffix": " Hospital", "is_faker_company": True, "display_name": "Hospital Name"},
    "diagnosis": {"type": "category", "constraint": ",".join(DIAGNOSES_LIST), "display_name": "Diagnosis"},
    "admission_date": {"type": "date", "constraint": "", "display_name": "Admission Date"}, # Needs relative date logic if "discharge_date" is also present
    "discharge_date": {"type": "date", "constraint": "", "display_name": "Discharge Date"},
    "room_number": {"type": "string", "constraint": "", "is_room_number_pattern": True, "display_name": "Room Number"}, # e.g., "101A"
    "insurance_provider": {"type": "category", "constraint": ",".join(INSURANCE_PROVIDERS_LIST), "display_name": "Insurance Provider"},
    "blood_type": {"type": "category", "constraint": ",".join(BLOOD_TYPES_LIST), "display_name": "Blood Type"},
    "medication": {"type": "category", "constraint": ",".join(MEDICATIONS_LIST), "display_name": "Medication"},

    # Finance Domain (Example, expand as needed)
    "transaction_id": {"type": "string", "constraint": "", "is_generic_numeric_id": True, "display_name": "Transaction ID"},
    "account_number": {"type": "string", "constraint": "digits:10-12", "is_digit_sequence": True, "display_name": "Account Number"}, # Custom constraint type
    "bank_name": {"type": "string", "constraint": "", "is_faker_company": True, "suffix": " Bank", "display_name": "Bank Name"},
    "balance": {"type": "float", "constraint": "-5000-50000", "display_name": "Balance"},
    "reference_number": {"type": "string", "constraint": "", "is_reference_number_pattern": True, "display_name": "Reference Number"},
    "amount": {"type": "float", "constraint": "100-10000", "display_name": "Amount"},
    "salary": {"type": "int", "constraint": "20000-200000", "display_name": "Salary"},
    "transaction_type": {"type": "category", "constraint": ",".join(TRANSACTION_TYPES_LIST), "display_name": "Transaction Type"},
    "credit_card_number": {"type": "string", "constraint": "", "is_faker_credit_card_number": True, "display_name": "Credit Card Number"},

    # Food Delivery Domain
    "restaurant_name": {"type": "string", "constraint": "", "is_faker_company": True, "suffix_from_list": RESTAURANT_TYPES_LIST, "display_name": "Restaurant Name"},
    "food_items": {"type": "category", "constraint": ",".join(FOOD_ITEMS_LIST), "display_name": "Food Items"}, # Could also be a multi-select or string for comma-separated items
    "order_total": {"type": "int", "constraint": "100-2000", "display_name": "Order Total (INR)"},
    "delivery_agent_name": {"type": "name", "constraint": "", "display_name": "Delivery Agent Name"},
    "delivery_time_minutes": {"type": "int", "constraint": "15-75", "display_name": "Delivery Time (Minutes)"},
    "delivery_rating": {"type": "int", "constraint": "1-5", "display_name": "Delivery Rating"},
    "delivery_address": {"type": "address", "constraint": "", "display_name": "Delivery Address"}, # Already covered by general address
    "payment_mode": {"type": "category", "constraint": ",".join(PAYMENT_METHODS_LIST), "display_name": "Payment Mode"}, # Re-use from e-commerce
    "delivery_status": {"type": "category", "constraint": ",".join(DELIVERY_STATUSES_LIST), "display_name": "Delivery Status"},

    # Education Domain
    "student_id": {"type": "string", "constraint": "", "is_generic_numeric_id": True, "display_name": "Student ID"},
    "school_name": {"type": "string", "constraint": "", "is_faker_company": True, "suffix": " School", "display_name": "School Name"},
    "teacher_name": {"type": "name", "constraint": "", "prefix_options": ["Mr.", "Ms.", "Dr."], "display_name": "Teacher Name"},
    "grade": {"type": "category", "constraint": ",".join(GRADES_LIST), "display_name": "Grade"},
    "subject": {"type": "category", "constraint": ",".join(SUBJECTS_LIST), "display_name": "Subject"},
    "attendance": {"type": "string", "constraint": "70-100", "is_percentage_pattern": True, "display_name": "Attendance"},
    "course_name": {"type": "string", "constraint": "", "is_faker_bs": True, "display_name": "Course Name"},

    # Employee Domain
    "employee_id": {"type": "string", "constraint": "", "is_generic_numeric_id": True, "display_name": "Employee ID"},
    "department": {"type": "category", "constraint": ",".join(DEPARTMENTS_LIST), "display_name": "Department"},
    "position": {"type": "category", "constraint": ",".join(POSITIONS_LIST), "display_name": "Position"},
    "employee_status": {"type": "category", "constraint": ",".join(EMPLOYEE_STATUS_LIST), "display_name": "Employee Status"},
    "performance_rating": {"type": "float", "constraint": "1.0-5.0", "display_name": "Performance Rating"},

    # Real Estate Domain
    "property_id": {"type": "string", "constraint": "", "is_generic_numeric_id": True, "display_name": "Property ID"},
    "area": {"type": "string", "constraint": "500-3000", "unit": "sq.ft.", "is_measurement_pattern": True, "display_name": "Area"},
    "property_type": {"type": "category", "constraint": ",".join(PROPERTY_TYPES_LIST), "display_name": "Property Type"},
    "listing_status": {"type": "category", "constraint": ",".join(PROPERTY_LISTING_STATUSES_LIST), "display_name": "Listing Status"},
    "year_built": {"type": "int", "constraint": "1990-2023", "display_name": "Year Built"},

    # E-commerce (some fields might be general or already covered)
    "order_id": {"type": "string", "constraint": "", "is_generic_numeric_id": True, "display_name": "Order ID"},
    "customer_name": {"type": "name", "constraint": "", "display_name": "Customer Name"},
    "product_name": {"type": "category", "constraint": ",".join(PRODUCT_NAMES_LIST), "display_name": "Product Name"},
    "price": {"type": "float", "constraint": "100-5000", "display_name": "Price"},
    "quantity": {"type": "int", "constraint": "1-10", "display_name": "Quantity"},
    "product_category": {"type": "category", "constraint": ",".join(PRODUCT_CATEGORIES_LIST), "display_name": "Product Category"},
    "discount": {"type": "float", "constraint": "0-0.5", "display_name": "Discount"}, # as a percentage
    "shipping_status": {"type": "category", "constraint": ",".join(SHIPPING_STATUSES_LIST), "display_name": "Shipping Status"},
    # ... Add ALL other canonical fields from FIELD_GENERATORS here, mapping them to a type and constraint

    # Academic/Research Domain
    "publication_title": {"type": "string", "constraint": "", "display_name": "Publication Title"}, # Will use _generate_string_value
    "author_names": {"type": "name", "constraint": "", "is_multi_name": True, "display_name": "Author Names"}, # Special handling for multiple authors
    "journal_name": {"type": "category", "constraint": ",".join(ACADEMIC_JOURNALS_LIST), "display_name": "Journal Name"},
    "publication_year": {"type": "int", "constraint": "2000-2024", "display_name": "Publication Year"},
    "citation_count": {"type": "int", "constraint": "0-1000", "display_name": "Citation Count"},
    "doi": {"type": "string", "constraint": "", "is_doi_pattern": True, "display_name": "DOI"},
    "keywords": {"type": "string", "constraint": "", "is_keywords_list": True, "display_name": "Keywords"}, # Comma-separated list of words
    "publication_type": {"type": "category", "constraint": ",".join(PUBLICATION_TYPES_LIST), "display_name": "Publication Type"},

    # Social Media Domain
    "username": {"type": "string", "constraint": "", "is_faker_user_name": True, "display_name": "Username"},
    "post_text": {"type": "string", "constraint": "", "display_name": "Post Text"}, # Will use _generate_string_value
    "like_count": {"type": "int", "constraint": "0-10000", "display_name": "Like Count"},
    "share_count": {"type": "int", "constraint": "0-5000", "display_name": "Share Count"},
    "comment_text": {"type": "string", "constraint": "", "display_name": "Comment Text"}, # Will use _generate_string_value
    "hashtags": {"type": "category", "constraint": ",".join(COMMON_HASHTAGS_LIST), "is_multi_category": True, "display_name": "Hashtags"},

    # IoT/Sensor Domain
    "sensor_id": {"type": "string", "constraint": "", "is_generic_alphanum_id": True, "prefix": "SENSOR-", "display_name": "Sensor ID"},
    "timestamp": {"type": "date", "constraint": "datetime", "display_name": "Timestamp"}, # Special constraint for datetime
    "temperature": {"type": "float", "constraint": "-20.0-50.0", "unit": "°C", "is_measurement_pattern": True, "display_name": "Temperature"},
    "humidity": {"type": "float", "constraint": "0.0-100.0", "unit": "%", "is_measurement_pattern": True, "display_name": "Humidity"},
    "latitude": {"type": "float", "constraint": "-90.0-90.0", "is_faker_latitude": True, "display_name": "Latitude"},
    "longitude": {"type": "float", "constraint": "-180.0-180.0", "is_faker_longitude": True, "display_name": "Longitude"},
    "sensor_type": {"type": "category", "constraint": ",".join(SENSOR_TYPES_LIST), "display_name": "Sensor Type"},

    # Animal Data Domain
    "animal_name": {"type": "animal_name", "constraint": "", "display_name": "Animal Name"}, # Changed type to "animal_name"
    "species": {"type": "category", "constraint": ",".join(SPECIES_LIST), "display_name": "Species"},
    "breed": {"type": "category", "constraint": ",".join(COMMON_DOG_BREEDS_LIST + COMMON_CAT_BREEDS_LIST), "display_name": "Breed"}, # Combine or make dynamic
    "animal_age": {"type": "int", "constraint": "0-25", "display_name": "Animal Age (Years)"},
    "habitat": {"type": "category", "constraint": ",".join(HABITAT_LIST), "display_name": "Habitat"},
    "animal_weight_kg": {"type": "float", "constraint": "0.1-1000", "display_name": "Weight (kg)"},
    "animal_color": {"type": "string", "constraint": "", "is_faker_color_name": True, "display_name": "Color"},

    # Logistics/Supply Chain Domain
    "shipment_id": {"type": "string", "constraint": "", "is_generic_alphanum_id": True, "prefix": "SHP-", "display_name": "Shipment ID"},
    "tracking_number": {"type": "string", "constraint": "", "is_tracking_number_pattern": True, "display_name": "Tracking Number"},
    "carrier_name": {"type": "string", "constraint": "", "is_faker_company": True, "suffix_from_list": LOGISTICS_CARRIER_SUFFIXES_LIST, "display_name": "Carrier Name"},
    "origin_location": {"type": "address", "constraint": "", "display_name": "Origin Location"}, # Reuses address type
    "destination_location": {"type": "address", "constraint": "", "display_name": "Destination Location"}, # Reuses address type
    "shipment_status_logistics": {"type": "category", "constraint": ",".join(LOGISTICS_SHIPMENT_STATUSES_LIST), "display_name": "Shipment Status"},
    "freight_cost": {"type": "float", "constraint": "50-5000", "display_name": "Freight Cost"},
    "estimated_delivery_date": {"type": "date", "constraint": "", "display_name": "Estimated Delivery Date"},
    "actual_delivery_date": {"type": "date", "constraint": "", "display_name": "Actual Delivery Date"},
    "package_weight_kg": {"type": "float", "constraint": "0.1-1000", "unit": "kg", "is_measurement_pattern": True, "display_name": "Package Weight (kg)"}, # Reused from animal, good generic
    "package_dimensions_cm": {"type": "string", "constraint": "", "is_dimension_pattern": True, "display_name": "Package Dimensions (cm)"},

    # Travel & Tourism Domain
    "booking_id": {"type": "string", "constraint": "", "is_generic_alphanum_id": True, "prefix": "BKG-", "display_name": "Booking ID"},
    "traveler_name": {"type": "name", "constraint": "", "display_name": "Traveler Name"},
    "destination_city_travel": {"type": "category", "constraint": ",".join(TRAVEL_DESTINATIONS_LIST), "is_faker_city_if_empty_constraint": True, "display_name": "Destination City"},
    "origin_city_travel": {"type": "category", "constraint": ",".join(TRAVEL_DESTINATIONS_LIST), "is_faker_city_if_empty_constraint": True, "display_name": "Origin City"},
    "travel_date": {"type": "date", "constraint": "", "display_name": "Travel Date"}, # Departure date
    "return_date": {"type": "date", "constraint": "", "display_name": "Return Date"}, # Needs to be after travel_date
    "flight_number": {"type": "string", "constraint": "", "is_flight_number_pattern": True, "display_name": "Flight Number"},
    "airline_name": {"type": "category", "constraint": ",".join(AIRLINE_NAMES_LIST), "is_faker_company_if_empty_constraint": True, "suffix": " Airlines", "display_name": "Airline Name"},
    "hotel_name": {"type": "string", "constraint": "", "is_faker_company": True, "suffix_options": [" Hotel", " Resort", " Inn", " Lodge", " Suites"], "display_name": "Hotel Name"},
    "room_type": {"type": "category", "constraint": ",".join(ROOM_TYPES_LIST), "display_name": "Room Type"},
    "booking_status_travel": {"type": "category", "constraint": ",".join(TRAVEL_BOOKING_STATUSES_LIST), "display_name": "Booking Status"},
    "total_travel_cost": {"type": "float", "constraint": "200-10000", "display_name": "Total Cost (USD)"}, # Assuming USD for now
    "travel_package_name": {"type": "string", "constraint": "", "is_faker_bs": True, "prefix": "Package: ", "display_name": "Travel Package Name"}, # Using bs for catchy names
    "hotel_amenities_included": {"type": "category", "constraint": ",".join(HOTEL_AMENITIES_LIST), "is_multi_category": True, "display_name": "Hotel Amenities"},
    "travel_activity": {"type": "category", "constraint": ",".join(TRAVEL_ACTIVITY_TYPES_LIST), "display_name": "Activity Booked"},

    # --- NEW DOMAINS ---
    # Sales/CRM
    "lead_id": {"type": "string", "constraint": "", "is_generic_alphanum_id": True, "prefix": "LEAD-", "display_name": "Lead ID"},
    "lead_source": {"type": "category", "constraint": ",".join(LEAD_SOURCES_LIST), "display_name": "Lead Source"},
    "deal_stage": {"type": "category", "constraint": ",".join(DEAL_STAGES_LIST), "display_name": "Deal Stage"},
    "opportunity_amount": {"type": "float", "constraint": "1000-1000000", "display_name": "Opportunity Amount"},
    "close_date": {"type": "date", "constraint": "", "display_name": "Close Date"},
    "account_type": {"type": "category", "constraint": "Prospect,Customer,Partner,Vendor", "display_name": "Account Type"},
    "industry": {"type": "string", "constraint": "", "is_faker_bs": True, "display_name": "Industry"}, # Using bs for variety

    # Marketing
    "campaign_id": {"type": "string", "constraint": "", "is_generic_alphanum_id": True, "prefix": "CAMP-", "display_name": "Campaign ID"},
    "campaign_name": {"type": "string", "constraint": "", "is_faker_catch_phrase": True, "display_name": "Campaign Name"},
    "campaign_type": {"type": "category", "constraint": ",".join(CAMPAIGN_TYPES_LIST), "display_name": "Campaign Type"},
    "ad_spend": {"type": "float", "constraint": "100-10000", "display_name": "Ad Spend"},
    "impressions": {"type": "int", "constraint": "1000-1000000", "display_name": "Impressions"},
    "clicks": {"type": "int", "constraint": "10-10000", "display_name": "Clicks"},
    "ctr": {"type": "float", "constraint": "0.001-0.1", "display_name": "Click-Through Rate (CTR)"}, # As a decimal
    "conversion_rate": {"type": "float", "constraint": "0.01-0.2", "display_name": "Conversion Rate"}, # As a decimal

    # Manufacturing/Inventory
    "sku": {"type": "string", "constraint": "", "is_generic_alphanum_id": True, "prefix": "SKU-", "display_name": "SKU"},
    "serial_number": {"type": "string", "constraint": "", "is_generic_alphanum_id": True, "length": 12, "display_name": "Serial Number"},
    "batch_number": {"type": "string", "constraint": "", "is_generic_alphanum_id": True, "prefix": "BATCH-", "length": 8, "display_name": "Batch Number"},
    "warehouse_location": {"type": "category", "constraint": ",".join(WAREHOUSE_LOCATIONS_LIST), "display_name": "Warehouse Location"},
    "stock_level": {"type": "int", "constraint": "0-10000", "display_name": "Stock Level"},
    "supplier_name": {"type": "string", "constraint": "", "is_faker_company": True, "display_name": "Supplier Name"},

    # Project Management
    "project_id": {"type": "string", "constraint": "", "is_generic_alphanum_id": True, "prefix": "PROJ-", "display_name": "Project ID"},
    "project_name": {"type": "string", "constraint": "", "is_faker_bs": True, "display_name": "Project Name"},
    "task_id": {"type": "string", "constraint": "", "is_generic_alphanum_id": True, "prefix": "TASK-", "display_name": "Task ID"},
    "task_name": {"type": "string", "constraint": "", "is_faker_catch_phrase": True, "display_name": "Task Name"},
    "assignee_name": {"type": "name", "constraint": "", "display_name": "Assignee"},
    "due_date": {"type": "date", "constraint": "", "display_name": "Due Date"},
    "project_status": {"type": "category", "constraint": ",".join(PROJECT_STATUS_LIST), "display_name": "Project Status"},
    "task_priority": {"type": "category", "constraint": ",".join(TASK_PRIORITY_LIST), "display_name": "Task Priority"},

    # Software Development
    "bug_id": {"type": "string", "constraint": "", "is_generic_numeric_id": True, "prefix": "BUG-", "display_name": "Bug ID"},
    "feature_id": {"type": "string", "constraint": "", "is_generic_numeric_id": True, "prefix": "FEAT-", "display_name": "Feature ID"},
    "commit_hash": {"type": "string", "constraint": "", "is_faker_sha256": True, "display_name": "Commit Hash"}, # sha256 is a good proxy
    "repository_url": {"type": "string", "constraint": "", "is_faker_url": True, "display_name": "Repository URL"},
    "issue_type": {"type": "category", "constraint": ",".join(ISSUE_TYPES_LIST), "display_name": "Issue Type"},
    "issue_status": {"type": "category", "constraint": ",".join(ISSUE_STATUSES_LIST), "display_name": "Issue Status"},


    # This is a crucial step for the refactor to work comprehensively.
}

# --- NEW: Field Schema Inference from Name ---
def infer_field_schema_from_name(field_name_str, default_pii_strategy="realistic_fake"):
    """
    Infers a basic field schema (type, constraint, etc.) from a field name string.
    """
    # Default display name from the input string
    display_name = field_name_str.replace("_", " ").title()
    field_schema = {
        "name": display_name,
        "type": "string",  # Default type
        "constraint": "",
        "pii_handling": default_pii_strategy,
        # Store a normalized version for internal use or as a fallback canonical key
        "_original_canonical": field_name_str.lower().replace(" ", "_")
    }
    field_lower = field_name_str.lower()

    # 1. Try to match to canonical schema first
    normalized_field_name = field_lower.replace(" ", "_")
    canonical_match = None

    if normalized_field_name in CANONICAL_FIELD_TO_SCHEMA_DETAILS_MAP:
        canonical_match = normalized_field_name
    else:
        # Sort synonyms by length (descending) to match longer, more specific synonyms first
        sorted_synonyms = sorted(FIELD_SYNONYM_TO_CANONICAL_MAP.keys(), key=len, reverse=True)
        for user_synonym in sorted_synonyms:
            # Check for exact match of the synonym
            if user_synonym == normalized_field_name: # More precise than 'in'
                potential_canonical = FIELD_SYNONYM_TO_CANONICAL_MAP[user_synonym]
                if potential_canonical in CANONICAL_FIELD_TO_SCHEMA_DETAILS_MAP:
                    canonical_match = potential_canonical
                    break
            # Fallback to 'in' if no exact match, but be cautious as it can be too broad
            # elif user_synonym in normalized_field_name:
            #     # This part might need more refinement to avoid overly greedy matches
            #     # For now, prioritize exact matches above
            #     pass


    if canonical_match:
        details = CANONICAL_FIELD_TO_SCHEMA_DETAILS_MAP[canonical_match].copy()
        field_schema.update(details)
        field_schema["name"] = display_name # Ensure original name (title cased) is used for display
        field_schema["_original_canonical"] = canonical_match # Store the matched canonical key
        return field_schema

    # 2. If no canonical match, apply general inference rules
    if any(kw in field_lower for kw in ["date", "time", "timestamp", "dob", "joining", "day", "month", "year_of_birth", "birth_date"]): # "year" is too generic alone
        field_schema["type"] = "date"
        # More specific date constraints
        if "dob" in field_lower or "birth_date" in field_lower or "year_of_birth" in field_lower :
            start_def = datetime.now() - timedelta(days=365*70) # Up to 70 years ago
            end_def = datetime.now() - timedelta(days=365*18)   # At least 18 years ago
        elif "joining" in field_lower or "hire_date" in field_lower:
            start_def = datetime.now() - timedelta(days=365*10) # Up to 10 years ago
            end_def = datetime.now()
        else: # Generic date
            start_def = datetime.now() - timedelta(days=365)
            end_def = datetime.now()
        field_schema["constraint"] = f"{start_def.strftime('%Y-%m-%d')} - {end_def.strftime('%Y-%m-%d')}"
    elif any(kw in field_lower for kw in ["id", "count", "age", "quantity", "salary", "year", "number", "level", "amount", "total", "seq", "num", "digit", "numeric", "integer", "score_value", "points"]):
        field_schema["type"] = "int"
        if "age" in field_lower: field_schema["constraint"] = "18-80"
        elif "year" in field_lower and not "year_of_birth" in field_lower : field_schema["constraint"] = f"{datetime.now().year - 10}-{datetime.now().year}" # For general year fields
        elif any(kw in field_lower for kw in ["salary", "amount", "total", "revenue", "income"]): field_schema["constraint"] = "1000-100000"
        elif any(kw in field_lower for kw in ["quantity", "count", "stock", "inventory"]): field_schema["constraint"] = "1-100"
        elif any(kw in field_lower for kw in ["level", "rank", "grade_numeric"]): field_schema["constraint"] = "1-10"
        elif any(kw in field_lower for kw in ["score_value", "points"]): field_schema["constraint"] = "0-100"
        else: field_schema["constraint"] = "0-1000"
    elif any(kw in field_lower for kw in ["rate", "percent", "ratio", "score_decimal", "efficiency", "integrity", "price", "value", "balance", "temp", "humidity", "factor", "decimal", "float", "measurement", "latitude", "longitude", "coordinate"]):
        field_schema["type"] = "float"
        if any(kw in field_lower for kw in ["price", "value", "balance", "cost"]): field_schema["constraint"] = "0.0-1000.0"
        elif any(kw in field_lower for kw in ["rate", "percent", "ratio", "factor", "coefficient"]): field_schema["constraint"] = "0.0-1.0" # Often proportions
        elif "temp" in field_lower: field_schema["constraint"] = "-10.0-40.0"
        elif "humidity" in field_lower: field_schema["constraint"] = "0.0-100.0"
        elif any(kw in field_lower for kw in ["latitude"]): field_schema["constraint"] = "-90.0-90.0"
        elif any(kw in field_lower for kw in ["longitude"]): field_schema["constraint"] = "-180.0-180.0"
        else: field_schema["constraint"] = "0.0-100.0"
    elif "email" in field_lower: field_schema["type"] = "email"
    elif any(kw in field_lower for kw in ["phone", "mobile", "contact number"]): field_schema["type"] = "phone"
    elif any(kw in field_lower for kw in ["address", "location", "city", "state", "country", "pincode", "zipcode", "street", "area", "place"]):
        field_schema["type"] = "address" # Generic address type
        if "city" in field_lower: field_schema["is_faker_city"] = True
        elif "state" in field_lower: field_schema["is_faker_state"] = True
        elif "country" in field_lower: field_schema["is_faker_country"] = True
        elif any(kw in field_lower for kw in ["pincode", "zipcode"]): field_schema["is_faker_postcode"] = True
    # More specific name checks to avoid overly broad matching
    elif ("name" in field_lower and not any(kw in field_lower for kw in ["user_name", "company_name", "product_name", "brand_name", "model_name", "sensor_name", "item_name", "file_name", "column_name", "field_name", "campaign_name", "project_name", "program_name", "course_name", "animal_name", "login_name", "screen_name", "domain_name", "event_name", "team_name", "group_name", "font_name", "color_name", "style_name"])) \
        or field_lower in ["person_name", "full_name", "customer_name", "employee_name", "student_name", "doctor_name", "author_name", "contact_name"]:
        field_schema["type"] = "name"
    elif "gender" in field_lower:
        field_schema["type"] = "category"; field_schema["constraint"] = "Male,Female,Other,Prefer not to say"
    elif "status" in field_lower:
        field_schema["type"] = "category"; field_schema["constraint"] = "Active,Inactive,Pending,Completed,Cancelled,Open,Closed,Resolved,Shipped,Delivered"
    elif any(kw in field_lower for kw in ["type", "category", "kind", "class", "group", "segment", "tier"]):
        field_schema["type"] = "category"; field_schema["constraint"] = "Type A,Type B,Type C,Type D" # Generic categories
    elif "flag" in field_lower or field_lower.startswith("is_") or field_lower.startswith("has_") or field_lower.endswith("_enabled") or field_lower.endswith("_available"):
        field_schema["type"] = "category"; field_schema["constraint"] = random.choice(BOOLEAN_REPRESENTATIONS_LIST)
    elif any(kw in field_lower for kw in ["url", "website", "link", "webpage", "site"]):
        field_schema["type"] = "string"; field_schema["is_faker_url"] = True
    elif "ip_address" in field_lower or field_lower == "ip":
        field_schema["type"] = "string"; field_schema["is_faker_ipv4"] = True
    elif any(kw in field_lower for kw in ["color", "colour"]):
        field_schema["type"] = "string"; field_schema["is_faker_color_name"] = True
    elif any(kw in field_lower for kw in ["job", "position", "role", "occupation", "designation"]):
        field_schema["type"] = "string"; field_schema["is_faker_job"] = True
    elif any(kw in field_lower for kw in ["company", "organization", "employer", "firm", "business", "vendor", "supplier", "client_company"]):
        field_schema["type"] = "string"; field_schema["is_faker_company"] = True
    elif "animal" in field_lower and "name" in field_lower: field_schema["type"] = "animal_name" # More specific
    elif "car" in field_lower and any(kw in field_lower for kw in ["dream", "favorite", "model", "type"]): field_schema["type"] = "string"; field_schema["_hint_vehicle"] = True
    elif "festival" in field_lower and any(kw in field_lower for kw in ["favorite", "preferred"]): field_schema["type"] = "string"; field_schema["_hint_event_name"] = True
    elif "birthstone" in field_lower:
        field_schema["type"] = "category"; field_schema["constraint"] = "Garnet,Amethyst,Aquamarine,Diamond,Emerald,Pearl,Ruby,Peridot,Sapphire,Opal,Topaz,Turquoise"
    elif "zodiac_sign" == field_lower or ("zodiac" in field_lower and "sign" in field_lower):
        field_schema["type"] = "category"; field_schema["constraint"] = "Aries,Taurus,Gemini,Cancer,Leo,Virgo,Libra,Scorpio,Sagittarius,Capricorn,Aquarius,Pisces"
    elif "blood_group" == field_lower or ("blood" in field_lower and ("group" in field_lower or "type" in field_lower)):
        field_schema["type"] = "category"; field_schema["constraint"] = ",".join(BLOOD_TYPES_LIST)
    elif any(kw in field_lower for kw in ["favorite", "dream", "preference", "choice", "opinion", "hobby", "interest", "skill"]): # For general preference-like text
        field_schema["type"] = "string"; field_schema["_hint_short_phrase"] = True
    # Infer alphanumeric ID for novel fields that look like codes/keys but weren't caught by canonicals
    elif field_schema["type"] == "string" and any(kw in field_lower for kw in ["code", "key", "token", "serial_no", "reference_id", "identifier_code", "tracking_id", "item_code", "product_code", "user_token", "api_key_value"]) and not any(kw in field_lower for kw in ["currency", "country_code", "zip_code", "area_code", "ifsc_code", "promo_code", "discount_code"]): # Avoid things that are already specific types
        field_schema["_is_inferred_alphanum_id"] = True
        # Try to make a somewhat meaningful prefix from the display name
        prefix_candidate = "".join(filter(str.isalnum, display_name.replace(" ", "")))
        field_schema["_inferred_prefix"] = prefix_candidate.upper()[:min(len(prefix_candidate), 4)] + "-" if prefix_candidate else "ID-"
        field_schema["length"] = random.randint(6,12) # Default length for inferred IDs

    return field_schema

# --- NEW: Domain Prompt to Predefined Schema Mapping ---
# Maps common domain phrases to a list of canonical field keys
DOMAIN_PROMPT_TO_SCHEMA_MAP = {
    "medical data": [
        "patient_id", "patient_name", "age", "gender", "admission_date",
        "discharge_date", "diagnosis", "doctor_name", "hospital_name", "blood_pressure_reading",
        "blood_type", "medication", "phone", "email"
    ],
    "patient data": [ # New entry for "patient data"
        "patient_id", "name", "age", "gender", "admission_date", "diagnosis", "phone", "email"
    ],
    "hospital data": [
        "patient_id", "patient_name", "age", "gender", "admission_date",
        "discharge_date", "diagnosis", "doctor_name", "hospital_name", "blood_pressure_reading",
        "room_number", "insurance_provider", "blood_type", "medication", "phone", "email"
    ],
    "healthcare data": [
        "patient_id", "patient_name", "age", "gender", "admission_date",
        "discharge_date", "diagnosis", "doctor_name", "hospital_name",
        "insurance_provider", "blood_type", "medication", "phone", "email"
    ],
    "e-commerce data": [
        "order_id", "customer_name", "email", "product_name", "quantity", "price",
        "order_date", "shipping_address", "shipping_status", "payment_method", "product_category"
    ],
    "retail data": [
        "order_id", "customer_name", "email", "product_name", "quantity", "price",
        "order_date", "shipping_address", "shipping_status", "payment_method", "product_category"
    ],
    "customer data": [ # New entry for "customer data"
        "id", "customer_name", "email", "phone", "address",
        "date", "category", "status" # Adding some generic fields
    ],
    "employee data": [
        "employee_id", "name", "email", "phone", "job", 
        "department", "salary", "date", 
        "age", "gender", "address"
    ],
    "hr data": [
        "employee_id", "name", "email", "phone", "job",
        "department", "salary", "date", "age", "gender", "address"
    ],
    "financial data": [
        "transaction_id", "account_number", "date", "amount", "transaction_type",
        "description", "bank_name", "currency", "balance", "ifsc", "upi"
    ],
    "finance data": [ # Added for more direct matching
        "transaction_id", "account_number", "date", "amount", "transaction_type",
        "description", "bank_name", "currency", "balance", "ifsc", "upi"
    ],
    "banking data": [
        "transaction_id", "account_number", "date", "amount", "transaction_type",
        "description", "bank_name", "currency", "balance", "ifsc", "upi", "customer_name"
    ],
    "social media data": [
        "id", "username", "post_text", "timestamp", "like_count",
        "share_count", "comment_text", "hashtags"
    ],
    "logistics data": [
        "shipment_id", "tracking_number", "carrier_name", "origin_location",
        "destination_location", "shipment_status_logistics", "estimated_delivery_date",
        "actual_delivery_date", "freight_cost", "package_weight_kg", "package_dimensions_cm"
    ],
    "supply chain data": [
        "shipment_id", "tracking_number", "carrier_name", "origin_location",
        "destination_location", "shipment_status_logistics", "estimated_delivery_date",
        "actual_delivery_date", "freight_cost", "package_weight_kg", "package_dimensions_cm", "product_name"
    ],
    "travel data": [
        "booking_id", "traveler_name", "destination_city_travel", "origin_city_travel",
        "travel_date", "return_date", "flight_number", "airline_name", "hotel_name",
        "room_type", "booking_status_travel", "total_travel_cost", "email", "phone"
    ],
    "tourism data": [
        "booking_id", "traveler_name", "destination_city_travel", "origin_city_travel",
        "travel_date", "return_date", "flight_number", "airline_name", "hotel_name",
        "room_type", "booking_status_travel", "total_travel_cost", "email", "phone", "travel_activity"
    ],
    "iot data": [
        "sensor_id", "timestamp", "temperature", "humidity", "latitude", "longitude", "sensor_type", "value"
    ],
    "sensor data": [
        "sensor_id", "timestamp", "temperature", "humidity", "latitude", "longitude", "sensor_type", "value"
    ],
    "academic data": [
        "id", "publication_title", "author_names", "journal_name", "publication_year", "doi", "keywords", "citation_count", "publication_type"
    ],
    "research data": [
        "id", "publication_title", "author_names", "journal_name", "publication_year", "doi", "keywords", "citation_count", "publication_type"
    ],
     "student data": [
        "student_id", "name", "age", "gender", "email", "phone", "address",
        "school_name", "grade", "subject", "marks", "attendance"
    ],
    "education data": [
        "student_id", "name", "age", "gender", "email", "school_name", "grade", "subject", "marks",
        "teacher_name", "course_name" 
    ],
    "real estate data": [
        "property_id", "property_type", "address", "city", "state", "pincode", "area",
        "price", "bedrooms", "bathrooms", "year_built", "status"
    ],
    "property data": [
        "property_id", "property_type", "address", "city", "state", "pincode", "area",
        "price", "bedrooms", "bathrooms", "year_built", "status"
    ],
    "food delivery data": [
        "order_id", "customer_name", "restaurant_name", "food_items", "order_total",
        "delivery_agent_name", "delivery_time_minutes", "delivery_rating", "delivery_address",
        "payment_mode", "delivery_status"
    ],
    "restaurant data": [
        "restaurant_name", "address", "city", "phone", "category", 
        "order_id", "food_items", "order_total", "delivery_status"
    ],

    # --- NEW DOMAINS MAPPINGS ---
    "sales data": [
        "lead_id", "customer_name", "email", "phone", "company", "lead_source",
        "deal_stage", "opportunity_amount", "close_date", "account_type", "industry"
    ],
    "crm data": [
        "lead_id", "customer_name", "email", "phone", "company", "lead_source",
        "deal_stage", "opportunity_amount", "close_date", "account_type", "industry", "last_contact_date"
    ],
    "marketing campaign data": [
        "campaign_id", "campaign_name", "campaign_type", "start_date", "end_date",
        "ad_spend", "impressions", "clicks", "ctr", "conversion_rate", "target_audience"
    ],
    "inventory data": [
        "sku", "product_name", "serial_number", "batch_number", "warehouse_location",
        "stock_level", "supplier_name", "last_stocked_date", "unit_cost"
    ],
    "manufacturing data": [
        "product_id", "product_name", "sku", "batch_number", "production_date", "quantity_produced",
        "machine_id", "operator_name", "quality_check_status"
    ],
    "project management data": [
        "project_id", "project_name", "task_id", "task_name", "assignee_name", "start_date",
        "due_date", "project_status", "task_priority", "estimated_hours", "actual_hours"
    ],
    "software development data": [
        "issue_id", "bug_id", "feature_id", "summary", "description", "reporter_name", "assignee_name",
        "issue_type", "issue_status", "priority", "created_date", "updated_date", "commit_hash", "repository_url"
    ],
    "website traffic data": [
        "session_id", "user_id", "timestamp_detailed", "ip_address", "url", "referrer_url",
        "user_agent", "country", "city", "page_views", "bounce_rate"
    ],
}

FIELD_GENERATORS = {
    "patient_id": lambda nr: [f"{random.randint(1000, 9999)}{random.randint(1000, 9999)}" for _ in range(nr)],
    "transaction_id": lambda nr: [f"{random.randint(1000, 9999)}{random.randint(1000, 9999)}" for _ in range(nr)],
    "order_id": lambda nr: [f"{random.randint(1000, 9999)}{random.randint(1000, 9999)}" for _ in range(nr)], # Will be caught by is_generic_numeric_id
    "student_id": lambda nr: [f"{random.randint(1000, 9999)}{random.randint(1000, 9999)}" for _ in range(nr)],
    "employee_id": lambda nr: [f"{random.randint(1000, 9999)}{random.randint(1000, 9999)}" for _ in range(nr)],
    "property_id": lambda nr: [f"{random.randint(1000, 9999)}{random.randint(1000, 9999)}" for _ in range(nr)],
    "id": lambda nr: [f"{random.randint(1000, 9999)}{random.randint(1000, 9999)}" for _ in range(nr)], # Will be caught by is_generic_numeric_id
    "patient_name": lambda nr: [fake.name() for _ in range(nr)],
    "customer_name": lambda nr: [fake.name() for _ in range(nr)],
    "student_name": lambda nr: [fake.name() for _ in range(nr)],
    "employee_name": lambda nr: [fake.name() for _ in range(nr)],
    "owner_name": lambda nr: [fake.name() for _ in range(nr)],
    "name": lambda nr: [fake.name() for _ in range(nr)],
    "age": lambda nr: [random.randint(18, 80) for _ in range(nr)],
    "gender": lambda nr: [random.choice(["Male", "Female", "Other"]) for _ in range(nr)],
    "doctor_name": lambda nr: [f"Dr. {fake.name()}" for _ in range(nr)],
    "hospital_name": lambda nr: [fake.company() + " Hospital" for _ in range(nr)],
    "diagnosis": lambda nr: [random.choice(DIAGNOSES_LIST) for _ in range(nr)],
    "admission_date": lambda nr: [fake.date_between(start_date="-1y", end_date="today").strftime("%Y-%m-%d") for _ in range(nr)],
    "discharge_date": lambda nr: [(fake.date_between(start_date="-1y", end_date="today") + timedelta(days=random.randint(1, 14))).strftime("%Y-%m-%d") for _ in range(nr)], # Simplified independent generation
    "room_number": lambda nr: [f"{random.randint(1, 10)}{random.choice(['A', 'B', 'C'])}" for _ in range(nr)],
    "insurance_provider": lambda nr: [random.choice(INSURANCE_PROVIDERS_LIST) for _ in range(nr)],
    "blood_type": lambda nr: [random.choice(BLOOD_TYPES_LIST) for _ in range(nr)],
    "medication": lambda nr: [random.choice(MEDICATIONS_LIST) for _ in range(nr)],
    "account_number": lambda nr: [str(random.randint(1000000000, 9999999999)) for _ in range(nr)],
    "amount": lambda nr: [round(random.uniform(1000, 10000), 2) for _ in range(nr)],
    "price": lambda nr: [round(random.uniform(100, 5000), 2) for _ in range(nr)], # Adjusted range for product price
    "salary": lambda nr: [round(random.uniform(20000, 200000), 2) for _ in range(nr)], # Adjusted range for salary
    "bank_name": lambda nr: [fake.company() + " Bank" for _ in range(nr)],
    "account_holder": lambda nr: [fake.name() for _ in range(nr)],
    "transaction_type": lambda nr: [random.choice(TRANSACTION_TYPES_LIST) for _ in range(nr)],
    "balance": lambda nr: [round(random.uniform(-5000, 50000), 2) for _ in range(nr)],
    "ifsc_code": lambda nr: [generate_ifsc() for _ in range(nr)],
    "branch": lambda nr: [fake.city() + " Branch" for _ in range(nr)],
    "upi_id": lambda nr: [generate_upi() for _ in range(nr)],
    "reference_number": lambda nr: [f"REF{random.randint(1000, 9999)}{random.randint(1000, 9999)}" for _ in range(nr)],
    "product_name": lambda nr: [random.choice(PRODUCT_NAMES_LIST) + " " + str(random.randint(1, 10)) for _ in range(nr)],
    "quantity": lambda nr: [random.randint(1, 5) for _ in range(nr)],
    "payment_method": lambda nr: [random.choice(PAYMENT_METHODS_LIST) for _ in range(nr)],
    "delivery_address": lambda nr: [fake.address().replace('\n', ', ') for _ in range(nr)],
    "customer_email": lambda nr: [fake.email() for _ in range(nr)],
    "customer_phone": lambda nr: [f"+91 {random.randint(7000, 9999)}{random.randint(100000, 999999)}" for _ in range(nr)],
    "product_category": lambda nr: [random.choice(PRODUCT_CATEGORIES_LIST) for _ in range(nr)],
    "discount": lambda nr: [round(random.uniform(0, 0.5), 2) for _ in range(nr)],
    "shipping_status": lambda nr: [random.choice(SHIPPING_STATUSES_LIST) for _ in range(nr)],
    "grade": lambda nr: [random.choice(GRADES_LIST) for _ in range(nr)],
    "school_name": lambda nr: [fake.company() + " School" for _ in range(nr)],
    "teacher_name": lambda nr: [f"Mr./Ms. {fake.last_name()}" for _ in range(nr)],
    "subject": lambda nr: [random.choice(SUBJECTS_LIST) for _ in range(nr)],
    "attendance": lambda nr: [f"{random.randint(70, 100)}%" for _ in range(nr)],
    "marks": lambda nr: [random.randint(50, 100) for _ in range(nr)],
    "parent_name": lambda nr: [fake.name() for _ in range(nr)],
    "department": lambda nr: [random.choice(DEPARTMENTS_LIST) for _ in range(nr)],
    "position": lambda nr: [random.choice(POSITIONS_LIST) for _ in range(nr)],
    "manager": lambda nr: [fake.name() for _ in range(nr)],
    "education": lambda nr: [random.choice(DEGREES_LIST) for _ in range(nr)],
    "skills": lambda nr: [", ".join(random.sample(SKILLS_LIST, random.randint(2, 4))) for _ in range(nr)],
    "performance_rating": lambda nr: [round(random.uniform(1, 5), 1) for _ in range(nr)],
    "property_type": lambda nr: [random.choice(PROPERTY_TYPES_LIST) for _ in range(nr)],
    "area": lambda nr: [f"{random.randint(500, 3000)} sq.ft." for _ in range(nr)],
    "location": lambda nr: [fake.city() for _ in range(nr)],
    "bedrooms": lambda nr: [random.randint(1, 5) for _ in range(nr)],
    "bathrooms": lambda nr: [random.randint(1, 3) for _ in range(nr)],
    "contact_phone": lambda nr: [f"+91 {random.randint(7000, 9999)}{random.randint(100000, 999999)}" for _ in range(nr)],
    "status": lambda nr: [random.choice(GENERAL_STATUSES_LIST + SHIPPING_STATUSES_LIST + REALESTATE_STATUSES_LIST) for _ in range(nr)], # Combined status
    "amenities": lambda nr: [", ".join(random.sample(AMENITIES_LIST, random.randint(1, 3))) for _ in range(nr)],
    "year_built": lambda nr: [random.randint(1990, 2023) for _ in range(nr)],
    "email": lambda nr: [fake.email() for _ in range(nr)],
    "phone": lambda nr: [f"+91 {random.randint(7000, 9999)}{random.randint(100000, 999999)}" for _ in range(nr)],
    "address": lambda nr: [fake.address().replace('\n', ', ') for _ in range(nr)],
    "date": lambda nr: [fake.date_between(start_date="-1y", end_date="today").strftime("%Y-%m-%d") for _ in range(nr)],
    "value": lambda nr: [round(random.uniform(10, 1000), 2) for _ in range(nr)], # Generic value
    "category": lambda nr: [random.choice(GENERAL_CATEGORIES_LIST) for _ in range(nr)],
    "description": lambda nr: [fake.sentence() for _ in range(nr)],
    "score": lambda nr: [round(random.uniform(0, 100), 1) for _ in range(nr)],
    "default_word": lambda nr: [fake.word() for _ in range(nr)] # Fallback generator
}
FIELD_GENERATORS.update({ # Adding some more direct Faker calls for new types
    "url": lambda nr: [fake.url() for _ in range(nr)],
    "ip_address": lambda nr: [fake.ipv4() for _ in range(nr)],
    "mac_address": lambda nr: [fake.mac_address() for _ in range(nr)],
    "file_name": lambda nr: [fake.file_name(extension=random.choice(FILE_EXTENSIONS_LIST)) for _ in range(nr)],
    "mime_type": lambda nr: [fake.mime_type() for _ in range(nr)],
})

# --- Synonym and Phrase Mapping for Fields ---
# Maps user-friendly terms/phrases to canonical field keys used in FIELD_GENERATORS
FIELD_SYNONYM_TO_CANONICAL_MAP = {
    # General
    "identifier": "id", "record id": "id",
    "full name": "name", "person name": "name",
    "years old": "age",
    "sex": "gender",
    "email address": "email",
    "phone number": "phone", "contact number": "phone", "mobile number": "phone",
    "location address": "address", "street address": "address",
    "specific date": "date", "transaction date": "date", "order date": "date", "join date": "date", "listing date": "date", "date of birth": "date", "dob": "date",
    "monetary value": "value",
    "type": "category", "classification": "category", # 'category' itself is a key
    "details": "description",
    "current status": "status", # 'status' itself is a key
    "rating": "score", # 'score' itself is a key
    "website address": "url", "web link": "url", "site url": "url",
    "internet protocol address": "ip_address", "ipv4 address": "ip_address",
    "media access control address": "mac_address",
    "software version": "version_number", "build number": "version_number",
    "yes/no flag": "boolean_flag", "true/false indicator": "boolean_flag", "is active": "boolean_flag",
    "event timestamp": "timestamp_detailed", "log time": "timestamp_detailed",
    "attachment name": "file_name", "document name": "file_name",

    # Hospital
    "patient id": "patient_id", "patient number": "patient_id", "medical record number": "patient_id", "mrn": "patient_id",
    "patient name": "patient_name",
    "doctor name": "doctor_name", "physician name": "doctor_name", "blood pressure": "blood_pressure_reading",
    "hospital name": "hospital_name", "clinic name": "hospital_name",
    "medical condition": "diagnosis", # 'diagnosis' itself is a key
    "admission date": "admission_date", "date of admission": "admission_date",
    "discharge date": "discharge_date", "date of discharge": "discharge_date",
    "room number": "room_number", "hospital room": "room_number",
    "insurance provider": "insurance_provider", "health insurance": "insurance_provider",
    "blood type": "blood_type", "blood group": "blood_type",
    "prescription": "medication", "drug name": "medication", # 'medication' itself is a key

    # Finance
    "transaction id": "transaction_id", "transaction number": "transaction_id", "txn id": "transaction_id",
    "account number": "account_number", "acct no": "account_number", "bank account number": "account_number",
    "transaction amount": "amount", # 'amount' itself is a key
    "bank name": "bank_name",
    "account holder": "account_holder", "account name": "account_holder",
    "transaction type": "transaction_type", "type of transaction": "transaction_type",
    "account balance": "balance", # 'balance' itself is a key
    "ifsc code": "ifsc_code", "ifsc": "ifsc_code",
    "bank branch": "branch", # 'branch' itself is a key
    "upi id": "upi_id", "upi address": "upi_id",
    "reference number": "reference_number", "ref no": "reference_number",

    # Ecommerce
    "order id": "order_id", "order number": "order_id",
    "customer name": "customer_name",
    "product name": "product_name", "item name": "product_name",
    "cost": "price", "product price": "price", # 'price' itself is a key
    "order total": "order_total", # Ensure "order total" maps to the canonical field
    "number of items": "quantity", # 'quantity' itself is a key
    "payment method": "payment_method", "mode of payment": "payment_method",
    "delivery address": "delivery_address", "shipping address": "delivery_address",
    "customer email": "customer_email",
    "customer phone": "customer_phone",
    "product category": "product_category", "item category": "product_category",
    "price reduction": "discount", # 'discount' itself is a key
    "shipping status": "shipping_status", "order status": "shipping_status",

    # Education (many are already direct matches or covered by general)
    "student id": "student_id", "student number": "student_id", "roll number": "student_id",
    "student name": "student_name",
    "class": "grade", "student grade": "grade", # 'grade' itself is a key
    "college name": "school_name", "university name": "school_name", # 'school_name' itself is a key
    "teacher name": "teacher_name", "instructor name": "teacher_name", "professor name": "teacher_name",
    "course name": "subject", # 'subject' itself is a key
    "student attendance": "attendance", # 'attendance' itself is a key
    "score obtained": "marks", "exam score": "marks", # 'marks' itself is a key
    "parent name": "parent_name", "guardian name": "parent_name",

    # Employee (many are already direct matches or covered by general)
    "employee id": "employee_id", "employee number": "employee_id", "staff id": "employee_id",
    "employee name": "employee_name", "staff name": "employee_name",
    "dept": "department", # 'department' itself is a key
    "compensation": "salary", "pay": "salary", "income": "salary", # 'salary' itself is a key
    "joining date": "date", "date of joining": "date", # Covered by general date
    "job title": "position", "designation": "position", "role": "position", # 'position' itself is a key
    "reporting manager": "manager", "supervisor": "manager", # 'manager' itself is a key
    "qualification": "education", "degree": "education", # 'education' itself is a key
    "abilities": "skills", "expertise": "skills", # 'skills' itself is a key
    "performance rating": "performance_rating", "appraisal score": "performance_rating",

    # Real Estate
    "property id": "property_id", "property number": "property_id", "listing id": "property_id",
    "type of property": "property_type", # 'property_type' itself is a key
    "size": "area", "square feet": "area", "sq ft": "area", # 'area' itself is a key
    "property location": "location", # 'location' itself is a key
    "number of bedrooms": "bedrooms", "beds": "bedrooms", # 'bedrooms' itself is a key
    "number of bathrooms": "bathrooms", "baths": "bathrooms", # 'bathrooms' itself is a key
    "seller name": "owner_name", # 'owner_name' itself is a key
    # "contact_phone" is direct
    "facilities": "amenities", # 'amenities' itself is a key
    "construction year": "year_built", # 'year_built' itself is a key

    # Add direct keys from FIELD_GENERATORS if they are not complex phrases

    # Food Delivery Synonyms
    "restaurant": "restaurant_name", "eatery name": "restaurant_name", "food place": "restaurant_name",
    "items ordered": "food_items", "dishes": "food_items", "menu items": "food_items",
    "bill amount": "order_total", "total cost": "order_total",
    "delivery boy name": "delivery_agent_name", "rider name": "delivery_agent_name", "delivery person": "delivery_agent_name",
    "restaurant name": "restaurant_name", # Added for exact phrase match
    "food items": "food_items",           # Added for exact phrase match
    "delivery agent": "delivery_agent_name", # Added for common phrase match
    "delivery agent name": "delivery_agent_name", # Added for exact phrase match
    "time to deliver": "delivery_time_minutes", "estimated delivery time": "delivery_time_minutes",
    "customer rating for delivery": "delivery_rating",
    "drop address": "delivery_address",
    "how paid": "payment_mode",
    "order current status": "delivery_status", "status of delivery": "delivery_status",
    **{key: key for key in CANONICAL_FIELD_TO_SCHEMA_DETAILS_MAP.keys()}, # Canonical names are their own primary synonym

    # Academic/Research Synonyms
    "paper title": "publication_title", "article name": "publication_title",
    "authors": "author_names", "researchers": "author_names",
    "published in": "journal_name",
    "year of publication": "publication_year",
    "citations": "citation_count", "cited by": "citation_count",
    "digital object identifier": "doi",
    "research keywords": "keywords", "tags": "keywords", # 'tags' also used in social media
    "type of publication": "publication_type",

    # Social Media Synonyms
    "user name": "username", "handle": "username", "screen name": "username",
    "post content": "post_text", "tweet text": "post_text",
    "likes": "like_count", "number of likes": "like_count",
    "shares": "share_count", "retweets": "share_count",
    "comment": "comment_text", "reply": "comment_text",
    # "hashtags" is direct

    # IoT/Sensor Synonyms
    "device id": "sensor_id",
    "reading time": "timestamp", "event time": "timestamp",
    "temp": "temperature",
    # "humidity", "latitude", "longitude", "sensor_type" are fairly direct

    # Animal Data Synonyms
    "animal": "animal_name", # General term, might default to name
    "animal species": "species", "type of animal": "species",
    "animal breed": "breed",
    "age of animal": "animal_age",
    "natural environment": "habitat", "lives in": "habitat",
    "animal weight": "animal_weight_kg",
    "fur color": "animal_color", "coat color": "animal_color",
}
FIELD_SYNONYM_TO_CANONICAL_MAP.update({
    # Logistics Synonyms
    "shipment number": "shipment_id", "consignment id": "shipment_id",
    "tracking id": "tracking_number", "awb number": "tracking_number", "air waybill": "tracking_number",
    "shipping company": "carrier_name", "freight carrier": "carrier_name", "courier": "carrier_name",
    "pickup location": "origin_location", "despatch point": "origin_location", "from address": "origin_location", "origin": "origin_location",
    "delivery location": "destination_location", "drop-off point": "destination_location", "to address": "destination_location", "destination": "destination_location",
    "logistics status": "shipment_status_logistics", "delivery progress": "shipment_status_logistics",
    "shipping cost": "freight_cost", "transportation charges": "freight_cost",
    "edd": "estimated_delivery_date", "expected delivery": "estimated_delivery_date",
    "add": "actual_delivery_date", "delivered on": "actual_delivery_date",
    "parcel weight": "package_weight_kg", "shipment weight": "package_weight_kg",
    "box size": "package_dimensions_cm", "package size": "package_dimensions_cm", "dimensions": "package_dimensions_cm",
})
FIELD_SYNONYM_TO_CANONICAL_MAP.update({
    # Travel & Tourism Synonyms
    "booking reference": "booking_id", "reservation id": "booking_id", "confirmation number": "booking_id",
    "passenger name": "traveler_name", "guest name": "traveler_name",
    "travel destination": "destination_city_travel", "going to": "destination_city_travel", "arrival city": "destination_city_travel",
    "departure city": "origin_city_travel", "flying from": "origin_city_travel", "source city": "origin_city_travel",
    "departure date": "travel_date", "check-in date": "travel_date", "start date": "travel_date",
    "arrival date": "return_date", "check-out date": "return_date", "end date": "return_date", # Can be ambiguous, but common
    "flight code": "flight_number", "flight id": "flight_number",
    "airline company": "airline_name", "carrier": "airline_name", # 'carrier' also in logistics
    "accommodation name": "hotel_name", "place to stay": "hotel_name", "resort name": "hotel_name",
    "room category": "room_type", "type of room": "room_type",
    "reservation status": "booking_status_travel", "booking state": "booking_status_travel",
    "trip cost": "total_travel_cost", "package price": "total_travel_cost", "total fare": "total_travel_cost",
    "tour package": "travel_package_name", "vacation package": "travel_package_name",
    "hotel features": "hotel_amenities_included", "included services": "hotel_amenities_included",
    "booked activity": "travel_activity", "excursion": "travel_activity", "tour name": "travel_activity",
})

FIELD_SYNONYM_TO_CANONICAL_MAP.update({
    # Sales/CRM Synonyms
    "lead identifier": "lead_id",
    "origin of lead": "lead_source",
    "sales stage": "deal_stage", "pipeline stage": "deal_stage",
    "deal size": "opportunity_amount", "potential revenue": "opportunity_amount",
    "expected close date": "close_date",
    "customer type": "account_type", "client type": "account_type",
    "business sector": "industry", "market segment": "industry",

    # Marketing Synonyms
    "marketing campaign id": "campaign_id",
    "promotion name": "campaign_name",
    "type of campaign": "campaign_type",
    "money spent on ads": "ad_spend", "advertising cost": "ad_spend",
    "number of views": "impressions",
    "number of clicks": "clicks",
    "click rate": "ctr",
    "conversion percentage": "conversion_rate",

    # Manufacturing/Inventory Synonyms
    "stock keeping unit": "sku",
    "product serial": "serial_number",
    "production batch": "batch_number",
    "storage location": "warehouse_location",
    "inventory count": "stock_level", "quantity on hand": "stock_level",
    "vendor name": "supplier_name",

    # Project Management Synonyms
    "project identifier": "project_id",
    "task identifier": "task_id",
    "assigned to": "assignee_name", "responsible person": "assignee_name",
    "deadline": "due_date", "target date": "due_date",
    "current project phase": "project_status",
    "task importance": "task_priority",

    # Software Development Synonyms
    "bug report id": "bug_id", "defect id": "bug_id",
    "git commit id": "commit_hash", "revision code": "commit_hash",
})
# Generate Synthetic Data Based on Input (Simplified prompt-based, distinct from domain-specific)
def generate_synthetic_data(description):
    # Seeding is now handled globally based on the 'use_fixed_seed' checkbox
    description_lower_original = description.lower() # Keep original lowercased prompt for row parsing

    # Default values
    num_rows = 10
    columns = []

    # --- NEW: Pre-process to remove common instructional phrases ---
    # Start with the original lowercased prompt
    processed_description_for_matching = description_lower_original
    instructional_phrases_to_remove = [
        "give me ", "generate ", "create ", "i want ", "a dataset of ", "dataset of ",
        "show me ", "provide ", "make ", "can i have ", "i need "
    ]
    for phrase_to_remove in instructional_phrases_to_remove:
        if processed_description_for_matching.startswith(phrase_to_remove):
            processed_description_for_matching = processed_description_for_matching[len(phrase_to_remove):].strip()
            break # Remove only the first matching prefix, then strip


    # --- Helper functions for parsing constraints from phrases ---
    def _parse_numeric_constraint_from_phrase(text_phrase):
        # Returns (constraint_str, remaining_text_str, type_hint_str) or (None, original_text_str, None)
        
        # Pattern for "between X and Y" or "from X to Y" (allows currency symbols)
        match_between = re.search(r"(?:between|from)\s*(?:[RrSs$₹€£¥]\.?\s*)?(-?\d+(?:\.\d+)?)\s*(?:and|to)\s*(?:[RrSs$₹€£¥]\.?\s*)?(-?\d+(?:\.\d+)?)", text_phrase, re.IGNORECASE)
        if match_between:
            min_val, max_val = match_between.groups()
            constraint_str = f"{min_val}-{max_val}"
            remaining_text = text_phrase.replace(match_between.group(0), "").strip()
            type_hint = "float" if "." in constraint_str else "int"
            return constraint_str, remaining_text, type_hint

        # Pattern for "X-Y" or "X - Y" (allows currency symbols)
        match_range_hyphen = re.search(r"(?:[RrSs$₹€£¥]\.?\s*)?(-?\d+(?:\.\d+)?)\s*-\s*(?:[RrSs$₹€£¥]\.?\s*)?(-?\d+(?:\.\d+)?)", text_phrase)
        if match_range_hyphen:
            min_val, max_val = match_range_hyphen.groups()
            constraint_str = f"{min_val}-{max_val}"
            remaining_text = text_phrase.replace(match_range_hyphen.group(0), "").strip()
            type_hint = "float" if "." in constraint_str else "int"
            return constraint_str, remaining_text, type_hint
        
        return None, text_phrase, None

    def _parse_category_constraint_from_phrase(text_phrase):
        # Returns (constraint_str, remaining_text_str, type_hint_str) or (None, original_text_str, None)
        match_category = re.search(r"(?:options|values|is one of|categories are|like)\s*:?\s*((?:[\w\s\-]+)(?:,\s*[\w\s\-]+)*)", text_phrase, re.IGNORECASE)
        if match_category:
            categories_str = match_category.group(1).strip()
            categories_str = ",".join([c.strip() for c in categories_str.split(',')]) # Clean individual items
            remaining_text = text_phrase.replace(match_category.group(0), "").strip()
            return categories_str, remaining_text, "category"
        return None, text_phrase, None
    # --- End of helper functions ---

    parsed_schema_fields = [] # This will hold {"name": display_name, "type": type, "constraint": constraint, ...}
    detected_pii = []
    detected_dpdp = []

    # Parse number of rows
    match_rows = re.search(r"(\d+)\s*(rows|records|entries|row|record|entry)", description_lower_original, re.IGNORECASE)
    if match_rows:
        num_rows = int(match_rows.group(1))
        # Remove row specifier from the description used for field/domain matching
        processed_description_for_matching = processed_description_for_matching.replace(match_rows.group(0), "", 1).strip()

    # PROMPT_FIELD_CONFIG was here, but it's deprecated and removed.

    # --- New Parsing Logic using FIELD_SYNONYM_TO_CANONICAL_MAP and CANONICAL_FIELD_TO_SCHEMA_DETAILS_MAP ---
    # Use the processed_description_for_matching for domain and field parsing
    remaining_description_for_parsing = processed_description_for_matching
    processed_canonical_fields = set()
    domain_matched = False

    # Sort domain keywords by length (descending) to match longer phrases first
    sorted_domain_keywords = sorted(DOMAIN_PROMPT_TO_SCHEMA_MAP.keys(), key=len, reverse=True)

    for domain_keyword in sorted_domain_keywords:
        if domain_keyword in remaining_description_for_parsing: # Check against the potentially cleaned description
            st.info(f"Recognized domain: '{domain_keyword}'. Generating predefined schema.")
            canonical_fields_for_domain = DOMAIN_PROMPT_TO_SCHEMA_MAP[domain_keyword]
            for canonical_field in canonical_fields_for_domain:
                if canonical_field in CANONICAL_FIELD_TO_SCHEMA_DETAILS_MAP:
                    schema_detail = CANONICAL_FIELD_TO_SCHEMA_DETAILS_MAP[canonical_field]
                    parsed_schema_fields.append({
                        "name": schema_detail.get("display_name", canonical_field.replace("_", " ").title()),
                        "type": schema_detail["type"],
                        "constraint": schema_detail["constraint"],
                        "pii_handling": st.session_state.get(DEFAULT_PII_STRATEGY_KEY, "realistic_fake"),
                        "_original_canonical": canonical_field
                    })
                    processed_canonical_fields.add(canonical_field) # Track processed fields
                    # Add to detected PII/DPDP lists
                    display_name = schema_detail.get("display_name", canonical_field.replace("_", " ").title())
                    if display_name in PII_FIELDS or canonical_field in PII_FIELDS:
                        if display_name not in detected_pii: detected_pii.append(display_name)
                    if is_dpdp_pii(display_name) or is_dpdp_pii(canonical_field):
                        if display_name not in detected_dpdp: detected_dpdp.append(display_name)
            domain_matched = True
            break # Stop after first domain match

    if not domain_matched:
        # Original logic for parsing individual fields if no domain was matched
        # Sort synonyms by length (descending) to match longer phrases first
        sorted_synonyms = sorted(FIELD_SYNONYM_TO_CANONICAL_MAP.keys(), key=len, reverse=True)

        # Split prompt by common delimiters like ",", "and", "with" to isolate potential field names
        potential_field_phrases = re.split(r'\s*(?:,|and|with)\s*', remaining_description_for_parsing)

        for phrase in potential_field_phrases:
            phrase = phrase.strip()
            if not phrase:
                continue
            
            parsed_custom_constraint = None
            field_name_part_for_synonym = phrase 
            parsed_type_hint_from_constraint = None

            # 1. Try to parse constraints, which might shorten field_name_part_for_synonym
            #    and provide a type hint.
            #    Order: numeric first, then category if numeric doesn't match.
            parsed_custom_constraint, temp_field_name_part, parsed_type_hint_from_constraint = _parse_numeric_constraint_from_phrase(phrase)
            if parsed_custom_constraint:
                field_name_part_for_synonym = temp_field_name_part.strip()
            else: # If not numeric, try category
                parsed_custom_constraint, temp_field_name_part, parsed_type_hint_from_constraint = _parse_category_constraint_from_phrase(phrase)
                if parsed_custom_constraint:
                    field_name_part_for_synonym = temp_field_name_part.strip()
            
            if not field_name_part_for_synonym: # If constraint parsing consumed the whole phrase (unlikely but possible)
                st.warning(f"Could not determine field name from phrase: '{phrase}'. Skipping.")
                continue
            
            # 2. Try to match the (potentially shortened) field_name_part_for_synonym against synonyms
            matched_canonical = None
            for user_synonym in sorted_synonyms:
                if user_synonym in field_name_part_for_synonym.lower(): 
                    canonical_field = FIELD_SYNONYM_TO_CANONICAL_MAP[user_synonym]
                    if canonical_field in CANONICAL_FIELD_TO_SCHEMA_DETAILS_MAP and canonical_field not in processed_canonical_fields:
                        matched_canonical = canonical_field
                        break # Found a match for this phrase part

            if matched_canonical:
                base_schema_detail = CANONICAL_FIELD_TO_SCHEMA_DETAILS_MAP[matched_canonical]
                
                final_type = base_schema_detail["type"]
                final_constraint = base_schema_detail["constraint"]

                if parsed_custom_constraint:
                    final_constraint = parsed_custom_constraint
                    if parsed_type_hint_from_constraint: # Override type if constraint implies it
                        final_type = parsed_type_hint_from_constraint
                
                parsed_schema_fields.append({
                    "name": base_schema_detail.get("display_name", matched_canonical.replace("_", " ").title()),
                    "type": final_type,
                    "constraint": final_constraint,
                    "pii_handling": st.session_state.get(DEFAULT_PII_STRATEGY_KEY, "realistic_fake"), 
                    "_original_canonical": matched_canonical
                })
                processed_canonical_fields.add(matched_canonical)
                display_name = base_schema_detail.get("display_name", matched_canonical.replace("_", " ").title())
                if display_name in PII_FIELDS or matched_canonical in PII_FIELDS:
                    if display_name not in detected_pii: detected_pii.append(display_name)
                if is_dpdp_pii(display_name) or is_dpdp_pii(matched_canonical):
                    if display_name not in detected_dpdp: detected_dpdp.append(display_name)
            else: # No canonical match, treat field_name_part_for_synonym as a novel field
                novel_field_name_candidate = field_name_part_for_synonym
                # Clean the candidate name by removing common instructional prefixes that might have been missed
                # or are part of a sub-phrase.
                instructional_prefixes_for_novel = [
                    "a field for ", "a column for ", "field for ", "column for ",
                    "data for ", "values for ", "entries for ",
                    "generate ", "create ", "include ", "add "
                ]
                for prefix_to_remove in instructional_prefixes_for_novel:
                    if novel_field_name_candidate.lower().startswith(prefix_to_remove):
                        novel_field_name_candidate = novel_field_name_candidate[len(prefix_to_remove):].strip()
                        # Update field_lower if novel_field_name_candidate changed
                        field_lower = novel_field_name_candidate.lower() # Re-assign field_lower
                        break # Remove only one prefix

                # Basic cleaning: remove "field like", "fields like", "column like", "columns like"
                novel_field_name_candidate = re.sub(r'\b(?:fields?|columns?)\s+like\b', '', novel_field_name_candidate, flags=re.IGNORECASE).strip()
                
                if not novel_field_name_candidate or novel_field_name_candidate in processed_canonical_fields or any(f['name'].lower() == novel_field_name_candidate.lower() for f in parsed_schema_fields):
                    continue

                inferred_type = "string" # Default
                inferred_constraint = parsed_custom_constraint if parsed_custom_constraint else ""
                field_lower = novel_field_name_candidate.lower()

                if parsed_type_hint_from_constraint: # Type hinted by parsed constraint
                    inferred_type = parsed_type_hint_from_constraint
                elif novel_field_name_candidate.isdigit(): # Check if the novel field name is purely digits
                    inferred_type = "int"
                else: # No constraint parsed, or constraint didn't hint type, so infer from name
                    if any(kw in field_lower for kw in ["date", "time", "timestamp", "dob", "joining"]):
                        inferred_type = "date"
                    elif any(kw in field_lower for kw in ["id", "count", "age", "quantity", "salary", "year", "number", "level", "hours", "duration"]):
                        inferred_type = "int"
                        if not inferred_constraint: # Only set default constraint if none was parsed
                            if "age" in field_lower: inferred_constraint = "18-80"
                            elif any(h_kw in field_lower for h_kw in ["hours", "duration"]): inferred_constraint = "0-100"
                            else: inferred_constraint = "0-1000"
                    elif any(kw in field_lower for kw in ["rate", "percent", "ratio", "score", "efficiency", "integrity", "price", "amount", "value", "balance", "temp", "humidity", "factor"]):
                        inferred_type = "float"
                        if not inferred_constraint: inferred_constraint = "0.0-100.0"
                    elif "email" in field_lower: inferred_type = "email"
                    elif "phone" in field_lower: inferred_type = "phone"
                    elif any(kw in field_lower for kw in ["address", "location", "city", "state", "country", "pincode", "zipcode"]): inferred_type = "address"
                    elif "name" in field_lower and not any(kw in field_lower for kw in ["user", "company", "product", "brand", "model", "sensor", "item", "file", "column", "field"]): inferred_type = "name"

                display_name = novel_field_name_candidate.replace("_", " ").title()
                parsed_schema_fields.append({
                    "name": display_name,
                    "type": inferred_type, # This will be used by generate_value
                    "constraint": inferred_constraint, # This will be used by generate_value
                    "pii_handling": st.session_state.get(DEFAULT_PII_STRATEGY_KEY, "realistic_fake"),
                    "_original_canonical": display_name.lower().replace(" ", "_"), # Treat as its own canonical for now
                    # Add a flag if this novel field seems like an alphanumeric ID
                    "_is_inferred_alphanum_id": True if inferred_type == "string" and \
                                                    any(kw in field_lower for kw in ["code", "identifier", "serial", "reference", "token", "key", "uid", "guid"]) and \
                                                    not any(kw in field_lower for kw in ["id","number"]) and \
                                                    not any(kw in field_lower for kw in ["email", "phone", "address", "name", "date", "url", "description", "text", "comment", "note", "message", "title", "subject", "label", "caption"]) \
                                                    else False,
                    "_inferred_prefix": (
                        "ORD-" if "order" in field_lower and any(kw in field_lower for kw in ["code", "identifier", "ref"]) else
                        "PROD-" if "product" in field_lower and any(kw in field_lower for kw in ["code", "identifier", "sku"]) else
                        "USR-" if "user" in field_lower and any(kw in field_lower for kw in ["code", "identifier", "key"]) else
                        "TK-" if "token" in field_lower else
                        "KEY-" if "key" in field_lower and "api" not in field_lower else # Avoid "api key" becoming KEY-
                        "SN-" if "serial" in field_lower else
                        "REF-" if "reference" in field_lower else
                        ""
                    ) if inferred_type == "string" and any(kw in field_lower for kw in ["code", "identifier", "serial", "reference", "token", "key", "uid", "guid"]) else ""
                })
                processed_canonical_fields.add(display_name.lower().replace(" ", "_"))
                if display_name in PII_FIELDS or any(pii_kw.lower() in display_name.lower() for pii_kw in PII_FIELDS): # Check PII for novel fields too
                    if display_name not in detected_pii: detected_pii.append(display_name)
                if is_dpdp_pii(display_name):
                    if display_name not in detected_dpdp: detected_dpdp.append(display_name)

    # Ensure we have at least some columns for simple generation
    if not parsed_schema_fields:
        # Check if the prompt was *only* for rows (e.g., "10 rows")
        # Remove digits and common row-specifying words to see if anything substantial is left
        prompt_content_check = re.sub(r'\d+', '', remaining_description_for_parsing).lower() # Check the content part
        prompt_content_check = prompt_content_check.replace("rows", "").replace("row", "").replace("records", "").replace("record", "").replace("entries", "").replace("entry", "").strip()
        
        if not prompt_content_check: # User likely just asked for a number of rows
            st.info(f"No specific fields requested. Generating a default dataset with {num_rows} rows (ID, Random Text, Random Number).")
            data = {
                "ID": [i + 1 for i in range(num_rows)],
                "Random Text": [fake.word() for _ in range(num_rows)],
                "Random Number": [random.randint(1, 100) for _ in range(num_rows)]
            }
            synthetic_df = pd.DataFrame(data)
            return synthetic_df, num_rows, [] # Return consistent tuple
        else: # User provided other words, but they weren't recognized
            st.error(
                "No valid columns detected from your prompt for simple generation. "
                "Please use keywords like name, age, salary, email, state, country, etc., or try the Smart Schema Editor for more control."
            )
            return None
    
    # Generate data using the parsed schema
    data = {}
    for field_schema_item in parsed_schema_fields:
        col_display_name = field_schema_item["name"]
        # Special handling for Faker direct calls if specified in CANONICAL_FIELD_TO_SCHEMA_DETAILS_MAP
        # This is a bridge for simple Faker calls that don't fit neatly into generate_value types
        # (e.g. city, state, country, job, company, currency)
        canonical_key = field_schema_item.get("_original_canonical")
        schema_details_for_canonical = CANONICAL_FIELD_TO_SCHEMA_DETAILS_MAP.get(canonical_key, {})

        # --- NEW: Dispatch Logic for Canonical-Specific Generation ---
        canonical_generator_map = {
            "is_faker_city": lambda: [fake.city() for _ in range(num_rows)],
            "is_faker_state": lambda: [fake.state() for _ in range(num_rows)],
            "is_faker_country": lambda: [fake.country() for _ in range(num_rows)],
            "is_faker_company_with_suffix_list": lambda: [f"{fake.company()} {random.choice(schema_details_for_canonical['suffix_from_list'])}" for _ in range(num_rows)], # Combined check
            "is_faker_company": lambda: [f"{schema_details_for_canonical.get('prefix', '').rstrip()} {fake.company()} {schema_details_for_canonical.get('suffix', '').lstrip()}".strip() for _ in range(num_rows)], # For company w/ prefix/suffix
            "is_faker_postcode": lambda: [fake.postcode() for _ in range(num_rows)],
            "is_faker_currency_code": lambda: [fake.currency_code() for _ in range(num_rows)],
            "is_faker_job": lambda: [fake.job() for _ in range(num_rows)],
            "is_generic_numeric_id": lambda: [f"{random.randint(1000, 9999)}{random.randint(1000, 9999)}" for _ in range(num_rows)],
            "is_room_number_pattern": lambda: [f"{random.randint(1, 20)}{random.choice(['A', 'B', 'C', 'D'])}" for _ in range(num_rows)],
            "is_percentage_pattern": lambda: [f"{random.randint(*map(int, schema_details_for_canonical.get('constraint', '0-100').split('-')))}%" for _ in range(num_rows)],
            "is_measurement_pattern": lambda: [f"{random.randint(*map(int, schema_details_for_canonical.get('constraint', '1-100').split('-')))} {schema_details_for_canonical.get('unit', '').strip()}" for _ in range(num_rows)],
            "is_reference_number_pattern": lambda: [f"REF{random.randint(10000, 99999)}{random.randint(100, 999)}" for _ in range(num_rows)],
            "is_faker_sentence": lambda: [fake.sentence() for _ in range(num_rows)],
            "is_faker_paragraph": lambda: [fake.paragraph(nb_sentences=3) for _ in range(num_rows)],
            "is_faker_user_name": lambda: [fake.user_name() for _ in range(num_rows)],
            "is_faker_latitude": lambda: [fake.latitude() for _ in range(num_rows)],
            "is_faker_longitude": lambda: [fake.longitude() for _ in range(num_rows)],
            "is_faker_color_name": lambda: [fake.color_name() for _ in range(num_rows)],
            "is_doi_pattern": lambda: [f"10.{random.randint(1000, 9999)}/{''.join(random.choices('abcdefghijklmnopqrstuvwxyz0123456789', k=6))}" for _ in range(num_rows)],
            "is_faker_url": lambda: [fake.url() for _ in range(num_rows)],
            "is_faker_ipv4": lambda: [fake.ipv4() for _ in range(num_rows)],
            "is_faker_mac_address": lambda: [fake.mac_address() for _ in range(num_rows)],
            "is_faker_file_name": lambda: [fake.file_name(extension=random.choice(FILE_EXTENSIONS_LIST)) for _ in range(num_rows)],
            "is_faker_mime_type": lambda: [fake.mime_type() for _ in range(num_rows)],
            "is_generic_alphanum_id": lambda: [f"{schema_details_for_canonical.get('prefix', '')}{''.join(random.choices('ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789', k=8))}" for _ in range(num_rows)],
            "is_keywords_list": lambda: [", ".join(fake.words(nb=random.randint(2, 5))) for _ in range(num_rows)],
            "is_tracking_number_pattern": lambda: [f"{''.join(random.choices('ABCDEFGHIJKLMNOPQRSTUVWXYZ', k=3))}{random.randint(100000000, 999999999)}{''.join(random.choices('ABCDEFGHIJKLMNOPQRSTUVWXYZ', k=2))}" for _ in range(num_rows)],
            "is_dimension_pattern": lambda: [f"{random.randint(10, 100)}x{random.randint(10, 100)}x{random.randint(5, 50)} cm" for _ in range(num_rows)],
            "is_flight_number_pattern": lambda: [f"{''.join(random.choices('ABCDEFGHIJKLMNOPQRSTUVWXYZ', k=2))}{random.randint(100, 9999)}" for _ in range(num_rows)],
            "is_multi_name": lambda: ["; ".join([fake.name() for _ in range(random.randint(1, 4))]) for _ in range(num_rows)],
            "is_multi_category": lambda: [", ".join(random.sample(schema_details_for_canonical['constraint'].split(','), k=random.randint(1, min(3, len(schema_details_for_canonical['constraint'].split(',')))))) for _ in range(num_rows)] if schema_details_for_canonical['constraint'] else [fake.word() for _ in range(num_rows)],
            "is_faker_city_if_empty_constraint": lambda: [random.choice(schema_details_for_canonical['constraint'].split(',')) for _ in range(num_rows)] if schema_details_for_canonical['constraint'] and schema_details_for_canonical['constraint'].strip() else [fake.word() for _ in range(num_rows)],
            "is_version_number_pattern": lambda: f"{random.randint(0,5)}.{random.randint(0,20)}.{random.randint(0,100)}",
            "is_faker_credit_card_number": lambda: [fake.credit_card_number() for _ in range(num_rows)],
            "is_digit_sequence": lambda: [''.join(random.choices('0123456789', k=random.randint(int(schema_details_for_canonical['constraint'].split(':')[1].split('-')[0]), int(schema_details_for_canonical['constraint'].split(':')[1].split('-')[1])))) for _ in range(num_rows)] if schema_details_for_canonical.get('constraint', '').startswith('digits:') else [''.join(random.choices('0123456789', k=10)) for _ in range(num_rows)], # Fallback to 10 digits if constraint malformed
        }
        
        # Check combined condition for company + suffix list *before* general company name generation
        if schema_details_for_canonical.get("suffix_from_list") and schema_details_for_canonical.get("is_faker_company"):
            generator_key = "is_faker_company_with_suffix_list"
        else:
            # Otherwise find the first key that matches in schema_details_for_canonical
            generator_key = next((key for key in canonical_generator_map if schema_details_for_canonical.get(key)), None)
        
        if generator_key:
            data[col_display_name] = canonical_generator_map[generator_key]()

        else:
            # For types like "name" that might have prefix/suffix, pass the schema_details
            if field_schema_item["type"] == "name":
                 # Temporarily override field_schema_item for generate_value to include details from CANONICAL_FIELD_TO_SCHEMA_DETAILS_MAP
                 # This is a bit of a hack; ideally generate_value always gets the full effective schema for the field.
                 effective_schema_for_name = {**field_schema_item, **schema_details_for_canonical}
                 data[col_display_name] = [_generate_value_from_schema(effective_schema_for_name) for _ in range(num_rows)]
            else:
                 data[col_display_name] = [_generate_value_from_schema(field_schema_item) for _ in range(num_rows)]

    # Create DataFrame
    synthetic_df = pd.DataFrame(data)

    # Display PII warning if detected
    if detected_pii:
        st.warning(f"⚠️ Detected possible PII fields: {', '.join(detected_pii)}. Ensure compliance with privacy regulations.")

    # Display DPDP-specific warnings
    if detected_dpdp:
        for field in detected_dpdp:
            st.markdown(f"""
            <div class="dpdp-warning">
            🔐 <strong>DPDP Compliance Note</strong>: "{field}" contains PII under India's DPDP Act
            </div>
            """, unsafe_allow_html=True)

    return synthetic_df, num_rows, parsed_schema_fields # Modified return

# --- Value Generation Dispatcher ---
VALUE_GENERATOR_FUNCTIONS = {
    # "string" is now handled directly in generate_value to pass full_field_schema
    "int": _generate_int_value,
    "float": _generate_float_value,
    "date": _generate_date_value,
    "category": _generate_category_value,
    "email": _generate_email_value,
    "phone": _generate_phone_value,
    "address": _generate_address_value,
    "name": _generate_name_value, # _generate_name_value will be enhanced
    "aadhaar": _generate_aadhaar_value,
    "pan": _generate_pan_value,
    "passport": _generate_passport_value,
    "voterid": _generate_voterid_value,
    "ifsc": _generate_ifsc_value,
    "upi": _generate_upi_value,
    "animal_name": _generate_animal_name_value, # Registering the new generator
}# ... (other imports)
from datetime import datetime # Already imported, ensure it's available

# ... (rest of your existing code)

# --- NEW: Explainability Report Function ---
def generate_explainability_pdf(generation_context_info, generated_dfs_info):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, f"{APP_NAME} - Data Generation Explainability Report", 0, 1, "C")
    pdf.set_font("Arial", size=10)
    pdf.cell(0, 7, f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", 0, 1, "C")
    pdf.ln(5)

    # --- Generation Overview ---
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, "1. Generation Overview", 0, 1)
    pdf.set_font("Arial", size=10)
    pdf.multi_cell(0, 5, f"Generation Method: {generation_context_info.get('method', 'N/A')}")
    pdf.multi_cell(0, 5, f"Reproducibility (Fixed Seed Used): {'Yes' if st.session_state.get('use_fixed_seed', False) else 'No'}")
    pdf.ln(3)

    # --- Input Configuration ---
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, "2. Input Configuration", 0, 1)
    pdf.set_font("Arial", size=10)

    if generation_context_info.get('method') == "Text Prompt":
        pdf.multi_cell(0, 5, f"User Prompt: {st.session_state.get('prompt', 'N/A')}")
    elif generation_context_info.get('method') == "Smart Schema Editor":
        pdf.multi_cell(0, 5, f"Number of Root Rows Requested: {generation_context_info.get('num_rows_root', 'N/A')}")
        default_pii_strategy = PII_HANDLING_STRATEGIES.get(st.session_state.get(DEFAULT_PII_STRATEGY_KEY, "realistic_fake"), "N/A")
        pdf.multi_cell(0, 5, f"Default PII Handling Strategy: {default_pii_strategy}")

        pdf.set_font("Arial", "B", 10)
        pdf.cell(0, 7, "Table Schemas:", 0, 1)
        pdf.set_font("Arial", size=9)
        for table_name, fields in st.session_state.get('table_schemas', {}).items():
            pdf.multi_cell(0, 5, f"  Table: {table_name}")
            for field in fields:
                pii_strat = PII_HANDLING_STRATEGIES.get(field.get('pii_handling', default_pii_strategy), "N/A")
                pdf.multi_cell(0, 5, f"    - Field: {field['name']}, Type: {FIELD_TYPES.get(field['type'], field['type'])}, Constraint: '{field['constraint']}', PII: {pii_strat}")
            pdf.ln(1)

        if st.session_state.get('relationships'):
            pdf.set_font("Arial", "B", 10)
            pdf.cell(0, 7, "Relationships:", 0, 1)
            pdf.set_font("Arial", size=9)
            for rel in st.session_state['relationships']:
                pdf.multi_cell(0, 5, f"  - {rel['parent_table']}.{rel['parent_pk']} (Parent) -> {rel['child_table']}.{rel['child_fk']} (Child)")
            pdf.ln(1)

        if st.session_state.get('edge_cases'):
            pdf.set_font("Arial", "B", 10)
            pdf.cell(0, 7, "Edge Cases Defined:", 0, 1)
            pdf.set_font("Arial", size=9)
            for i, rule in enumerate(st.session_state['edge_cases']):
                pdf.multi_cell(0, 5, f"  Rule {i+1} (Applied to ~{rule.get('percentage', 0)}% of relevant rows):")
                for cond in rule.get('conditions', []):
                    pdf.multi_cell(0, 5, f"    - If {cond.get('table', 'N/A')}.{cond.get('field', 'N/A')} {cond.get('operator', 'N/A')} '{cond.get('value', 'N/A')}'")
            pdf.ln(1)

    elif generation_context_info.get('method') == "File-based Generation":
        pdf.multi_cell(0, 5, f"Original Uploaded File: {st.session_state.get('uploaded_file_name_tab3', 'N/A')}")
        pdf.multi_cell(0, 5, f"Number of Synthetic Rows Generated: {generation_context_info.get('num_synthetic_rows_from_file', 'N/A')}")
        pdf.multi_cell(0, 5, "Synthesis Strategy: PII fields were faked. Numeric columns based on original min/max values. Other categorical/object columns sampled from original distributions.")
    pdf.ln(3)

    # --- Data Generation Process Highlights ---
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, "3. Data Generation Process Highlights", 0, 1)
    pdf.set_font("Arial", size=10)
    pdf.multi_cell(0, 5, "PII Handling: Sensitive data types (e.g., names, emails, phone numbers, IDs) are generated using realistic but fake values via the Faker library, or processed according to the specified PII handling strategy (Masked, Redacted, Scrambled).")
    pdf.multi_cell(0, 5, "Core Engine: Data generation leverages Python libraries including Faker (for realistic fake data), Pandas (for data manipulation), and NumPy/random (for numerical and choice-based generation).")

    # Add DPDP specific note
    dpdp_fields_handled_notes = []
    if generation_context_info.get('method') == "Text Prompt" and st.session_state.get('inferred_schema_from_prompt'):
        # For prompt-based, the schema is a single list of field dicts
        for field in st.session_state.get('inferred_schema_from_prompt', []):
            if is_dpdp_pii(field['name']):
                # For prompt-based, PII strategy is global initially
                strategy = PII_HANDLING_STRATEGIES.get(st.session_state.get(DEFAULT_PII_STRATEGY_KEY, "realistic_fake"), "N/A")
                dpdp_fields_handled_notes.append(f"- Field '{field['name']}' (inferred from prompt) identified as DPDP PII and handled via default '{strategy}' strategy.")
    elif generation_context_info.get('method') == "Smart Schema Editor" and st.session_state.get('table_schemas'):
        for table_name, fields in st.session_state.get('table_schemas', {}).items():
            for field in fields:
                if is_dpdp_pii(field['name']):
                    strategy = PII_HANDLING_STRATEGIES.get(field.get('pii_handling', st.session_state.get(DEFAULT_PII_STRATEGY_KEY, "realistic_fake")), "N/A")
                    dpdp_fields_handled_notes.append(f"- Field '{field['name']}' (in table '{table_name}') identified as DPDP PII and handled via '{strategy}' strategy.")
    elif generation_context_info.get('method') == "File-based Generation" and st.session_state.get('uploaded_df_for_schema') is not None:
        original_df_columns = st.session_state.uploaded_df_for_schema.columns
        for col_name in original_df_columns:
            if is_dpdp_pii(col_name):
                 dpdp_fields_handled_notes.append(f"- Field '{col_name}' (from uploaded file) identified as DPDP PII and handled by faking values during synthetic generation.")

    if dpdp_fields_handled_notes:
        pdf.set_font("Arial", "B", 9)
        pdf.multi_cell(0, 5, "DPDP Compliance Note:")
        pdf.set_font("Arial", size=9)
        for note in dpdp_fields_handled_notes:
            pdf.multi_cell(0, 5, f"  {note}")
    # Add more details here if you implement more sophisticated logging
    pdf.ln(3)

    # --- NEW: Differential Privacy Section (if applicable) ---
    if generation_context_info.get('dp_applied', False): # Add a flag to context if DP was used
        pdf.set_font("Arial", "B", 12)
        pdf.cell(0, 10, "3.5. Differential Privacy Parameters (Conceptual)", 0, 1)
        pdf.set_font("Arial", size=10)
        pdf.multi_cell(0, 5, f"Epsilon (ε) Used: {generation_context_info.get('dp_epsilon', 'N/A')}")
        pdf.multi_cell(0, 5, f"Numerical Data Mechanism: {generation_context_info.get('dp_mechanism_numeric', 'N/A')}")
        pdf.multi_cell(0, 5, f"Categorical Data Mechanism: {generation_context_info.get('dp_mechanism_categorical', 'N/A')}")
        pdf.multi_cell(0, 5, "Note: Applying Differential Privacy involves a trade-off. Lower epsilon values provide stronger privacy guarantees but may reduce the statistical utility of the synthetic data. The reported parameters are based on user selection for conceptual application.")
        pdf.ln(3)




    # --- Output Summary ---
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, "4. Output Summary", 0, 1)
    pdf.set_font("Arial", size=10)
    if generated_dfs_info:
        for table_name, info in generated_dfs_info.items():
            pdf.multi_cell(0, 5, f"Table: {table_name}, Rows: {info['rows']}, Columns: {info['cols']}")
    else:
        pdf.multi_cell(0, 5, "No dataset information available or generation was not completed for all tables.")
    pdf.ln(3)

    # --- Notes on Trust & Transparency ---
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, "5. Notes on Trust & Transparency", 0, 1)
    pdf.set_font("Arial", "I", 10)
    pdf.multi_cell(0, 5, "- The generated data is synthetic and intended for purposes like software testing, data analysis prototyping, and machine learning model training where real sensitive data cannot be used.")
    pdf.multi_cell(0, 5, "- While NullByte AI strives to create realistic and diverse datasets, synthetic data may not capture all nuances and complex correlations of real-world data.")
    pdf.multi_cell(0, 5, "- Scores from the Ethical AI Dashboard (Bias, PII Risk, Compliance) are illustrative and aim to provide guidance. Thorough validation is recommended for specific use cases.")
    pdf.multi_cell(0, 5, "- Always ensure compliance with relevant data privacy regulations (e.g., GDPR, CCPA, DPDP Act) when handling any data, including synthetic data that might resemble PII structures.")

    # Output the PDF to a byte string.
    # 'latin-1' is commonly used for FPDF string output. If you encounter encoding issues with special characters,
    # you might need to ensure your FPDF setup correctly handles UTF-8 and adjust encoding here if necessary.
    pdf_bytes = pdf.output(dest='S').encode('latin-1')
    return pdf_bytes

# ... (rest of your existing code)

# --- NEW: Reusable Edge Case UI Function ---
def render_edge_case_ui(ui_key_prefix, edge_cases_list_key, table_schemas_for_fields, available_table_names_for_conditions):
    """
    Renders the UI for defining edge cases.
    Args:
        ui_key_prefix (str): A prefix for Streamlit keys to ensure uniqueness (e.g., "tab2", "playground").
        edge_cases_list_key (str): The session state key for the list of edge cases.
        table_schemas_for_fields (dict): A dictionary of table schemas {table_name: [field_defs]} to populate field dropdowns.
        available_table_names_for_conditions (list): List of table names that can be selected in edge case conditions.
    """
    st.subheader("🧪 Edge Case Injection")
    st.markdown("Define specific scenarios to inject into a percentage of your dataset.")

    if edge_cases_list_key not in st.session_state:
        st.session_state[edge_cases_list_key] = []
    
    OPERATORS = ['==', '!=', '>', '<', '>=', '<='] # Should be defined globally or passed if not already

    edge_cases_list = st.session_state[edge_cases_list_key]

    for i, edge_rule in enumerate(edge_cases_list):
        st.markdown(f"**Edge Case Rule {i+1}**")
        rule_cols = st.columns([2, 1, 1])
        edge_rule['percentage'] = rule_cols[0].number_input(
            "Percentage of rows to affect",
            min_value=0.0, max_value=100.0,
            value=edge_rule.get('percentage', 1.0),
            step=0.1, format="%.1f",
            key=f"{ui_key_prefix}_edge_perc_{i}"
        )
        if rule_cols[1].button("🗑️ Delete Rule", key=f"{ui_key_prefix}_del_edge_rule_{i}"):
            edge_cases_list.pop(i)
            st.rerun()

        if 'conditions' not in edge_rule:
            edge_rule['conditions'] = []

        for j, condition in enumerate(edge_rule['conditions']):
            cond_cols = st.columns([2, 3, 2, 2, 1])
            
            selected_table_index = 0
            if condition.get('table') in available_table_names_for_conditions:
                selected_table_index = available_table_names_for_conditions.index(condition['table'])

            condition['table'] = cond_cols[0].selectbox(
                "Table", available_table_names_for_conditions,
                index=selected_table_index,
                key=f"{ui_key_prefix}_edge_table_{i}_{j}"
            )
            
            selected_table_for_cond = condition.get('table')
            schema_field_names_for_cond_table = []
            if selected_table_for_cond and selected_table_for_cond in table_schemas_for_fields:
                schema_field_names_for_cond_table = [f['name'] for f in table_schemas_for_fields[selected_table_for_cond] if f['name']]

            selected_field_index = 0
            if condition.get('field') in schema_field_names_for_cond_table:
                selected_field_index = schema_field_names_for_cond_table.index(condition['field'])

            condition['field'] = cond_cols[1].selectbox(
                "Field", schema_field_names_for_cond_table,
                index=selected_field_index,
                key=f"{ui_key_prefix}_edge_field_{i}_{j}"
            )
            condition['operator'] = cond_cols[2].selectbox(
                "Operator", OPERATORS,
                index=OPERATORS.index(condition['operator']) if condition.get('operator') in OPERATORS else 0,
                key=f"{ui_key_prefix}_edge_op_{i}_{j}"
            )
            condition['value'] = cond_cols[3].text_input(
                "Value", value=condition.get('value', ''),
                key=f"{ui_key_prefix}_edge_val_{i}_{j}"
            )
            if cond_cols[4].button("➖", key=f"{ui_key_prefix}_del_edge_cond_{i}_{j}"):
                edge_rule['conditions'].pop(j)
                st.rerun()

        if st.button("➕ Add Condition to Rule", key=f"{ui_key_prefix}_add_edge_cond_{i}"):
            if available_table_names_for_conditions:
                default_table_for_new_cond = available_table_names_for_conditions[0]
                default_fields_for_new_cond_table = [f['name'] for f in table_schemas_for_fields.get(default_table_for_new_cond, []) if f['name']]
                default_field_for_new_cond = default_fields_for_new_cond_table[0] if default_fields_for_new_cond_table else ""
                edge_rule['conditions'].append({'table': default_table_for_new_cond, 'field': default_field_for_new_cond, 'operator': '==', 'value': ''})
                st.rerun()
            else:
                st.warning("Please define a schema or select a template with fields before adding edge case conditions.")
        st.markdown("---")

    if st.button("➕ Add New Edge Case Rule", key=f"{ui_key_prefix}_add_new_edge_rule"):
        edge_cases_list.append({'percentage': 1.0, 'conditions': []})
        st.rerun()

def get_field_pii_strategy(field_schema, global_default_strategy):
    """Determines the PII handling strategy for a field."""
    # Field-specific strategy overrides global default
    return field_schema.get("pii_handling", global_default_strategy)

def _generate_value_from_schema(field_schema, edge_condition=None):
    """Generate a random value based on field type and constraint using a dispatch dictionary."""
    field_type = field_schema["type"]
    constraint = field_schema.get("constraint", "") # Ensure constraint exists
    # field_name = field_schema["name"] # field_name is available in field_schema
    
    default_pii_strategy = st.session_state.get(DEFAULT_PII_STRATEGY_KEY, "realistic_fake")
    current_pii_strategy = get_field_pii_strategy(field_schema, default_pii_strategy)

    # --- NEW: Check for special Faker direct calls and patterns based on flags in field_schema ---
    # These flags would have been populated if field_schema came from CANONICAL_FIELD_TO_SCHEMA_DETAILS_MAP
    # or inferred during the renaming process.
    # Note: This is a representative subset. For full functionality, all relevant 'is_faker_...'
    
    # NEW: Handle inferred alphanumeric ID for novel fields first
    if field_schema.get("_is_inferred_alphanum_id"):
        prefix = field_schema.get("_inferred_prefix", "")
        length = field_schema.get("length", random.randint(6, 10)) # Default length for inferred IDs
        return f"{prefix}{''.join(random.choices('ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789', k=length))}"

    # and 'is_..._pattern' flags from CANONICAL_FIELD_TO_SCHEMA_DETAILS_MAP should be handled here.

    if field_schema.get("is_faker_city"): return fake.city()
    if field_schema.get("is_blood_pressure_pattern"):
        systolic = random.randint(90, 170) # Example typical ranges
        diastolic = random.randint(60, 110)
        return f"{systolic}/{diastolic} mmHg"
    if field_schema.get("is_faker_state"): return fake.state()
    if field_schema.get("is_faker_country"): return fake.country()
    if field_schema.get("is_faker_postcode"): return fake.postcode()
    if field_schema.get("is_faker_currency_code"): return fake.currency_code()
    if field_schema.get("is_faker_job"): return fake.job()
    if field_schema.get("is_faker_company_with_suffix_list") and field_schema.get("suffix_from_list"):
        return f"{fake.company()} {random.choice(field_schema['suffix_from_list'])}"
    if field_schema.get("is_faker_company"): # Must be after more specific company checks
        prefix = field_schema.get('prefix', '').rstrip()
        suffix = field_schema.get('suffix', '').lstrip()
        company_name = fake.company()
        # Construct name with prefix/suffix carefully
        parts = []
        if prefix: parts.append(prefix)
        parts.append(company_name)
        if suffix: parts.append(suffix)
        return " ".join(parts)
    if field_schema.get("is_faker_url"): return fake.url()
    if field_schema.get("is_faker_ipv4"): return fake.ipv4()
    if field_schema.get("is_faker_mac_address"): return fake.mac_address()
    if field_schema.get("is_faker_user_name"): return fake.user_name()
    if field_schema.get("is_faker_latitude"): return fake.latitude()
    if field_schema.get("is_faker_longitude"): return fake.longitude()
    if field_schema.get("is_faker_color_name"): return fake.color_name()
    if field_schema.get("is_faker_credit_card_number"): return fake.credit_card_number()
    if field_schema.get("is_faker_file_name"):
        ext_list = field_schema.get("constraint") # Assuming constraint might hold extensions for file_name
        extensions = ext_list.split(',') if ext_list else FILE_EXTENSIONS_LIST # Fallback to global
        return fake.file_name(extension=random.choice(extensions))

    # Pattern-based generators
    if field_schema.get("is_generic_numeric_id"): return f"{field_schema.get('prefix', '')}{random.randint(1000, 9999)}{random.randint(1000, 9999)}"
    if field_schema.get("is_generic_alphanum_id"):
        length = field_schema.get("length", 8)
        return f"{field_schema.get('prefix', '')}{''.join(random.choices('ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789', k=length))}"
    if field_schema.get("is_room_number_pattern"): return f"{random.randint(1, 20)}{random.choice(['A', 'B', 'C', 'D'])}"
    if field_schema.get("is_version_number_pattern"): return f"{random.randint(0,5)}.{random.randint(0,20)}.{random.randint(0,100)}"
    if field_schema.get("is_doi_pattern"): return f"10.{random.randint(1000, 9999)}/{''.join(random.choices('abcdefghijklmnopqrstuvwxyz0123456789', k=random.randint(5,10)))}"
    if field_schema.get("is_tracking_number_pattern"): return f"{''.join(random.choices('ABCDEFGHIJKLMNOPQRSTUVWXYZ', k=3))}{random.randint(100000000, 999999999)}{''.join(random.choices('ABCDEFGHIJKLMNOPQRSTUVWXYZ', k=2))}"
    if field_schema.get("is_flight_number_pattern"): return f"{''.join(random.choices('ABCDEFGHIJKLMNOPQRSTUVWXYZ', k=2))}{random.randint(100, 9999)}"
    if field_schema.get("is_digit_sequence"): # e.g., constraint "digits:10-12"
        min_len, max_len = 10, 12 # defaults
        if field_schema.get('constraint', '').startswith('digits:'):
            parts = field_schema['constraint'].split(':')[1].split('-')
            min_len, max_len = int(parts[0]), int(parts[1])
        return ''.join(random.choices('0123456789', k=random.randint(min_len, max_len)))

    # Special handling for name with prefix/suffix
    if field_type == "name":
        base_name = fake.name()
        # Handle multi_name for author lists - this logic might be better placed in the main generation loop
        # if field_schema.get("is_multi_name"): # This flag is now handled in the main loop
        #     return "; ".join([fake.name() for _ in range(random.randint(1,4))])

        prefix_options = field_schema.get("prefix_options")
        prefix = field_schema.get("prefix", "")
        if prefix_options and isinstance(prefix_options, list) and prefix_options:
            prefix = random.choice(prefix_options) # Choose one from options

        if prefix and not prefix.endswith(" "):
            prefix += " "

        suffix = field_schema.get("suffix", "")
        if suffix and not suffix.startswith(" "):
            suffix = " " + suffix

        val = f"{prefix}{base_name}{suffix}".strip()
        return _apply_pii_strategy_to_value(val, "name", current_pii_strategy)
    elif field_type == "date" and constraint == "datetime": # Special handling for datetime
        # Generate a datetime object and format it
        dt_obj = fake.date_time_this_year()
        return dt_obj.strftime("%Y-%m-%d %H:%M:%S")
    elif field_type == "date" and constraint == "datetime_utc": # Special handling for datetime with UTC
        dt_obj = fake.date_time_this_year(tzinfo=timezone.utc) # Use timezone.utc
        return dt_obj.isoformat() # ISO format includes timezone

    # If type is string and not handled by specific flags/patterns above,
    # call _generate_string_value directly, passing the full_field_schema for hints.
    if field_type == "string":
        return _generate_string_value(constraint, field_schema["name"], current_pii_strategy, edge_condition=edge_condition, full_field_schema=field_schema)

    generator_func = VALUE_GENERATOR_FUNCTIONS.get(field_type)
    if generator_func:
        # Pass field_name for context within generator functions
        return generator_func(constraint, field_schema["name"], current_pii_strategy, edge_condition=edge_condition)

    st.warning(f"Unknown field type '{field_type}' for field '{field_schema['name']}'. Defaulting to N/A.")
    return "N/A" # Fallback for unknown types

def generate_value(field_name_str: str):
    """
    Generates a single synthetic value for a given field name string,
    intelligently inferring type and context for business, tech, or industrial use cases.

    Args:
        field_name_str (str): The name of the field to generate a value for.

    Returns:
        Any: A single synthetic value appropriate for the field name.

    Example Usage:
        # Business/Finance Examples:
        generate_value("subscription_price")  # Returns: 49.99 (float in USD)
        generate_value("monthly_recurring_revenue")  # Returns: 12500.75 (float in USD)
        generate_value("customer_lifetime_value")  # Returns: 2500.50 (float in USD)
        generate_value("churn_rate")  # Returns: 0.15 (float as percentage)
        generate_value("conversion_rate")  # Returns: 0.023 (float as percentage)
        
        # Tech/SaaS Examples:
        generate_value("api_endpoint")  # Returns: /v1/users/get (string)
        generate_value("database_connection_string")  # Returns: postgresql://user:pass@host:5432/db (string)
        generate_value("software_version")  # Returns: 2.1.3 (string)
        generate_value("deployment_environment")  # Returns: production (category)
        generate_value("server_status")  # Returns: running (category)
        
        # E-commerce Examples:
        generate_value("product_sku")  # Returns: SKU-A42X-B7 (string)
        generate_value("order_status")  # Returns: shipped (category)
        generate_value("shipping_cost")  # Returns: 12.99 (float in USD)
        generate_value("discount_percentage")  # Returns: 15.0 (float as percentage)
        
        # Analytics Examples:
        generate_value("daily_active_users")  # Returns: 1250 (integer)
        generate_value("bounce_rate")  # Returns: 0.45 (float as percentage)
        generate_value("average_session_duration")  # Returns: 180 (integer in seconds)
        generate_value("page_load_time")  # Returns: 1.2 (float in seconds)
        
        # HR/Employee Examples:
        generate_value("employee_id")  # Returns: EMP-7321 (string)
        generate_value("department")  # Returns: Engineering (category)
        generate_value("job_title")  # Returns: Senior Software Engineer (string)
        generate_value("salary_range")  # Returns: 80000-120000 (string)
        
        # Marketing Examples:
        generate_value("campaign_id")  # Returns: CAMP-2024-Q1 (string)
        generate_value("lead_source")  # Returns: Website (category)
        generate_value("email_open_rate")  # Returns: 0.35 (float as percentage)
        generate_value("click_through_rate")  # Returns: 0.12 (float as percentage)
    """
    # First, try to infer the schema from the field name
    inferred_schema = infer_field_schema_from_name(field_name_str)
    
    # Add business-specific enhancements to the schema
    field_lower = field_name_str.lower()
    
    # Business/Finance specific enhancements
    if any(kw in field_lower for kw in ["revenue", "income", "price", "cost", "amount", "value"]):
        if "rate" in field_lower or "percentage" in field_lower:
            inferred_schema["type"] = "float"
            inferred_schema["constraint"] = "0.0-1.0"  # Represent as decimal
        else:
            inferred_schema["type"] = "float"
            inferred_schema["constraint"] = "10.0-10000.0"  # Business-appropriate range
    
    # Tech/SaaS specific enhancements
    elif any(kw in field_lower for kw in ["api", "endpoint", "url", "path"]):
        inferred_schema["type"] = "string"
        inferred_schema["_hint_api_path"] = True
    elif any(kw in field_lower for kw in ["version", "release"]):
        inferred_schema["type"] = "string"
        inferred_schema["_hint_version"] = True
    elif any(kw in field_lower for kw in ["environment", "env"]):
        inferred_schema["type"] = "category"
        inferred_schema["constraint"] = "development,staging,production"
    
    # Analytics specific enhancements
    elif any(kw in field_lower for kw in ["rate", "percentage", "ratio"]):
        inferred_schema["type"] = "float"
        inferred_schema["constraint"] = "0.0-1.0"
    elif any(kw in field_lower for kw in ["duration", "time"]):
        if "millisecond" in field_lower:
            inferred_schema["type"] = "float"
            inferred_schema["constraint"] = "0.0-1000.0"
        elif "second" in field_lower:
            inferred_schema["type"] = "float"
            inferred_schema["constraint"] = "0.0-60.0"
        else:
            inferred_schema["type"] = "int"
            inferred_schema["constraint"] = "0-3600"
    
    # Generate the value based on the enhanced schema
    return _generate_value_from_schema(inferred_schema)

# --- Helper for Dependency Handling ---
def _get_dependent_value(current_row_data, dependent_field_name_lower): # Already imported, ensure it's available
    """Retrieves a value from current_row_data based on a lowercase dependent field name."""
    key = next((k for k in current_row_data if k.lower() == dependent_field_name_lower), None)
    return current_row_data.get(key) if key else None

def _handle_discharge_date_dependency(field_schema, current_row_data, edge_condition):
    admission_date_str = _get_dependent_value(current_row_data, "admission date")
    if admission_date_str and isinstance(admission_date_str, str):
        try:
            admission_date = datetime.strptime(admission_date_str, "%Y-%m-%d")
            if edge_condition and edge_condition.get('operator') == '==':
                return str(edge_condition['value'])
            discharge_date = admission_date + timedelta(days=random.randint(1, 30))
            return discharge_date.strftime("%Y-%m-%d")
        except ValueError:
            st.warning(f"Admission Date '{admission_date_str}' for '{field_schema['name']}' dependency is not a valid date. Generating independently.")
    else:
        st.warning(f"Admission Date not found or invalid for '{field_schema['name']}' dependency. Generating independently.")
    return None # Indicates dependency could not be resolved, fall back

def _handle_city_dependency(field_schema, current_row_data, edge_condition):
    state_val = _get_dependent_value(current_row_data, "state")
    cities_for_state = STATE_CITY_MAP.get(state_val)
    if cities_for_state:
        if edge_condition and edge_condition.get('operator') == '==':
            edge_val = str(edge_condition['value'])
            return edge_val if edge_val in cities_for_state else random.choice(cities_for_state)
        return random.choice(cities_for_state)
    return None # Fallback

def _handle_currency_dependency(field_schema, current_row_data, edge_condition):
    country_val = _get_dependent_value(current_row_data, "country")
    currency_code = COUNTRY_CURRENCY_MAP.get(country_val)
    if currency_code:
        if edge_condition and edge_condition.get('operator') == '==':
            return str(edge_condition['value'])
        return currency_code
    return None # Fallback

# --- Dispatch Dictionary for Dependency Handlers ---
# Maps (field_to_generate_lower, depends_on_field_lower) to handler function
DEPENDENCY_HANDLERS = {
    ("discharge date", "admission date"): _handle_discharge_date_dependency, # Already imported, ensure it's available
    ("city", "state"): _handle_city_dependency,
    ("currency", "country"): _handle_currency_dependency,
    # Add more dependency handlers here as needed
    # Example: ("dependent field name lower", "primary field name lower"): _handle_custom_dependency
}


def generate_value_with_dependencies(field_schema, current_row_data, edge_condition=None):
    """
    Generate a value for a field, considering dependencies on other fields in the current_row_data.
    Exclusively used by the Smart Schema Editor.
    """
    field_type = field_schema["type"]
    constraint = field_schema["constraint"]
    field_name = field_schema["name"] # Original case field name
    field_name_lower = field_name.lower() # Lowercase for matching

    default_pii_strategy = st.session_state.get(DEFAULT_PII_STRATEGY_KEY, "realistic_fake")
    current_pii_strategy = get_field_pii_strategy(field_schema, default_pii_strategy)

    # --- NEW: Dependency Handling using Dispatch Dictionary ---
    for (fn_to_gen_lower, dep_fn_lower), handler_func in DEPENDENCY_HANDLERS.items():
        if field_name_lower == fn_to_gen_lower:
            # Check if the dependent field actually exists in the current row data
            if any(k.lower() == dep_fn_lower for k in current_row_data.keys()):
                generated_value = handler_func(field_schema, current_row_data, edge_condition)
                if generated_value is not None: # Handler successfully generated a value
                    return generated_value
                # If handler returned None, it means it couldn't satisfy the dependency (e.g., invalid dependent value)
                # We will then fall through to standard/non-dependent generation for this field.
            break # Found the field to generate, no need to check other dependency rules for it

    # --- Fallback to standard generation if no specific dependency handled ---
    # Special handling for name with prefix/suffix (needs field_schema)
    if field_type == "name":
        base_name = fake.name()
        prefix_options = field_schema.get("prefix_options")
        prefix = field_schema.get("prefix", "")
        if prefix_options and isinstance(prefix_options, list) and prefix_options:
            prefix = random.choice(prefix_options)
        if prefix and not prefix.endswith(" "): prefix += " "
        suffix = field_schema.get("suffix", "")
        if suffix and not suffix.startswith(" "): suffix = " " + suffix
        val = f"{prefix}{base_name}{suffix}".strip()
        return _apply_pii_strategy_to_value(val, "name", current_pii_strategy)
    elif field_type == "date" and constraint == "datetime":
        dt_obj = fake.date_time_this_year()
        return dt_obj.strftime("%Y-%m-%d %H:%M:%S")
    elif field_type == "date" and constraint == "datetime_utc":
        dt_obj = fake.date_time_this_year(tzinfo=timezone.utc)
        return dt_obj.isoformat()

    # If type is string and not handled by specific flags/patterns above,
    # call _generate_string_value directly, passing the full_field_schema for hints.
    if field_type == "string":
        return _generate_string_value(constraint, field_name, current_pii_strategy, edge_condition=edge_condition, full_field_schema=field_schema)

    generator_func = VALUE_GENERATOR_FUNCTIONS.get(field_type)
    if generator_func:
        return generator_func(constraint, field_name, current_pii_strategy, edge_condition=edge_condition)

    st.warning(f"Unknown field type '{field_type}' for field '{field_name}' in dependency generator. Defaulting to N/A.")
    return "N/A"

def generate_value_for_field(field_name_str):
    """
    Generates a single synthetic value for a given field name string,
    even if the field name is random or never seen before.
    """
    inferred_schema = infer_field_schema_from_name(field_name_str)
    return _generate_value_from_schema(inferred_schema)


def get_generation_order(table_schemas, relationships):
    """Determines the order to generate tables based on dependencies (PK/FK)."""
    from collections import defaultdict, deque

    graph = defaultdict(list)
    in_degree = defaultdict(int)
    all_tables = list(table_schemas.keys())

    for table in all_tables:
        in_degree[table] = 0 # Initialize in_degree for all tables

    for rel in relationships:
        parent = rel['parent_table']
        child = rel['child_table']
        if parent in all_tables and child in all_tables:
            graph[parent].append(child)
            in_degree[child] += 1
        else:
            st.warning(f"Relationship {parent}->{child} involves undefined tables. Skipping this relationship for ordering.")

    queue = deque([table for table in all_tables if in_degree[table] == 0])
    generation_order = []

    while queue:
        table = queue.popleft()
        generation_order.append(table)
        for neighbor in graph[table]:
            in_degree[neighbor] -= 1
            if in_degree[neighbor] == 0:
                queue.append(neighbor)

    if len(generation_order) != len(all_tables):
        st.error("Cycle detected in table relationships or some tables are unreachable. Cannot determine generation order.")
        # Identify missing tables for better error message
        missing_tables = set(all_tables) - set(generation_order)
        st.error(f"Problematic tables might include: {', '.join(missing_tables)}")
        return None

    return generation_order

# --- NEW: Simplified Generation for Single Table Scenario Playground ---
def generate_single_table_data_with_edge_cases(schema_fields, num_rows, edge_cases_list, pii_strategy_global, table_name_for_conditions):
    """Generates data for a single table, applying edge cases."""
    if not schema_fields:
        st.error("Schema is empty. Cannot generate data for the playground.")
        return None

    table_rows_data = []
    for i in range(num_rows):
        row_data = {} # For potential intra-row dependencies
        applied_edge_rule_for_row = None
        
        # Determine if an edge case rule applies to this row
        potential_rules_for_row = [
            rule for rule in edge_cases_list
            if rule.get('percentage', 0.0) > 0 and random.random() < (rule.get('percentage', 0.0) / 100.0)
        ]
        if potential_rules_for_row:
            applied_edge_rule_for_row = random.choice(potential_rules_for_row)

        for field_schema in schema_fields:
            field_name = field_schema["name"]
            field_specific_edge_condition = None
            if applied_edge_rule_for_row:
                for cond in applied_edge_rule_for_row.get('conditions', []):
                    # For playground, table name in condition must match the fixed playground table name
                    if cond.get('table') == table_name_for_conditions and cond.get('field') == field_name:
                        field_specific_edge_condition = cond
                        break
            # Use generate_value_with_dependencies, passing field_schema, current row_data, and edge_condition
            row_data[field_name] = generate_value_with_dependencies(field_schema, row_data, edge_condition=field_specific_edge_condition)
        
        table_rows_data.append(row_data)

    df = pd.DataFrame(table_rows_data)
    
    # Apply PII Scrambling if needed
    for field_s in schema_fields:
        is_sensitive = field_s["type"] in ["email", "phone", "aadhaar", "pan", "passport", "voterid", "ifsc", "upi", "name", "address"]
        if is_sensitive and field_s.get("pii_handling") == "scramble_column" and field_s["name"] in df.columns:
            col_to_scramble = df[field_s["name"]].copy()
            if col_to_scramble.nunique() > 1: # Only scramble if there's more than one unique value
                np.random.shuffle(col_to_scramble.values)
                df[field_s["name"]] = col_to_scramble
    return df

def generate_hierarchical_data(table_schemas, relationships, num_rows_root, edge_cases_all, pii_strategy_global):
    """Generates data for multiple related tables."""
    generation_order = get_generation_order(table_schemas, relationships)
    if not generation_order:
        return None # Error already shown by get_generation_order

    generated_data_frames = {}
    min_children_per_parent = 1 # Configurable: min number of child records per parent
    max_children_per_parent = 3 # Configurable: max number of child records per parent

    for table_name in generation_order:
        current_schema_fields = table_schemas[table_name]
        table_rows_data = []
        
        parent_relationships_for_this_table = [r for r in relationships if r['child_table'] == table_name]
        num_rows_for_this_table = 0
        parent_pk_map_for_rows = [] # Stores [{fk_field: parent_pk_value}, ...] for each row

        if not parent_relationships_for_this_table: # It's a root table
            num_rows_for_this_table = num_rows_root
            for _ in range(num_rows_for_this_table):
                parent_pk_map_for_rows.append({})
        else:
            # Simplified: Iterate through the first parent's rows and generate children
            primary_parent_rel = parent_relationships_for_this_table[0]
            parent_df = generated_data_frames.get(primary_parent_rel['parent_table'])
            if parent_df is None or parent_df.empty:
                st.error(f"Parent table '{primary_parent_rel['parent_table']}' for '{table_name}' has no data. Cannot generate child rows.")
                continue

            for _, parent_row in parent_df.iterrows():
                num_children = random.randint(min_children_per_parent, max_children_per_parent)
                for _ in range(num_children):
                    fk_map = {}
                    for rel in parent_relationships_for_this_table:
                        current_parent_df = generated_data_frames.get(rel['parent_table'])
                        if current_parent_df is not None and rel['parent_pk'] in current_parent_df.columns:
                            # If this rel is the primary_parent_rel, use parent_row directly
                            # Otherwise, for secondary parents, pick a random PK (simplification for now)
                            pk_val_to_use = parent_row[rel['parent_pk']] if rel == primary_parent_rel else random.choice(current_parent_df[rel['parent_pk']].tolist())
                            fk_map[rel['child_fk']] = pk_val_to_use
                        else:
                            st.warning(f"Could not find PK '{rel['parent_pk']}' in generated parent table '{rel['parent_table']}' for FK '{rel['child_fk']}' in '{table_name}'. FK will be None.")
                            fk_map[rel['child_fk']] = None 
                    parent_pk_map_for_rows.append(fk_map)
            num_rows_for_this_table = len(parent_pk_map_for_rows)

        for i in range(num_rows_for_this_table):
            row_data = {}
            current_fk_map = parent_pk_map_for_rows[i]
            for fk_field, pk_value in current_fk_map.items():
                row_data[fk_field] = pk_value

            applied_edge_rule_for_row = None
            potential_rules_for_row = [
                rule for rule in edge_cases_all 
                if rule.get('percentage', 0.0) > 0 and random.random() < (rule.get('percentage', 0.0) / 100.0)
            ]
            if potential_rules_for_row:
                applied_edge_rule_for_row = random.choice(potential_rules_for_row)

            for field_schema in current_schema_fields:
                field_name = field_schema["name"]
                if field_name in row_data: continue # Skip FKs already set
                
                field_specific_edge_condition = None
                if applied_edge_rule_for_row:
                    for cond in applied_edge_rule_for_row.get('conditions', []):
                        if cond.get('table') == table_name and cond.get('field') == field_name:
                            field_specific_edge_condition = cond
                            break
                row_data[field_name] = generate_value_with_dependencies(field_schema, row_data, edge_condition=field_specific_edge_condition)
            
            table_rows_data.append(row_data)

        df = pd.DataFrame(table_rows_data)
        
        for field_s in current_schema_fields: # PII Scrambling
            is_sensitive = field_s["type"] in ["email", "phone", "aadhaar", "pan", "passport", "voterid", "ifsc", "upi", "name", "address"]
            if is_sensitive and field_s.get("pii_handling") == "scramble_column" and field_s["name"] in df.columns:
                col_to_scramble = df[field_s["name"]].copy()
                if col_to_scramble.nunique() > 1:
                    np.random.shuffle(col_to_scramble.values)
                    df[field_s["name"]] = col_to_scramble

        generated_data_frames[table_name] = df
        if df.empty and num_rows_for_this_table > 0 :
            st.warning(f"Generated empty DataFrame for table '{table_name}' despite expecting {num_rows_for_this_table} rows.")
        elif not df.empty:
            detected_dpdp_for_table = [f_schema["name"] for f_schema in current_schema_fields if is_dpdp_pii(f_schema["name"])]
            if detected_dpdp_for_table:
                st.info(f"DPDP sensitive fields in '{table_name}': {', '.join(detected_dpdp_for_table)}")

    return generated_data_frames

# The generate_domain_specific_data function is now effectively merged into generate_synthetic_data

# --- Schema Inference Rules for pre-populating Smart Schema Editor ---
SCHEMA_INFERENCE_RULES = [
    (r"(?:email)", "email", ""),
    (r"(?:name)", "name", ""),
    (r"(?:phone|contact)", "phone", ""),
    (r"(?:address|location)", "address", ""),
    (r"(?:age)", "int", "18-60"),
    (r"(?:salary|income|pay)", "int", "20000-500000"),
    (r"(?:gender)", "category", "Male, Female, Other"),
    (r"(?:aadhaar)", "aadhaar", ""),
    (r"(?:pan)", "pan", ""),
    (r"(?:passport)", "passport", ""),
    (r"(?:voter)", "voterid", ""), # voter or voter id
    (r"(?:ifsc)", "ifsc", ""),
    (r"(?:upi)", "upi", ""),
    (r"(?:date|dob|joining_date|order_date)", "date", f"{fake.date_this_year(before_today=True, after_today=False).strftime('%Y-%m-%d')} - {fake.date_this_year(before_today=False, after_today=True).strftime('%Y-%m-%d')}"),
    (r"(?:price|amount|value|cost)", "float", "10.00-1000.00"),
    (r"(?:id|number|code)$", "string", ""), # Ends with id, number, or code
    (r"(?:status|type|category)$", "category", "Type A, Type B, Type C"), # Ends with status, type, or category
]

# --- Schema Templates ---
SCHEMA_TEMPLATES = {
    "None (Custom Schema)": { # Special value to indicate no template or custom editing
        "description": "Start with a blank schema in the Smart Schema Editor.",
        "fields": []
    },
    "E-commerce Customer Orders": {
        "description": "A standard schema for tracking customer orders, including customer details, product information, and order status.",
        "fields": [
            {"name": "Order ID", "type": "string", "constraint": "", "_canonical_suggestion": "order_id"},
            {"name": "Customer Name", "type": "name", "constraint": ""},
            {"name": "Customer Email", "type": "email", "constraint": ""},
            {"name": "Product Name", "type": "category", "constraint": "Laptop, Smartphone, Headphones, Charger, Case"},
            {"name": "Quantity", "type": "int", "constraint": "1-5"},
            {"name": "Price Per Unit", "type": "float", "constraint": "10.99-1299.99"},
            {"name": "Order Date", "type": "date", "constraint": ""},
            {"name": "Shipping Address", "type": "address", "constraint": ""},
            {"name": "Order Status", "type": "category", "constraint": "Pending, Shipped, Delivered, Cancelled, Returned"},
            {"name": "Payment Method", "type": "category", "constraint": "Credit Card, Debit Card, UPI, Net Banking"},
        ]
    },
    "Basic Employee Records": {
        "description": "Essential fields for managing employee information, such as ID, contact details, department, and salary.",
        "fields": [
            {"name": "Employee ID", "type": "string", "constraint": "", "_canonical_suggestion": "employee_id"},
            {"name": "Full Name", "type": "name", "constraint": ""},
            {"name": "Email", "type": "email", "constraint": ""},
            {"name": "Phone Number", "type": "phone", "constraint": ""},
            {"name": "Department", "type": "category", "constraint": "HR, Engineering, Marketing, Sales, Finance, Operations"},
            {"name": "Position", "type": "string", "constraint": ""},
            {"name": "Salary", "type": "int", "constraint": "30000-250000"},
            {"name": "Joining Date", "type": "date", "constraint": ""},
            {"name": "Age", "type": "int", "constraint": "22-60"},
        ]
    },
    "Healthcare Patient Data (Demo)": {
        "description": "A demonstration schema for patient records in a healthcare setting, including admission details and diagnosis.",
        "fields": [
            {"name": "Patient ID", "type": "string", "constraint": "", "_canonical_suggestion": "patient_id"},
            {"name": "Patient Name", "type": "name", "constraint": ""},
            {"name": "Age", "type": "int", "constraint": "0-90"},
            {"name": "Gender", "type": "category", "constraint": "Male, Female, Other"},
            {"name": "Admission Date", "type": "date", "constraint": ""},
            {"name": "Discharge Date", "type": "date", "constraint": ""},
            {"name": "Diagnosis", "type": "category", "constraint": "Flu, Common Cold, Hypertension, Diabetes, Injury"},
            {"name": "Attending Doctor", "type": "name", "constraint": "", "_canonical_suggestion": "doctor_name"},
            {"name": "Contact Phone", "type": "phone", "constraint": ""},
        ]
    },
    "Financial Transactions (Simplified)": {
        "description": "A simplified schema for financial transactions, covering transaction ID, account details, amount, and type.",
        "fields": [
            {"name": "Transaction ID", "type": "string", "constraint": "", "_canonical_suggestion": "transaction_id"},
            {"name": "Account Number", "type": "string", "constraint": ""},
            {"name": "Transaction Date", "type": "date", "constraint": ""},
            {"name": "Amount", "type": "float", "constraint": "1.00-50000.00"},
            {"name": "Transaction Type", "type": "category", "constraint": "Credit, Debit, Transfer, Payment"},
            {"name": "Description", "type": "string", "constraint": ""},
        ]
    },
    "Social Media Posts": {
        "description": "Schema for social media post data, including user information, post content, engagement metrics, and timestamps.",
        "fields": [
            {"name": "Post ID", "type": "string", "constraint": "", "_canonical_suggestion": "id"},
            {"name": "Username", "type": "string", "constraint": "", "_canonical_suggestion": "username"},
            {"name": "Post Text", "type": "string", "constraint": "", "_canonical_suggestion": "post_text"},
            {"name": "Timestamp", "type": "date", "constraint": "datetime", "_canonical_suggestion": "timestamp"},
            {"name": "Likes", "type": "int", "constraint": "0-10000", "_canonical_suggestion": "like_count"},
            {"name": "Shares", "type": "int", "constraint": "0-5000", "_canonical_suggestion": "share_count"},
            {"name": "Hashtags", "type": "string", "constraint": "", "_canonical_suggestion": "hashtags"}
        ]
    },
    "IoT Sensor Readings": {
        "description": "Data structure for readings from IoT sensors, including sensor ID, timestamp, and various environmental measurements.",
        "fields": [
            {"name": "Sensor ID", "type": "string", "constraint": "", "_canonical_suggestion": "sensor_id"},
            {"name": "Timestamp", "type": "date", "constraint": "datetime", "_canonical_suggestion": "timestamp"},
            {"name": "Temperature (°C)", "type": "float", "constraint": "-10.0-40.0", "_canonical_suggestion": "temperature"},
            {"name": "Humidity (%)", "type": "float", "constraint": "20.0-80.0", "_canonical_suggestion": "humidity"},
            {"name": "Latitude", "type": "float", "constraint": "-90.0-90.0", "_canonical_suggestion": "latitude"},
            {"name": "Longitude", "type": "float", "constraint": "-180.0-180.0", "_canonical_suggestion": "longitude"},
            {"name": "Sensor Type", "type": "category", "constraint": "Temperature,Humidity,Pressure,Light", "_canonical_suggestion": "sensor_type"}
        ]
    },
    "Academic Publications": {
        "description": "Schema for academic publication metadata, including title, authors, journal, publication year, and citation information.",
        "fields": [
            {"name": "Publication ID", "type": "string", "constraint": "", "_canonical_suggestion": "id"},
            {"name": "Title", "type": "string", "constraint": "", "_canonical_suggestion": "publication_title"},
            {"name": "Authors", "type": "string", "constraint": "", "_canonical_suggestion": "author_names"},
            {"name": "Journal", "type": "category", "constraint": "Nature, Science, Cell, The Lancet, PLOS One", "_canonical_suggestion": "journal_name"},
            {"name": "Publication Year", "type": "int", "constraint": "2000-2024", "_canonical_suggestion": "publication_year"},
            {"name": "DOI", "type": "string", "constraint": "", "_canonical_suggestion": "doi"},
            {"name": "Keywords", "type": "string", "constraint": "", "_canonical_suggestion": "keywords"},
            {"name": "Citation Count", "type": "int", "constraint": "0-500", "_canonical_suggestion": "citation_count"}
        ]
    },
    "Logistics Shipment Tracking": {
        "description": "Fields for tracking logistics shipments, including shipment ID, carrier, origin, destination, status, and costs.",
        "fields": [
            {"name": "Shipment ID", "type": "string", "constraint": "", "_canonical_suggestion": "shipment_id"},
            {"name": "Tracking Number", "type": "string", "constraint": "", "_canonical_suggestion": "tracking_number"},
            {"name": "Carrier Name", "type": "string", "constraint": "", "_canonical_suggestion": "carrier_name"},
            {"name": "Origin", "type": "address", "constraint": "", "_canonical_suggestion": "origin_location"},
            {"name": "Destination", "type": "address", "constraint": "", "_canonical_suggestion": "destination_location"},
            {"name": "Status", "type": "category", "constraint": ",".join(LOGISTICS_SHIPMENT_STATUSES_LIST), "_canonical_suggestion": "shipment_status_logistics"},
            {"name": "Estimated Delivery", "type": "date", "constraint": "", "_canonical_suggestion": "estimated_delivery_date"},
            {"name": "Actual Delivery", "type": "date", "constraint": "", "_canonical_suggestion": "actual_delivery_date"},
            {"name": "Freight Cost (USD)", "type": "float", "constraint": "50-5000", "_canonical_suggestion": "freight_cost"},
            {"name": "Package Weight (kg)", "type": "float", "constraint": "0.1-1000", "_canonical_suggestion": "package_weight_kg"},
            {"name": "Package Dimensions (cm)", "type": "string", "constraint": "", "_canonical_suggestion": "package_dimensions_cm"},
        ]
    },
    "Travel Booking Records": {
        "description": "Schema for travel booking information, covering booking ID, traveler details, destination, flight, hotel, and costs.",
        "fields": [
            {"name": "Booking ID", "type": "string", "constraint": "", "_canonical_suggestion": "booking_id"},
            {"name": "Traveler Name", "type": "name", "constraint": "", "_canonical_suggestion": "traveler_name"},
            {"name": "Destination", "type": "category", "constraint": ",".join(TRAVEL_DESTINATIONS_LIST[:10]), "_canonical_suggestion": "destination_city_travel"},
            {"name": "Origin", "type": "category", "constraint": ",".join(TRAVEL_DESTINATIONS_LIST[10:20]), "_canonical_suggestion": "origin_city_travel"},
            {"name": "Travel Date", "type": "date", "constraint": "", "_canonical_suggestion": "travel_date"},
            {"name": "Return Date", "type": "date", "constraint": "", "_canonical_suggestion": "return_date"},
            {"name": "Flight Number", "type": "string", "constraint": "", "_canonical_suggestion": "flight_number"},
            {"name": "Airline", "type": "category", "constraint": ",".join(AIRLINE_NAMES_LIST[:5]), "_canonical_suggestion": "airline_name"},
            {"name": "Hotel Name", "type": "string", "constraint": "", "_canonical_suggestion": "hotel_name"},
            {"name": "Room Type", "type": "category", "constraint": ",".join(ROOM_TYPES_LIST[:4]), "_canonical_suggestion": "room_type"},
            {"name": "Booking Status", "type": "category", "constraint": ",".join(TRAVEL_BOOKING_STATUSES_LIST), "_canonical_suggestion": "booking_status_travel"},
            {"name": "Total Cost (USD)", "type": "float", "constraint": "200-8000", "_canonical_suggestion": "total_travel_cost"},
            {"name": "Package Name", "type": "string", "constraint": "", "_canonical_suggestion": "travel_package_name"},
            {"name": "Booked Activity", "type": "category", "constraint": ",".join(TRAVEL_ACTIVITY_TYPES_LIST[:5]), "_canonical_suggestion": "travel_activity"},
        ]
    },
    # --- NEW TEMPLATES ---
    "Sales CRM Leads": {
        "description": "Schema for tracking sales leads in a CRM system, including lead source, deal stage, and opportunity details.",
        "fields": [
            {"name": "Lead ID", "type": "string", "constraint": "", "_canonical_suggestion": "lead_id"},
            {"name": "Lead Name", "type": "name", "constraint": "", "_canonical_suggestion": "name"}, # Generic name
            {"name": "Company Name", "type": "string", "constraint": "", "_canonical_suggestion": "company"},
            {"name": "Email", "type": "email", "constraint": ""},
            {"name": "Phone", "type": "phone", "constraint": ""},
            {"name": "Lead Source", "type": "category", "constraint": ",".join(LEAD_SOURCES_LIST), "_canonical_suggestion": "lead_source"},
            {"name": "Deal Stage", "type": "category", "constraint": ",".join(DEAL_STAGES_LIST), "_canonical_suggestion": "deal_stage"},
            {"name": "Opportunity Amount", "type": "float", "constraint": "5000-500000", "_canonical_suggestion": "opportunity_amount"},
            {"name": "Expected Close Date", "type": "date", "constraint": "", "_canonical_suggestion": "close_date"},
            {"name": "Assigned To", "type": "name", "constraint": "", "_canonical_suggestion": "assignee_name"},
        ]
    },
    "Marketing Campaign Performance": {
        "description": "Data for analyzing marketing campaign performance, including ad spend, impressions, clicks, and conversion metrics.",
        "fields": [
            {"name": "Campaign ID", "type": "string", "constraint": "", "_canonical_suggestion": "campaign_id"},
            {"name": "Campaign Name", "type": "string", "constraint": "", "_canonical_suggestion": "campaign_name"},
            {"name": "Campaign Type", "type": "category", "constraint": ",".join(CAMPAIGN_TYPES_LIST), "_canonical_suggestion": "campaign_type"},
            {"name": "Start Date", "type": "date", "constraint": "", "_canonical_suggestion": "date"},
            {"name": "End Date", "type": "date", "constraint": "", "_canonical_suggestion": "date"},
            {"name": "Ad Spend (USD)", "type": "float", "constraint": "100-20000", "_canonical_suggestion": "ad_spend"},
            {"name": "Impressions", "type": "int", "constraint": "10000-5000000", "_canonical_suggestion": "impressions"},
            {"name": "Clicks", "type": "int", "constraint": "100-50000", "_canonical_suggestion": "clicks"},
            {"name": "CTR", "type": "float", "constraint": "0.005-0.05", "_canonical_suggestion": "ctr"},
            {"name": "Conversions", "type": "int", "constraint": "10-1000"}, # No direct canonical, simple int
            {"name": "Conversion Rate", "type": "float", "constraint": "0.01-0.15", "_canonical_suggestion": "conversion_rate"},
        ]
    },
    "Software Issue Tracking": {
        "description": "Schema for tracking software bugs, feature requests, and tasks, including status, priority, and assignee.",
        "fields": [
            {"name": "Issue ID", "type": "string", "constraint": "", "_canonical_suggestion": "id"}, # Generic ID
            {"name": "Title", "type": "string", "constraint": "", "_canonical_suggestion": "description"}, # Using description for title
            {"name": "Issue Type", "type": "category", "constraint": ",".join(ISSUE_TYPES_LIST), "_canonical_suggestion": "issue_type"},
            {"name": "Status", "type": "category", "constraint": ",".join(ISSUE_STATUSES_LIST), "_canonical_suggestion": "issue_status"},
            {"name": "Priority", "type": "category", "constraint": ",".join(TASK_PRIORITY_LIST), "_canonical_suggestion": "task_priority"},
            {"name": "Reporter", "type": "name", "constraint": "", "_canonical_suggestion": "name"},
            {"name": "Assignee", "type": "name", "constraint": "", "_canonical_suggestion": "assignee_name"},
            {"name": "Created Date", "type": "date", "constraint": "datetime", "_canonical_suggestion": "timestamp"},
            {"name": "Updated Date", "type": "date", "constraint": "datetime", "_canonical_suggestion": "timestamp"},
            {"name": "Commit Hash", "type": "string", "constraint": "", "_canonical_suggestion": "commit_hash"},
        ]
    }
}

# Smart Schema Editor Section
def show_smart_schema_editor(synthetic_df=None, num_rows=10):
    st.subheader("🧠 Smart Schema Editor")
    st.markdown("Define and customize your data schema with field types and constraints.")

    # --- Logic to load schema from Text Prompt if requested ---
    if st.session_state.get('load_prompt_schema_to_editor'):
        inferred_schema = st.session_state.get('inferred_schema_from_prompt')
        num_rows_prompt_val = st.session_state.get('num_rows_from_prompt', 10)

        if inferred_schema:
            target_table_name = "PromptInferredTable" # Name for the table loaded from prompt
            
            # Ensure this table exists, or create it
            if target_table_name not in st.session_state.table_schemas:
                st.session_state.table_schemas[target_table_name] = []
            
            st.session_state.table_schemas[target_table_name] = inferred_schema # Load the schema
            st.session_state.active_table_name = target_table_name # Make it active
            st.session_state.num_rows_smart_schema_editor = num_rows_prompt_val # Pre-fill row count
            
            st.session_state.initial_schema_populated = True # Mark as populated
            st.session_state.selected_template_name = "None (Custom Schema)" # Reset template selection

            # Clear flags and temporary data
            st.session_state.load_prompt_schema_to_editor = False
            st.session_state.inferred_schema_from_prompt = None
            st.session_state.num_rows_from_prompt = None
            
            # Clear other generated data to avoid confusion and focus on the editor
            st.session_state.prompt_generated_df = None 
            st.session_state.uploaded_df_for_schema = None
            st.session_state.generated_data_frames = {}
            st.session_state.active_display_table_name = None
            st.session_state.playground_generated_df = None
            st.session_state.newly_generated_df_tab3 = None
            st.session_state.synthetic_df_from_file_tab3 = None
            
            st.success(f"Schema from text prompt loaded into table '{target_table_name}'. You can now edit it.")
            st.rerun()
        else:
            # If flag is true but no schema, just reset flag to avoid loop
            st.session_state.load_prompt_schema_to_editor = False
            st.warning("No schema data found from prompt to load.")

    # --- Table Management ---
    st.markdown("---")
    st.subheader("Tables & Relationships")
    
    if not st.session_state.table_schemas: # Ensure at least one table
        st.session_state.table_schemas["Table1"] = []
        st.session_state.active_table_name = "Table1"
        st.session_state.initial_schema_populated = False

    if st.session_state.active_table_name is None and st.session_state.table_schemas:
        st.session_state.active_table_name = list(st.session_state.table_schemas.keys())[0]

    table_management_cols = st.columns([2,1,1])
    with table_management_cols[0]:
        new_table_name = st.text_input("New Table Name", key="new_table_name_input")
    with table_management_cols[1]:
        if st.button("➕ Add Table", key="add_table_button", use_container_width=True):
            if new_table_name and new_table_name not in st.session_state.table_schemas:
                st.session_state.table_schemas[new_table_name] = []
                st.session_state.active_table_name = new_table_name
                st.session_state.initial_schema_populated = False
                st.rerun()
            elif not new_table_name: st.warning("Table name cannot be empty.")
            else: st.warning(f"Table '{new_table_name}' already exists.")
    
    active_table_schema = st.session_state.table_schemas.get(st.session_state.active_table_name, [])

    # Pre-populate from synthetic_df if schema is empty and df is available, and no template action pending
    # This logic now applies to the active_table_schema
    if not st.session_state.initial_schema_populated and synthetic_df is not None and not active_table_schema and st.session_state.active_table_name:
            for col_name_from_df in synthetic_df.columns:
                inferred_type = "string" # Default
                inferred_constraint = ""
                # Attempt to find a canonical mapping for the column name
                # This helps in pre-filling schema with more specific types if the column name is recognized
                
                # Normalize col_name_from_df for matching (e.g., lower, replace space with underscore)
                normalized_col_name = col_name_from_df.lower().replace(" ", "_")
                canonical_match = None

                # Check direct match in canonical map
                if normalized_col_name in CANONICAL_FIELD_TO_SCHEMA_DETAILS_MAP:
                    canonical_match = normalized_col_name
                else: # Check synonyms
                    for syn, can in FIELD_SYNONYM_TO_CANONICAL_MAP.items():
                        if syn == normalized_col_name: # Prioritize exact synonym match
                            if can in CANONICAL_FIELD_TO_SCHEMA_DETAILS_MAP:
                                canonical_match = can
                                break
                        elif syn in normalized_col_name: # Broader match (less reliable)
                             if can in CANONICAL_FIELD_TO_SCHEMA_DETAILS_MAP:
                                 canonical_match = can # Take first broader match
                                 # break # Don't break, allow more specific synonym match later
                
                field_details_for_schema = {"name": col_name_from_df}

                if canonical_match and canonical_match in CANONICAL_FIELD_TO_SCHEMA_DETAILS_MAP:
                    details = CANONICAL_FIELD_TO_SCHEMA_DETAILS_MAP[canonical_match]
                    field_details_for_schema["type"] = details["type"]
                    field_details_for_schema["constraint"] = details["constraint"]
                    # Carry over relevant flags for generation if needed by schema editor (e.g. for display hints)
                    # For now, type and constraint are primary.
                else: # Fallback to regex-based inference if no canonical match
                    for pattern, f_type, f_constraint in SCHEMA_INFERENCE_RULES:
                        if re.search(pattern, col_name_from_df, re.IGNORECASE):
                            field_details_for_schema["type"] = f_type
                            field_details_for_schema["constraint"] = f_constraint
                            break 
                
                # Ensure default type and constraint if still not set
                field_details_for_schema.setdefault("type", "string")
                field_details_for_schema.setdefault("constraint", "")
                field_details_for_schema["pii_handling"] = st.session_state.get(DEFAULT_PII_STRATEGY_KEY, "realistic_fake")

                st.session_state.table_schemas[st.session_state.active_table_name].append(field_details_for_schema)
            st.session_state.initial_schema_populated = True # Mark as populated

    # --- Template Selector ---
    template_options = list(SCHEMA_TEMPLATES.keys())
    if 'selected_template_name' not in st.session_state:
        # Default to "None (Custom Schema)" or the first option
        st.session_state.selected_template_name = template_options[0] 

    st.markdown("---")
    newly_selected_template = st.selectbox(
        f"Load Schema Template for '{st.session_state.active_table_name}' (optional, will replace its schema)",
        options=template_options,
        index=template_options.index(st.session_state.selected_template_name),
        key="template_selector_key"
    )

    if newly_selected_template != st.session_state.selected_template_name:
        st.session_state.selected_template_name = newly_selected_template
        if newly_selected_template != "None (Custom Schema)":
            template_data = SCHEMA_TEMPLATES[newly_selected_template]
            st.session_state.table_schemas[st.session_state.active_table_name] = [] # Clear active table's schema
            default_global_pii_strategy = st.session_state.get(DEFAULT_PII_STRATEGY_KEY, "realistic_fake")
            for field_template in template_data.get("fields", []): # Access the 'fields' key
                new_field = field_template.copy()
                # If template suggests a canonical type, try to get its full details
                canonical_suggestion = new_field.pop("_canonical_suggestion", None)
                if canonical_suggestion and canonical_suggestion in CANONICAL_FIELD_TO_SCHEMA_DETAILS_MAP:
                    canonical_details = CANONICAL_FIELD_TO_SCHEMA_DETAILS_MAP[canonical_suggestion]
                    new_field["type"] = canonical_details.get("type", new_field["type"])
                    new_field["constraint"] = canonical_details.get("constraint", new_field["constraint"])
                    # Potentially copy other relevant flags from canonical_details to new_field if needed by schema editor

                is_sensitive = new_field["type"] in ["email", "phone", "aadhaar", "pan", "passport", "voterid", "ifsc", "upi", "name", "address"]
                new_field["pii_handling"] = new_field.get("pii_handling", default_global_pii_strategy if is_sensitive else "realistic_fake")
                st.session_state.table_schemas[st.session_state.active_table_name].append(new_field)
            st.session_state.generated_data_frames = {} # Clear previously generated multi-table data
            st.session_state.initial_schema_populated = True # Mark as populated (by template)
            # Clear data from other generation paths
            st.session_state.prompt_generated_df = None
            st.session_state.uploaded_df_for_schema = None
            st.session_state.playground_generated_df = None
            st.session_state.newly_generated_df_tab3 = None
            st.session_state.synthetic_df_from_file_tab3 = None
        else: # User selected "None (Custom Schema)"
            st.session_state.table_schemas[st.session_state.active_table_name] = [] # Clear active table's schema
            st.session_state.generated_data_frames = {}
            st.session_state.initial_schema_populated = False # Allow re-inference if data is available
        st.rerun()

    table_selector_cols = st.columns([2,1])
    with table_selector_cols[0]:
        if st.session_state.table_schemas:
            selected_active_table = st.selectbox(
                "Edit Schema For Table:",
                options=list(st.session_state.table_schemas.keys()),
                index=list(st.session_state.table_schemas.keys()).index(st.session_state.active_table_name) if st.session_state.active_table_name in st.session_state.table_schemas else 0,
                key="active_table_selector"
            )
            if selected_active_table != st.session_state.active_table_name:
                st.session_state.active_table_name = selected_active_table
                st.session_state.initial_schema_populated = False 
                st.rerun()
    with table_selector_cols[1]:
        if st.session_state.active_table_name and len(st.session_state.table_schemas) > 1:
            if st.button(f"🗑️ Delete Table '{st.session_state.active_table_name}'", key="delete_active_table", use_container_width=True):
                del st.session_state.table_schemas[st.session_state.active_table_name]
                st.session_state.relationships = [
                    r for r in st.session_state.relationships 
                    if r['parent_table'] != st.session_state.active_table_name and r['child_table'] != st.session_state.active_table_name
                ]
                st.session_state.active_table_name = list(st.session_state.table_schemas.keys())[0] if st.session_state.table_schemas else None
                st.session_state.initial_schema_populated = False
                st.rerun()
        elif len(st.session_state.table_schemas) <= 1:
            st.caption("Cannot delete the last table.")

    st.markdown("---")
    st.subheader(f"Field Definitions for Table: '{st.session_state.active_table_name}'")

    if not st.session_state.active_table_name:
        st.info("Please add or select a table to define its fields.")
        return 

    if st.session_state.active_table_name not in st.session_state.table_schemas:
        st.session_state.table_schemas[st.session_state.active_table_name] = []

    current_active_schema_fields = st.session_state.table_schemas[st.session_state.active_table_name]

    # Display existing schema fields
    for i, field in enumerate(current_active_schema_fields):
        is_sensitive_field = field["type"] in ["email", "phone", "aadhaar", "pan", "passport", "voterid", "ifsc", "upi", "name", "address"]
        cols = st.columns([3, 2, 2, 2, 1] if is_sensitive_field else [3, 2, 3, 1]) # Adjust columns based on sensitivity

        with cols[0]:
            current_active_schema_fields[i]["name"] = cols[0].text_input(
                f"Field Name",
                value=field["name"],
                key=f"field_name_{st.session_state.active_table_name}_{i}"
            )

        with cols[1]:
            current_active_schema_fields[i]["type"] = cols[1].selectbox(
                f"Field Type",
                options=list(FIELD_TYPES.keys()),
                format_func=lambda x: FIELD_TYPES[x],
                index=list(FIELD_TYPES.keys()).index(field["type"]) if field["type"] in FIELD_TYPES else 0,
                key=f"field_type_{st.session_state.active_table_name}_{i}"
            )

        field_type = current_active_schema_fields[i]["type"]
        with cols[2]:
            if field_type == "int" or field_type == "float":
                min_val_default, max_val_default = (1, 100) if field_type == "int" else (1.0, 1000.0)
                parsed_min, parsed_max = min_val_default, max_val_default
                current_constraint = field.get("constraint", "")
                if current_constraint:
                    match = re.match(r"^\s*(-?\d+(?:\.\d+)?)\s*-\s*(-?\d+(?:\.\d+)?)\s*$", current_constraint)
                    if match:
                        try:
                            parsed_min_str, parsed_max_str = match.groups()
                            if field_type == "int":
                                parsed_min = int(float(parsed_min_str))
                                parsed_max = int(float(parsed_max_str))
                            else: # float
                                parsed_min = float(parsed_min_str)
                                parsed_max = float(parsed_max_str)
                            if parsed_min > parsed_max:
                                parsed_min, parsed_max = min_val_default, max_val_default
                                st.warning(f"Invalid range in constraint '{current_constraint}' for '{field['name']}'. Reverted to default.")
                        except ValueError:
                            parsed_min, parsed_max = min_val_default, max_val_default
                            st.warning(f"Could not parse constraint '{current_constraint}' for '{field['name']}'. Reverted to default.")

                constraint_sub_cols = st.columns(2)
                num_input_step = 1 if field_type == "int" else 0.01
                num_input_format = "%d" if field_type == "int" else "%.2f"

                min_val_ui = constraint_sub_cols[0].number_input(
                    "Min", value=parsed_min, key=f"field_constraint_min_{st.session_state.active_table_name}_{i}",
                    step=num_input_step, format=num_input_format
                )
                max_val_ui = constraint_sub_cols[1].number_input(
                    "Max", value=parsed_max, key=f"field_constraint_max_{st.session_state.active_table_name}_{i}",
                    step=num_input_step, format=num_input_format
                )

                if field_type == "int":
                    min_val_ui = int(min_val_ui)
                    max_val_ui = int(max_val_ui)

                if min_val_ui > max_val_ui:
                    st.error(f"Min value ({min_val_ui}) cannot be greater than Max value ({max_val_ui}) for field '{field['name']}'. Constraint not updated to this invalid range.")
                else:
                    current_active_schema_fields[i]["constraint"] = f"{min_val_ui}-{max_val_ui}"

            elif field_type == "date":
                start_date_default = datetime.now().date() - timedelta(days=365)
                end_date_default = datetime.now().date()
                parsed_start_date, parsed_end_date = start_date_default, end_date_default
                current_constraint = field.get("constraint", "")
                if current_constraint:
                    match = re.match(r"^\s*(\d{4}-\d{2}-\d{2})\s*-\s*(\d{4}-\d{2}-\d{2})\s*$", current_constraint)
                    if match:
                        try:
                            start_str, end_str = match.groups()
                            parsed_start_date = datetime.strptime(start_str, "%Y-%m-%d").date()
                            parsed_end_date = datetime.strptime(end_str, "%Y-%m-%d").date()
                            if parsed_start_date > parsed_end_date:
                                parsed_start_date, parsed_end_date = start_date_default, end_date_default
                                st.warning(f"Invalid date range in constraint '{current_constraint}' for '{field['name']}'. Reverted to default.")
                        except ValueError:
                            parsed_start_date, parsed_end_date = start_date_default, end_date_default
                            st.warning(f"Could not parse date constraint '{current_constraint}' for '{field['name']}'. Reverted to default.")
                
                constraint_sub_cols = st.columns(2)
                start_date_ui = constraint_sub_cols[0].date_input("Start Date", value=parsed_start_date, key=f"field_constraint_start_{st.session_state.active_table_name}_{i}")
                end_date_ui = constraint_sub_cols[1].date_input("End Date", value=parsed_end_date, key=f"field_constraint_end_{st.session_state.active_table_name}_{i}")

                if start_date_ui > end_date_ui:
                    st.error(f"Start date ({start_date_ui}) cannot be after End date ({end_date_ui}) for field '{field['name']}'. Constraint not updated to this invalid range.")
                else:
                    current_active_schema_fields[i]["constraint"] = f"{start_date_ui.strftime('%Y-%m-%d')} - {end_date_ui.strftime('%Y-%m-%d')}"
            elif field_type == "category":
                current_active_schema_fields[i]["constraint"] = st.text_input(f"Categories (comma-sep)", value=field.get("constraint", ""), placeholder="e.g., Option A, Option B", key=f"field_constraint_cat_{st.session_state.active_table_name}_{i}")
            else: # Default text input for other types like string, email, phone etc. where constraints are less structured or not typically ranges
                current_active_schema_fields[i]["constraint"] = st.text_input(f"Constraint/Format", value=field.get("constraint", ""), placeholder="Optional constraint", key=f"field_constraint_other_{st.session_state.active_table_name}_{i}")

        if is_sensitive_field:
            with cols[3]:
                # Ensure pii_handling key exists for old schema items
                if "pii_handling" not in current_active_schema_fields[i]:
                     current_active_schema_fields[i]["pii_handling"] = st.session_state.get(DEFAULT_PII_STRATEGY_KEY, "realistic_fake")

                current_active_schema_fields[i]["pii_handling"] = cols[3].selectbox(
                    "PII Handling",
                    options=list(PII_HANDLING_STRATEGIES.keys()),
                    format_func=lambda x: PII_HANDLING_STRATEGIES[x],
                    index=list(PII_HANDLING_STRATEGIES.keys()).index(current_active_schema_fields[i]["pii_handling"]),
                    key=f"pii_handling_{st.session_state.active_table_name}_{i}",
                    help="Choose how to handle this sensitive field."
                )

        delete_button_col_index = 4 if is_sensitive_field else 3
        with cols[delete_button_col_index]:
            if cols[delete_button_col_index].button("Delete", key=f"delete_{st.session_state.active_table_name}_{i}"):
                current_active_schema_fields.pop(i)
                st.rerun()

    # Add new field button
    if st.button("➕ Add Field"):
        current_active_schema_fields.append({
            "name": f"Field{len(current_active_schema_fields) + 1}",
            "type": "string",
            "constraint": "",
            "pii_handling": st.session_state.get(DEFAULT_PII_STRATEGY_KEY, "realistic_fake") # Default for new fields
        })
        st.rerun()

    # --- Edge Case Injection UI ---
    # This section needs to be updated to select a table for the field condition
    render_edge_case_ui(
        ui_key_prefix="tab2_smart_schema",
        edge_cases_list_key='edge_cases', # Uses the main 'edge_cases' for Smart Schema Editor
        table_schemas_for_fields=st.session_state.table_schemas,
        available_table_names_for_conditions=list(st.session_state.table_schemas.keys())
    )

    # --- Relationship Definition UI ---
    st.markdown("---")
    st.subheader("🔗 Define Table Relationships (One-to-Many)")

    if not st.session_state.table_schemas or len(st.session_state.table_schemas) < 1:
        st.info("Define at least one table with fields to set up relationships.")
    else:
        all_table_names = list(st.session_state.table_schemas.keys())
        
        rel_cols = st.columns(4)
        parent_table_options = all_table_names
        selected_parent_table = rel_cols[0].selectbox("Parent Table (One)", parent_table_options, key="rel_parent_table")
        
        parent_pk_options = []
        if selected_parent_table and selected_parent_table in st.session_state.table_schemas:
            parent_pk_options = [f['name'] for f in st.session_state.table_schemas[selected_parent_table] if f['name']]
        selected_parent_pk = rel_cols[1].selectbox("Parent Primary Key (PK)", parent_pk_options, key="rel_parent_pk")

        child_table_options = [t for t in all_table_names if t != selected_parent_table] 
        selected_child_table = rel_cols[2].selectbox("Child Table (Many)", child_table_options, key="rel_child_table")

        child_fk_options = []
        if selected_child_table and selected_child_table in st.session_state.table_schemas:
            child_fk_options = [f['name'] for f in st.session_state.table_schemas[selected_child_table] if f['name']]
        selected_child_fk = rel_cols[3].selectbox("Child Foreign Key (FK)", child_fk_options, key="rel_child_fk")

        if st.button("🔗 Add Relationship", key="add_relationship_button"):
            if selected_parent_table and selected_parent_pk and selected_child_table and selected_child_fk:
                if selected_parent_table == selected_child_table:
                    st.error("Parent and Child tables cannot be the same for a simple one-to-many relationship.")
                else:
                    new_relationship = {
                        "parent_table": selected_parent_table,
                        "parent_pk": selected_parent_pk,
                        "child_table": selected_child_table,
                        "child_fk": selected_child_fk,
                    }
                    if new_relationship not in st.session_state.relationships:
                        st.session_state.relationships.append(new_relationship)
                        st.success(f"Relationship added: {selected_parent_table}.{selected_parent_pk} -> {selected_child_table}.{selected_child_fk}")
                        st.rerun()
                    else:
                        st.warning("This relationship already exists.")
            else:
                st.error("All fields are required to define a relationship.")

        if st.session_state.relationships:
            st.markdown("**Existing Relationships:**")
            for i, rel in enumerate(st.session_state.relationships):
                rel_text = f"{i+1}. **{rel['parent_table']}** (`{rel['parent_pk']}`) → **{rel['child_table']}** (`{rel['child_fk']}`)"
                del_rel_cols = st.columns([8,1])
                del_rel_cols[0].markdown(rel_text)
                if del_rel_cols[1].button("🗑️", key=f"del_rel_{i}"):
                    st.session_state.relationships.pop(i)
                    st.rerun()

    # --- Number of rows and PII Strategy (Global for now) ---
    st.markdown("---")
    num_rows_help_text = "Number of rows for root/parent tables. Child table rows will be derived."
    # num_rows = st.number_input("Number of rows to generate", min_value=1, value=max(1, num_rows), step=1) # Old way
    st.number_input(
        "Number of rows to generate", 
        min_value=1, 
        step=1, 
        key="num_rows_smart_schema_editor", # Uses session state
        help=num_rows_help_text)

    # Default PII Handling Strategy
    st.session_state[DEFAULT_PII_STRATEGY_KEY] = st.selectbox(
        "Default PII Handling Strategy (for new sensitive fields)",
        options=list(PII_HANDLING_STRATEGIES.keys()),
        format_func=lambda x: PII_HANDLING_STRATEGIES[x],
        index=list(PII_HANDLING_STRATEGIES.keys()).index(st.session_state.get(DEFAULT_PII_STRATEGY_KEY, "realistic_fake")),
        help="Default strategy for new sensitive fields or those not individually set."
    )

    # Generate button
    if st.button("🔄 Generate Data from All Schemas", use_container_width=True, key="generate_all_schemas_button"):
        if not st.session_state.table_schemas:
            st.error("No tables defined. Please add tables and define their schemas.")
        else:
            valid_relationships = True
            for rel in st.session_state.relationships:
                parent_fields = [f['name'] for f in st.session_state.table_schemas.get(rel['parent_table'], [])]
                child_fields = [f['name'] for f in st.session_state.table_schemas.get(rel['child_table'], [])]
                if rel['parent_pk'] not in parent_fields:
                    st.error(f"Relationship Error: PK '{rel['parent_pk']}' not found in table '{rel['parent_table']}'.")
                    valid_relationships = False
                if rel['child_fk'] not in child_fields:
                    st.error(f"Relationship Error: FK '{rel['child_fk']}' not found in table '{rel['child_table']}'.")
                    valid_relationships = False
            
            if valid_relationships:
                st.session_state.generated_data_frames = generate_hierarchical_data(
                    st.session_state.table_schemas,
                    st.session_state.relationships,
                    st.session_state.num_rows_smart_schema_editor, # Use session state value
                    st.session_state.edge_cases,
                    st.session_state.get(DEFAULT_PII_STRATEGY_KEY, "realistic_fake")
                )
                # Clear data from other generation paths
                st.session_state.prompt_generated_df = None
                st.session_state.uploaded_df_for_schema = None
                st.session_state.playground_generated_df = None
                st.session_state.newly_generated_df_tab3 = None
                st.session_state.synthetic_df_from_file_tab3 = None
                if st.session_state.generated_data_frames:
                    st.success("Hierarchical data generated successfully!")
                    if not st.session_state.active_display_table_name or st.session_state.active_display_table_name not in st.session_state.generated_data_frames:
                        st.session_state.active_display_table_name = list(st.session_state.generated_data_frames.keys())[0]
                else:
                    st.error("Failed to generate hierarchical data. Check errors above.")
            else:
                st.error("Please fix relationship errors before generating data.")

    # Display generated data if available
    if st.session_state.generated_data_frames:
        st.subheader("📊 Generated Datasets")
        
        display_table_options = list(st.session_state.generated_data_frames.keys())
        # Ensure active_display_table_name is valid
        if st.session_state.active_display_table_name not in display_table_options and display_table_options:
            st.session_state.active_display_table_name = display_table_options[0]
        
        st.session_state.active_display_table_name = st.selectbox(
            "Select table to display:",
            options=display_table_options,
            index=display_table_options.index(st.session_state.active_display_table_name) if st.session_state.active_display_table_name in display_table_options else 0,
            key="display_table_selector"
        )

        if st.session_state.active_display_table_name and st.session_state.active_display_table_name in st.session_state.generated_data_frames:
            df_to_display = st.session_state.generated_data_frames[st.session_state.active_display_table_name]
            st.dataframe(df_to_display)

            csv_selected = df_to_display.to_csv(index=False)
            excel_buffer_selected = io.BytesIO()
            with pd.ExcelWriter(excel_buffer_selected, engine='openpyxl') as writer_excel_sel:
                df_to_display.to_excel(writer_excel_sel, index=False, sheet_name=st.session_state.active_display_table_name)
            excel_data_selected = excel_buffer_selected.getvalue()

            dl_cols = st.columns(3)
            with dl_cols[0]:
                st.download_button(
                    label=f"Download CSV ({st.session_state.active_display_table_name})",
                    data=csv_selected,
                    file_name=f"{st.session_state.active_display_table_name}.csv",
                    mime="text/csv",
                    key=f"download_csv_selected_{st.session_state.active_display_table_name}"
                )
            with dl_cols[1]:
                st.download_button(
                    label=f"Download Excel ({st.session_state.active_display_table_name})",
                    data=excel_data_selected,
                    file_name=f"{st.session_state.active_display_table_name}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"download_excel_selected_{st.session_state.active_display_table_name}"
                )

            if len(st.session_state.generated_data_frames) > 0: # Show only if there are tables
                with dl_cols[2]:
                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                        for table_n, table_df in st.session_state.generated_data_frames.items():
                            zip_file.writestr(f"{table_n}.csv", table_df.to_csv(index=False))
                    
                    st.download_button( # This button directly triggers download
                        label="Download All Tables (ZIP)",
                        data=zip_buffer.getvalue(),
                        file_name="all_tables_synthetic_data.zip",
                        mime="application/zip",
                        key="download_all_zip_final"
                    )
    # Fallback to old single schema_df display if it exists and no multi-table data generated
    elif "schema_df" in st.session_state and st.session_state.schema_df is not None and not st.session_state.generated_data_frames:
        st.subheader(f"📊 Generated Data (Legacy Single Table: {st.session_state.active_table_name or 'N/A'})")
        st.dataframe(st.session_state.schema_df)
        # ... (old download buttons for single schema_df, if you want to keep them as a fallback)


# --- Helper functions for Ethical AI Dashboard Scores ---
def calculate_bias_score(df):
    """Calculates a bias score based on categorical column distributions."""
    if df is None or df.empty:
        return 50 # Default score if no data

    categorical_cols = df.select_dtypes(include='object').columns.tolist()
    # Attempt to find more categoricals if 'object' type is not sufficient
    if not categorical_cols:
         categorical_cols = [col for col in df.columns if df[col].nunique() < 20 and df[col].nunique() > 1 and df[col].nunique() < len(df)]


    if not categorical_cols:
        return 100 # No categorical columns to assess bias, so perfectly unbiased in this context

    column_scores = []
    for col in categorical_cols:
        counts = df[col].value_counts()
        num_categories = len(counts)
        if num_categories <= 1:
            column_scores.append(100) # Perfectly uniform or single category
            continue

        # Handle potential NaN values in counts before calculating probabilities
        probabilities = counts / len(df[col].dropna())
        # Ensure probabilities sum to 1 (or close to it) and handle potential floating point issues
        probabilities = probabilities[probabilities > 0] # Remove categories with 0 count after dropna
        probabilities = probabilities / probabilities.sum() # Re-normalize

        if len(probabilities) <= 1: # After cleaning, might have only one category left
             column_scores.append(100)
             continue

        entropy = -np.sum(probabilities * np.log2(probabilities + 1e-9)) # Add epsilon to avoid log(0)
        max_entropy = np.log2(num_categories)

        normalized_entropy = entropy / max_entropy if max_entropy > 0 else 1.0
        column_scores.append(normalized_entropy * 100)

    return np.mean(column_scores) if column_scores else 50

MIN_SAMPLES_FOR_DRIFT_TEST = 20 # Minimum samples required in each series for a reliable test
DRIFT_P_VALUE_THRESHOLD = 0.05 # Standard p-value threshold

def detect_numerical_drift(series1, series2, column_name):
    """
    Detects drift between two numerical series using the Kolmogorov-Smirnov (K-S) test.
    Returns: (bool: drift_detected, str: message)
    """
    s1_clean = series1.dropna()
    s2_clean = series2.dropna()

    if len(s1_clean) < MIN_SAMPLES_FOR_DRIFT_TEST or len(s2_clean) < MIN_SAMPLES_FOR_DRIFT_TEST:
        return False, f"Insufficient data for drift test in '{column_name}' (s1: {len(s1_clean)}, s2: {len(s2_clean)} samples)."

    if s1_clean.nunique() == 1 and s2_clean.nunique() == 1 and s1_clean.iloc[0] == s2_clean.iloc[0]:
        return False, f"No drift in '{column_name}'; both series are constant and identical."
    
    try:
        ks_statistic, p_value = stats.ks_2samp(s1_clean, s2_clean)
        if p_value < DRIFT_P_VALUE_THRESHOLD:
            return True, f"Drift detected in '{column_name}' (K-S test, p={p_value:.3g}). Distributions likely differ."
        else:
            return False, f"No significant drift detected in '{column_name}' (K-S test, p={p_value:.3g})."
    except Exception as e:
        return False, f"Error during K-S test for '{column_name}': {e}"

def detect_categorical_drift(series1, series2, column_name):
    """
    Detects drift between two categorical series using the Chi-squared test.
    Returns: (bool: drift_detected, str: message)
    """
    s1_clean = series1.dropna()
    s2_clean = series2.dropna()

    if len(s1_clean) < MIN_SAMPLES_FOR_DRIFT_TEST or len(s2_clean) < MIN_SAMPLES_FOR_DRIFT_TEST:
        return False, f"Insufficient data for drift test in '{column_name}' (s1: {len(s1_clean)}, s2: {len(s2_clean)} samples)."

    if s1_clean.nunique() == 0 or s2_clean.nunique() == 0:
         return False, f"No data to compare for drift in '{column_name}' after cleaning."

    s1_counts = s1_clean.value_counts()
    s2_counts = s2_clean.value_counts()

    combined_index = sorted(list(set(s1_counts.index) | set(s2_counts.index)))

    if not combined_index:
        return False, f"No common categories or data to compare for '{column_name}'."

    observed_df = pd.DataFrame({
        's1': s1_counts.reindex(combined_index, fill_value=0),
        's2': s2_counts.reindex(combined_index, fill_value=0)
    })
    
    # Remove categories that are zero in both (shouldn't happen if combined_index is from actual counts)
    observed_df = observed_df.loc[(observed_df['s1'] > 0) | (observed_df['s2'] > 0)]

    if observed_df.shape[0] < 2 or observed_df.shape[1] < 2: # Need at least a 2x2 table for chi2
        # This can happen if one series is all one value and the other is all another, or one is empty.
        # Check if distributions are identical by comparing normalized value counts
        s1_norm_counts = s1_clean.value_counts(normalize=True).sort_index()
        s2_norm_counts = s2_clean.value_counts(normalize=True).sort_index()
        if s1_norm_counts.equals(s2_norm_counts):
            return False, f"No drift in '{column_name}'; distributions are identical (small sample/categories)."
        else:
            return True, f"Drift detected in '{column_name}'; distributions differ (small sample/categories)."

    try:
        chi2, p_value, dof, expected = stats.chi2_contingency(observed_df.values)
        if p_value < DRIFT_P_VALUE_THRESHOLD:
            # Check for low expected frequencies
            if (expected < 5).any().any(): # If any cell in expected frequencies is < 5
                 return True, f"Drift detected in '{column_name}' (Chi-squared, p={p_value:.3g}). Note: Some expected frequencies are low (<5), test may be less reliable."
            return True, f"Drift detected in '{column_name}' (Chi-squared, p={p_value:.3g}). Distributions likely differ."
        else:
            return False, f"No significant drift detected in '{column_name}' (Chi-squared, p={p_value:.3g})."
    except ValueError as e: # Catches errors like "The internally computed table of expected frequencies has a zero element at..."
        # This can happen if a whole row/column sum is 0 in the contingency table.
        # Fallback to comparing normalized value counts for equality.
        s1_norm_counts = s1_clean.value_counts(normalize=True).sort_index()
        s2_norm_counts = s2_clean.value_counts(normalize=True).sort_index()
        if s1_norm_counts.equals(s2_norm_counts):
            return False, f"No drift in '{column_name}'; distributions appear identical (Chi-squared error: {e})."
        else:
            return True, f"Drift detected in '{column_name}'; distributions differ (Chi-squared error: {e})."
    except Exception as e:
        return False, f"Error during Chi-squared test for '{column_name}': {e}"

def calculate_bias_score(df): # Original calculate_bias_score function was here, moved the new one above it.
    """Calculates a bias score based on categorical column distributions."""
    if df is None or df.empty:
        return 50 # Default score if no data

    categorical_cols = df.select_dtypes(include='object').columns.tolist()
    # Attempt to find more categoricals if 'object' type is not sufficient
    if not categorical_cols:
         categorical_cols = [col for col in df.columns if df[col].nunique() < 20 and df[col].nunique() > 1 and df[col].nunique() < len(df)]


    if not categorical_cols:
        return 100 # No categorical columns to assess bias, so perfectly unbiased in this context

    column_scores = []
    for col in categorical_cols:
        counts = df[col].value_counts()
        num_categories = len(counts)
        if num_categories <= 1:
            column_scores.append(100) # Perfectly uniform or single category
            continue

        # Handle potential NaN values in counts before calculating probabilities
        probabilities = counts / len(df[col].dropna())
        # Ensure probabilities sum to 1 (or close to it) and handle potential floating point issues
        probabilities = probabilities[probabilities > 0] # Remove categories with 0 count after dropna
        probabilities = probabilities / probabilities.sum() # Re-normalize

        if len(probabilities) <= 1: # After cleaning, might have only one category left
             column_scores.append(100)
             continue

        entropy = -np.sum(probabilities * np.log2(probabilities + 1e-9)) # Add epsilon to avoid log(0)
        max_entropy = np.log2(num_categories)

        normalized_entropy = entropy / max_entropy if max_entropy > 0 else 1.0
        column_scores.append(normalized_entropy * 100)

    return np.mean(column_scores) if column_scores else 50

def calculate_compliance_score(pii_risk_level, dpdp_risk_level, bias_score_val):
    """Calculates a compliance score."""
    score = 100

    # Deductions for PII/DPDP risk
    if pii_risk_level == "High": score -= 25
    elif pii_risk_level == "Medium": score -= 10 # Assuming we might add a Medium state

    if dpdp_risk_level == "High": score -= 35 # Higher penalty for DPDP
    elif dpdp_risk_level == "Medium": score -= 15

    # Deduction for bias (if bias score is low, deduct more)
    if bias_score_val < 70: # If bias score is less than 70 (meaning more biased)
        score -= (70 - bias_score_val) * 0.5 # Scale down the penalty from bias

    return max(0, min(100, round(score))) # Ensure score is between 0 and 100

# --- Helper functions for synthesizing data from uploaded file ---
def _synthesize_numeric_column_from_upload(original_series, column_name, num_rows_to_generate):
    """Generates a new numeric series with num_rows_to_generate, based on original_series characteristics."""
    if not pd.api.types.is_numeric_dtype(original_series):
        st.warning(f"Attempted to synthesize non-numeric column '{column_name}' as numeric. Skipping.")
        return pd.Series([np.nan] * num_rows_to_generate, name=column_name)

    valid_series = original_series.dropna()
    if valid_series.empty:
        return pd.Series([np.nan] * num_rows_to_generate, name=column_name)

    min_val, max_val = valid_series.min(), valid_series.max()
    is_integer_type = pd.api.types.is_integer_dtype(original_series)

    generated_values = []
    for _ in range(num_rows_to_generate):
        if min_val == max_val: # Only one unique non-NaN value
            val = min_val
        elif is_integer_type:
            # Ensure min_val and max_val are integers for randint
            val = random.randint(int(round(min_val)), int(round(max_val)))
        else:
            val = random.uniform(min_val, max_val)
        
        if any(keyword in column_name.lower() for keyword in ['age', 'salary', 'price', 'cost', 'amount', 'quantity', 'marks', 'bedrooms', 'bathrooms']):
            val = max(1 if is_integer_type else 0.01, val) # Ensure positivity
        
        # If original was integer, round the result if it became float due to positivity adjustment
        if is_integer_type and isinstance(val, float):
            val = round(val)
            
        generated_values.append(val)
    
    # Attempt to cast back to original dtype if possible, esp. for integers
    new_series = pd.Series(generated_values, name=column_name)
    return new_series.astype(original_series.dtype, errors='ignore')


PII_FIELD_SYNTHESIZERS_FOR_UPLOAD = {
    # Maps keywords found in column names to (field_type_for_generate_value, constraint_for_generate_value)
    'name': ("name", ""),
    'email': ("email", ""),
    'phone': ("phone", ""),
    'address': ("address", ""),
    'aadhaar': ("aadhaar", ""),
    'pan': ("pan", ""),
    'passport': ("passport", ""),
    'voter': ("voterid", ""), # Catches "voter" or "voter id"
    'ifsc': ("ifsc", ""),
    'upi': ("upi", ""),
    # SSN is in PII_FIELDS but not DPDP_PII_FIELDS, add if needed: 'ssn': ("ssn", "")
}

def _synthesize_categorical_column_from_upload(original_series, column_name, num_rows_to_generate):
    """Synthesizes a categorical/object column with num_rows_to_generate. Fakes PII, samples others."""
    # Check for PII first
    for keyword, (gen_type, gen_constraint) in PII_FIELD_SYNTHESIZERS_FOR_UPLOAD.items():
        if keyword in column_name.lower() and (column_name in PII_FIELDS or is_dpdp_pii(column_name)):
            default_pii_strategy = st.session_state.get(DEFAULT_PII_STRATEGY_KEY, "realistic_fake")
            field_schema_for_gen = {"type": gen_type, "constraint": gen_constraint, "name": column_name, "pii_handling": default_pii_strategy}
            return pd.Series([_generate_value_from_schema(field_schema_for_gen) for _ in range(num_rows_to_generate)], name=column_name)

    # Non-PII: sample from original distribution
    cleaned_series = original_series.dropna()
    if cleaned_series.empty:
        return pd.Series([np.nan] * num_rows_to_generate, name=column_name)

    value_counts = cleaned_series.value_counts(normalize=True)
    if value_counts.empty: # Should be caught by cleaned_series.empty, but safeguard
        return pd.Series([np.nan] * num_rows_to_generate, name=column_name)
        
    unique_values = value_counts.index
    probabilities = value_counts.values
    
    # Ensure probabilities sum to 1 (can be off due to floating point issues)
    probabilities = probabilities / np.sum(probabilities)

    generated_values = np.random.choice(unique_values, size=num_rows_to_generate, p=probabilities)
    
    new_series = pd.Series(generated_values, name=column_name)
    return new_series.astype(original_series.dtype, errors='ignore')


# Tabs for different workflows
tab1, tab2, tab3, tab4, tab5, tab_advanced_lab = st.tabs(["Text-based Generation", "Smart Schema Editor", "File-based Generation", "Scenario Playground", "Community Gallery", "🔬 Advanced Lab"])

with tab1:
    # Main Content for Text-based Generation
    if prompt:
        should_regenerate_tab1 = False
        if st.session_state.prompt_generated_df is None:
            should_regenerate_tab1 = True
        elif prompt != st.session_state.get('last_processed_prompt_tab1', ""):
            should_regenerate_tab1 = True

        if should_regenerate_tab1:
            # synthetic_df = generate_synthetic_data(prompt) # Old
            synthetic_df_tab1, num_rows_generated_tab1, inferred_schema_for_editor_tab1 = generate_synthetic_data(prompt) # New

            st.session_state.prompt_generated_df = synthetic_df_tab1 # Store for tab2 access
            st.session_state.num_rows_from_prompt = num_rows_generated_tab1 # Store for editor
            st.session_state.inferred_schema_from_prompt = inferred_schema_for_editor_tab1 # Store for editor
            st.session_state.last_processed_prompt_tab1 = prompt # Update last processed prompt

            if synthetic_df_tab1 is not None:
                # Clear data from other generation paths only when new prompt data is generated
                st.session_state.uploaded_df_for_schema = None
                st.session_state.generated_data_frames = {} # From Smart Schema
                st.session_state.active_display_table_name = None
                st.session_state.playground_generated_df = None
                st.session_state.newly_generated_df_tab3 = None
                st.session_state.synthetic_df_from_file_tab3 = None
            # If synthetic_df_tab1 is None, prompt_generated_df will be None, and the UI below won't render.

        # Always work with st.session_state.prompt_generated_df from this point onwards in Tab 1
        if st.session_state.prompt_generated_df is not None:
            st.subheader("📊 Generated Synthetic Data")
            # st.dataframe(synthetic_df) # Old way

            # --- Column Name Editing ---
            st.markdown("---")
            st.subheader("✏️ Edit Column Names (Optional)")

            # Use a form to batch column name changes
            with st.form(key="column_rename_form_prompt"):
                current_df_for_rename = st.session_state.prompt_generated_df
                original_column_names_list = list(current_df_for_rename.columns)
                
                proposed_new_names_inputs = {} # To store the st.text_input widgets' current values

                num_cols_to_edit = len(original_column_names_list) # Define num_cols_to_edit
                if num_cols_to_edit > 0:
                    # Dynamically create columns for text inputs for better UI
                    # Max 4 input fields per row for readability
                    num_ui_cols = min(num_cols_to_edit, 4) 
                    ui_columns_for_inputs = st.columns(num_ui_cols)
                    
                    for i, original_col_name in enumerate(original_column_names_list):
                        # Cycle through the UI columns
                        current_ui_col = ui_columns_for_inputs[i % num_ui_cols]
                        with current_ui_col:
                            # The text_input widgets are part of the form.
                            # Their values will be in st.session_state on submit.
                            st.text_input(
                                label=f"Rename '{original_col_name}'",
                                value=original_col_name, # Default to current name
                                key=f"txt_rename_prompt_idx_{i}" # Unique key using index
                            )
                
                submit_rename_button = st.form_submit_button("Apply Column Name Changes")

            if submit_rename_button:
                actual_rename_map = {}
                final_name_list_check = [] # For checking duplicates among new names
                valid_rename = True

                for i, original_name in enumerate(original_column_names_list):
                    widget_key = f"txt_rename_prompt_idx_{i}"
                    new_name = st.session_state[widget_key].strip() # Get submitted value from session_state

                    if not new_name:
                        st.error(f"Column name for original '{original_name}' cannot be empty.")
                        valid_rename = False
                        break
                    if new_name in final_name_list_check:
                        st.error(f"Duplicate column name '{new_name}' proposed. Please use unique names.")
                        valid_rename = False
                        break
                    final_name_list_check.append(new_name)
                    if new_name != original_name:
                        actual_rename_map[original_name] = new_name
                
                if valid_rename and actual_rename_map: # Proceed only if valid and there are changes
                    # Update DataFrame
                    df_to_update = st.session_state.prompt_generated_df.copy()
                    
                    # --- NEW: Intelligent Data Re-generation based on new column name ---
                    updated_schema_fields_for_tab1 = []
                    original_df_columns_before_rename = list(st.session_state.prompt_generated_df.columns) # For mapping to schema

                    # First, apply the structural rename to the DataFrame
                    df_to_update.rename(columns=actual_rename_map, inplace=True)

                    if st.session_state.inferred_schema_from_prompt and len(st.session_state.inferred_schema_from_prompt) == len(original_df_columns_before_rename):
                        for i, original_col_name_for_schema_lookup in enumerate(original_df_columns_before_rename):
                            current_field_schema_dict = st.session_state.inferred_schema_from_prompt[i].copy() # Work on a copy
                            
                            new_col_name_for_data = actual_rename_map.get(original_col_name_for_schema_lookup, original_col_name_for_schema_lookup)
                            
                            # If the name actually changed, re-infer schema and re-generate data
                            if new_col_name_for_data != original_col_name_for_schema_lookup:
                                st.write(f"Re-evaluating column: '{original_col_name_for_schema_lookup}' renamed to '{new_col_name_for_data}'") # Debug
                                
                                # 1. Infer schema for the new_col_name_for_data
                                effective_field_schema_for_regen = {"name": new_col_name_for_data, "type": "string", "constraint": "", "pii_handling": st.session_state.get(DEFAULT_PII_STRATEGY_KEY, "realistic_fake")}
                                
                                new_name_lower = new_col_name_for_data.lower()
                                new_canonical_match = None
                                sorted_synonyms_for_regen = sorted(FIELD_SYNONYM_TO_CANONICAL_MAP.keys(), key=len, reverse=True)
                                for syn in sorted_synonyms_for_regen: # Find best canonical match
                                    if syn == new_name_lower:
                                        if FIELD_SYNONYM_TO_CANONICAL_MAP[syn] in CANONICAL_FIELD_TO_SCHEMA_DETAILS_MAP:
                                            new_canonical_match = FIELD_SYNONYM_TO_CANONICAL_MAP[syn]; break
                                if not new_canonical_match:
                                    for syn in sorted_synonyms_for_regen:
                                        if syn in new_name_lower:
                                            if FIELD_SYNONYM_TO_CANONICAL_MAP[syn] in CANONICAL_FIELD_TO_SCHEMA_DETAILS_MAP:
                                                new_canonical_match = FIELD_SYNONYM_TO_CANONICAL_MAP[syn]; break
                                
                                if new_canonical_match:
                                    canonical_details = CANONICAL_FIELD_TO_SCHEMA_DETAILS_MAP[new_canonical_match]
                                    effective_field_schema_for_regen.update(canonical_details) # Add all flags and details
                                    effective_field_schema_for_regen["type"] = canonical_details.get("type", "string") # Ensure type is set
                                    effective_field_schema_for_regen["constraint"] = canonical_details.get("constraint", "") # Ensure constraint is set
                                else: # Fallback to basic keyword inference for novel names
                                    if any(kw in new_name_lower for kw in ["date", "time", "timestamp", "dob", "joining"]): effective_field_schema_for_regen["type"] = "date"
                                    elif any(kw in new_name_lower for kw in ["id", "count", "age", "quantity", "salary", "year", "number", "level"]):
                                        effective_field_schema_for_regen["type"] = "int"
                                        if "age" in new_name_lower: effective_field_schema_for_regen["constraint"] = "18-80"
                                        else: effective_field_schema_for_regen["constraint"] = "0-1000"
                                    elif any(kw in new_name_lower for kw in ["rate", "percent", "ratio", "score", "price", "amount", "value", "balance"]):
                                        effective_field_schema_for_regen["type"] = "float"; effective_field_schema_for_regen["constraint"] = "0.0-100.0"
                                    elif "email" in new_name_lower: effective_field_schema_for_regen["type"] = "email"
                                    elif "phone" in new_name_lower: effective_field_schema_for_regen["type"] = "phone"
                                    elif any(kw in new_name_lower for kw in ["address", "location", "city", "state", "country", "pincode", "zipcode"]): effective_field_schema_for_regen["type"] = "address"
                                    elif "name" in new_name_lower and not any(kw in new_name_lower for kw in ["user", "company", "product", "brand", "model", "sensor", "item", "file", "column", "field"]): effective_field_schema_for_regen["type"] = "name"
                                
                                # Ensure the name in the schema is the new name for generation
                                effective_field_schema_for_regen["name"] = new_col_name_for_data

                                # 2. Re-generate data for this column in df_to_update
                                num_rows_for_regen = len(df_to_update)
                                df_to_update[new_col_name_for_data] = [generate_value(effective_field_schema_for_regen.copy()) for _ in range(num_rows_for_regen)] # Pass a copy
                                st.write(f"Data for '{new_col_name_for_data}' re-generated with type '{effective_field_schema_for_regen['type']}' and constraint '{effective_field_schema_for_regen['constraint']}'.") # Debug
                                
                                # Update the schema list item
                                current_field_schema_dict.update(effective_field_schema_for_regen)
                            else: # Name didn't change, just ensure the name in schema dict is current
                                current_field_schema_dict["name"] = new_col_name_for_data
                            
                            updated_schema_fields_for_tab1.append(current_field_schema_dict)
                        
                        st.session_state.prompt_generated_df = df_to_update
                        st.session_state.inferred_schema_from_prompt = updated_schema_fields_for_tab1
                        st.success("Column names updated, relevant data re-generated, and schema updated!")
                    else:
                        st.session_state.prompt_generated_df = df_to_update # Update DF even if schema couldn't be
                        st.warning("DataFrame column names updated and data re-generated. Could not update schema field names due to a mismatch or missing schema. The original schema structure (with old names) will be sent to the editor if you proceed.")
                    st.rerun()
                elif valid_rename and not actual_rename_map:
                    st.info("No column names were changed.")
                # If not valid_rename, errors are already displayed by the checks above.

            st.markdown("---")
            st.subheader("✏️ Edit Data Values (Optional)")
            # Use st.data_editor for cell editing. It operates on st.session_state.prompt_generated_df
            edited_df_from_editor = st.data_editor(
                st.session_state.prompt_generated_df, 
                key="prompt_data_editor_main_tab1", # A consistent key for the editor, ensure it's unique if used elsewhere
                num_rows="dynamic", # Allows adding/deleting rows
                use_container_width=True
            )
            # Check if the DataFrame was actually changed by the data_editor
            if not edited_df_from_editor.equals(st.session_state.prompt_generated_df):
                 st.session_state.prompt_generated_df = edited_df_from_editor
                 # Note: num_rows_from_prompt might become inconsistent if rows are added/deleted.
                 # The schema editor uses its own row count, so this is acceptable for now.
                 st.success("Data edits applied. These edits are local to this tab. The 'Send to Editor' button will use the original inferred schema structure (with updated names if you applied them).")
                 # No st.rerun() needed here as st.data_editor updates reactively if its output is assigned.

            # --- Button to send schema to Smart Editor ---
            if st.session_state.inferred_schema_from_prompt: # Check if schema exists (it might have been name-updated)
                if st.button("⚙️ Send Inferred Schema to Smart Editor for Refinement", key="send_prompt_schema_to_editor"):
                    st.session_state.load_prompt_schema_to_editor = True
                    # The actual loading happens in show_smart_schema_editor on rerun
                    st.success("Schema will be loaded into the Smart Schema Editor. Please navigate to Tab 2.")
                    st.rerun() # Trigger rerun to activate logic in Tab 2
            else:
                st.caption("No detailed schema was inferred from the prompt (or it was lost). Cannot send to editor.")

            # Dataset Summary
            st.markdown(f"**Rows:** {st.session_state.prompt_generated_df.shape[0]} | **Columns:** {st.session_state.prompt_generated_df.shape[1]}")

            # --- Ethical AI Dashboard for Prompt-Generated Data ---
            st.subheader("🛡️ Ethical AI Dashboard (Prompt Data)")
            dash_col1, dash_col2, dash_col3, dash_col4 = st.columns(4)

            pii_cols_prompt = [col for col in st.session_state.prompt_generated_df.columns if col in PII_FIELDS or is_dpdp_pii(col)] # Combine PII and DPDP for general risk
            dpdp_cols_prompt_specific = [col for col in st.session_state.prompt_generated_df.columns if is_dpdp_pii(col)]

            pii_risk_level_prompt = "High" if pii_cols_prompt else "Low"
            dpdp_risk_level_prompt = "High" if dpdp_cols_prompt_specific else "Low"
            
            bias_score_prompt = calculate_bias_score(st.session_state.prompt_generated_df)
            compliance_score_prompt = calculate_compliance_score(pii_risk_level_prompt, dpdp_risk_level_prompt, bias_score_prompt)

            dash_col1.metric("Bias Score", f"{bias_score_prompt:.0f} / 100",
                             help="Measures fairness in categorical distributions. Higher (closer to 100) is better (more balanced).")
            dash_col2.metric("PII Risk", pii_risk_level_prompt,
                             help=f"Detected PII-like fields: {', '.join(pii_cols_prompt)}" if pii_cols_prompt else "No common PII fields detected.")
            dash_col3.metric("DPDP Risk", dpdp_risk_level_prompt,
                             help=f"Detected DPDP-specific fields: {', '.join(dpdp_cols_prompt_specific)}" if dpdp_cols_prompt_specific else "No specific DPDP fields detected.")
            dash_col4.metric("Compliance Score", f"{compliance_score_prompt:.0f}%",
                             help="An estimated score for overall compliance readiness based on PII, DPDP, and Bias.")

            # --- Bias Detection for Prompt-Generated Data ---
            st.subheader("📊 Bias Detection (Prompt Data)")
            categorical_cols_prompt = st.session_state.prompt_generated_df.select_dtypes(include='object').columns.tolist()
            # Attempt to find more categoricals if 'object' type is not sufficient
            if not categorical_cols_prompt:
                 categorical_cols_prompt = [col for col in st.session_state.prompt_generated_df.columns if st.session_state.prompt_generated_df[col].nunique() < 20 and st.session_state.prompt_generated_df[col].nunique() > 1]

            if categorical_cols_prompt:
                selected_col_prompt = st.selectbox(
                    "Select a categorical column from generated data to check bias",
                    categorical_cols_prompt,
                    key="bias_checker_prompt"
                )
                if selected_col_prompt:
                    value_counts_prompt = st.session_state.prompt_generated_df[selected_col_prompt].value_counts(normalize=True) * 100
                    fig_prompt = px.bar(
                        value_counts_prompt,
                        x=value_counts_prompt.index,
                        y=value_counts_prompt.values,
                        labels={"x": selected_col_prompt, "y": "Percentage"},
                        title=f"Distribution in '{selected_col_prompt}' (Prompt Generated Data)"
                    )
                    st.plotly_chart(fig_prompt)
            else:
                st.info("No suitable categorical columns found in the prompt-generated data for bias checking.")

            # --- Compliance Report Generator for Prompt-Generated Data ---
            st.subheader("📄 Download Compliance Report (Prompt Data)")
            if st.button("Generate & Download PDF Report (Prompt Data)", key="pdf_prompt_data"):
                pdf = FPDF()
                pdf.add_page()
                pdf.set_font("Arial", size=12)
                pdf.cell(200, 10, txt=f"{APP_NAME} - Synthetic Data Report (From Text Prompt)", ln=True, align="C")
                pdf.ln(10)
                pdf.cell(200, 10, txt=f"Dataset Source: Text Prompt", ln=True)
                pdf.cell(200, 10, txt=f"Generated Rows: {st.session_state.prompt_generated_df.shape[0]}, Columns: {st.session_state.prompt_generated_df.shape[1]}", ln=True)
                pdf.cell(200, 10, txt=f"Bias Score: {bias_score_prompt:.0f} / 100", ln=True)
                pdf.cell(200, 10, txt=f"PII Risk: {pii_risk_level_prompt}", ln=True)
                if pii_cols_prompt:
                    pdf.cell(200, 10, txt=f"Detected PII Fields: {', '.join(pii_cols_prompt)}", ln=True)
                pdf.cell(200, 10, txt=f"DPDP Risk: {dpdp_risk_level_prompt}", ln=True)
                if dpdp_cols_prompt_specific:
                    pdf.cell(200, 10, txt=f"Detected DPDP-Specific Fields: {', '.join(dpdp_cols_prompt_specific)}", ln=True)
                pdf.cell(200, 10, txt=f"Compliance Readiness Score: {compliance_score_prompt:.0f}%", ln=True)
                pdf.ln(10) # Add some space before the note
                pdf.cell(200, 10, txt="Note: Scores are illustrative. Conduct thorough validation.", ln=True) # Add the note

                pdf_bytes = pdf.output(dest='S').encode('latin-1') # Output to string and encode
                st.download_button(
                    label="Download PDF Now",
                    data=pdf_bytes,
                    file_name="prompt_data_compliance_report.pdf",
                    mime="application/pdf",
                    key="download_pdf_prompt_data_final"
                )

            # Download options for generated data
            csv_prompt = st.session_state.prompt_generated_df.to_csv(index=False)

            # Use openpyxl to create Excel file
            excel_buffer = io.BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer_excel:
                st.session_state.prompt_generated_df.to_excel(writer_excel, index=False, sheet_name='Sheet1')
            excel_data = excel_buffer.getvalue()

            col1, col2 = st.columns(2)
            col1, col2, col3 = st.columns(3)
            with col1:
                st.download_button(
                    label="Download CSV",
                    data=csv_prompt,
                    file_name="synthetic_data_prompt.csv", # Use the intended file name
                    mime="text/csv",
                    key="download_csv_prompt"
                )
            with col2:
                st.download_button(
                    label="Download Excel",
                    data=excel_data,
                    file_name="synthetic_data_prompt.xlsx", # Use the intended file name
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_excel_prompt"
                )
            with col3:
                explain_context_prompt = {
                    "method": "Text Prompt",
                    "prompt": st.session_state.get('last_processed_prompt_tab1', 'N/A')
                }
                explain_dfs_info_prompt = {
                    "PromptGeneratedTable": {"rows": st.session_state.prompt_generated_df.shape[0], "cols": st.session_state.prompt_generated_df.shape[1]}
                }
                # Add DP info to context if it was applied
                if st.session_state.get('advanced_lab_selection') == "🛡️ Differential Privacy":
                    explain_context_prompt['dp_applied'] = True
                    explain_context_prompt['dp_epsilon'] = st.session_state.get('dp_epsilon', 'N/A')
                    explain_context_prompt['dp_mechanism_numeric'] = st.session_state.get('dp_mechanism_numeric', 'N/A')
                    explain_context_prompt['dp_mechanism_categorical'] = st.session_state.get('dp_mechanism_categorical', 'N/A')

                explain_pdf_data_prompt = generate_explainability_pdf(explain_context_prompt, explain_dfs_info_prompt)
                st.download_button(
                    label="Download Explainability Report",
                    data=explain_pdf_data_prompt,
                    file_name="explainability_report_prompt.pdf",
                    mime="application/pdf",
                    key="download_explain_pdf_prompt"
                )

with tab2:
    # Smart Schema Editor tab
    synth_df = None
    rows = 10
    # Check if there's data from prompt generation first
    if 'prompt_generated_df' in st.session_state and st.session_state.prompt_generated_df is not None:
         synth_df = st.session_state.prompt_generated_df
         rows = max(1, len(synth_df)) # Use synth_df which is now populated
    # If not, check if there's data from file upload
    elif 'uploaded_df_for_schema' in st.session_state and st.session_state.uploaded_df_for_schema is not None:
        synth_df = st.session_state.uploaded_df_for_schema
        rows = max(1, len(synth_df))
    # If schema_df (old single table) exists from a previous schema generation, use its row count
    elif 'schema_df' in st.session_state and st.session_state.schema_df is not None:
         rows = max(1, len(st.session_state.schema_df))


    show_smart_schema_editor(synth_df, rows)

with tab3:
    st.header("File-based Synthetic Data Generation")
    uploaded_file = st.file_uploader("Upload your CSV or XLSX dataset", type=["csv", "xlsx"], key="file_uploader_tab3")

    if uploaded_file:
        try:
            # File Size Check
            MAX_APP_FILE_SIZE = 50 * 1024 * 1024 # 50 MB
            if uploaded_file.size > MAX_APP_FILE_SIZE:
                st.error(f"File is too large ({uploaded_file.size / (1024*1024):.1f} MB). Maximum allowed size is {MAX_APP_FILE_SIZE / (1024*1024):.1f} MB.")
                # Clear related session state if file is rejected
                st.session_state.uploaded_df_for_schema = None
                st.session_state.file_action_tab3 = None
                st.session_state.newly_generated_df_tab3 = None
                st.session_state.synthetic_df_from_file_tab3 = None
                st.session_state.uploaded_file_name_tab3 = None # Clear the filename
                st.stop() # Stop further processing in this tab for this run

            if uploaded_file.name.endswith(".csv"):
                df = pd.read_csv(uploaded_file)
            else:
                df = pd.read_excel(uploaded_file)
        except pd.errors.ParserError as pe:
            st.error(f"Error parsing the file: {pe}. Please ensure the file is a valid CSV or Excel file.")
            df = None
        except MemoryError:
            st.error("Memory Error: The file is too large or complex to process with available memory. Try a smaller file.")
            df = None
        except Exception as e:
            st.error(f"An unexpected error occurred while processing the file: {e}")
            df = None

        if df is None:
            # Ensure related states are reset if file processing failed
            st.session_state.uploaded_df_for_schema = None
            st.session_state.file_action_tab3 = None
            st.session_state.newly_generated_df_tab3 = None
            st.session_state.synthetic_df_from_file_tab3 = None
            # return # Stop further processing in this tab if file load failed
        else:
            # If df is loaded successfully, proceed
            st.session_state.uploaded_df_for_schema = df # Store for tab2 access
            # PII Detection on Uploaded File
            pii_columns = [col for col in df.columns if col in PII_FIELDS]
            dpdp_columns = [col for col in df.columns if is_dpdp_pii(col)]

            if pii_columns:
                st.warning(f"⚠️ Uploaded dataset contains potential PII columns: {', '.join(pii_columns)}")

            # Display DPDP-specific warnings
            if dpdp_columns:
                for field in dpdp_columns:
                    st.markdown(f"""
                    <div class="dpdp-warning">
                    🔐 <strong>DPDP Compliance Note</strong>: "{field}" contains PII under India's DPDP Act
                    </div>
                    """, unsafe_allow_html=True)

            st.subheader("📊 Preview of Original Data")
            st.dataframe(df.head())

            # Reset choice if a new file is uploaded
            if st.session_state.get('uploaded_file_name_tab3') != uploaded_file.name:
                st.session_state.file_action_tab3 = None
                st.session_state.uploaded_file_name_tab3 = uploaded_file.name
                # Clear synthetic data derived from a *previous* file
                st.session_state.newly_generated_df_tab3 = None
                st.session_state.synthetic_df_from_file_tab3 = None
                # Also clear data from other generation tabs as a new base file is being introduced
                st.session_state.prompt_generated_df = None
                st.session_state.generated_data_frames = {}
                st.session_state.active_display_table_name = None
                st.session_state.playground_generated_df = None
                st.session_state.num_rows_for_file_upload_tab3 = len(df) # Reset num_rows default

            st.markdown("---")
            if st.session_state.file_action_tab3 is None:
                st.subheader("What would you like to do with the uploaded file?")
                choice_col1, choice_col2 = st.columns(2)
                with choice_col1:
                    if st.button("Generate Synthetic Data", key="choice_generate_data_tab3", use_container_width=True):
                        st.session_state.file_action_tab3 = "generate"
                        st.rerun()
                with choice_col2:
                    if st.button("Check Compliance of Original File", key="choice_check_compliance_tab3", use_container_width=True):
                        st.session_state.file_action_tab3 = "compliance"
                        st.rerun()
            
            if st.session_state.file_action_tab3 == "generate":
                st.subheader("⚙️ Configure Synthetic Data Generation")
                st.number_input( # The widget now directly manages st.session_state.num_rows_for_file_upload_tab3
                    "Number of *additional* synthetic rows to generate:",
                    min_value=1,
                    help="These rows will be generated based on the schema of your uploaded file and appended to it.",
                    step=max(1, len(df) // 10 if len(df) > 10 else 1),
                    key="num_rows_for_file_upload_tab3" # Use the session state variable itself as the key
                )

                if st.button("Generate Synthetic Data", key="generate_synthetic_from_file_tab3", use_container_width=True):
                    num_to_generate = st.session_state.num_rows_for_file_upload_tab3
                    new_synthetic_data = {}
                    for col_name in df.columns: # Use columns from the original uploaded df
                        original_column_series = df[col_name] # Base characteristics on original column
                        if pd.api.types.is_numeric_dtype(original_column_series):
                            new_synthetic_data[col_name] = _synthesize_numeric_column_from_upload(original_column_series, col_name, num_to_generate)
                        elif pd.api.types.is_object_dtype(original_column_series) or pd.api.types.is_categorical_dtype(original_column_series):
                            new_synthetic_data[col_name] = _synthesize_categorical_column_from_upload(original_column_series, col_name, num_to_generate)
                        else:
                            if not original_column_series.dropna().empty:
                                # For other types, sample with replacement from the original column to generate new rows
                                new_values = np.random.choice(original_column_series.dropna().values, size=num_to_generate, replace=True)
                                new_synthetic_data[col_name] = pd.Series(new_values, name=col_name).astype(original_column_series.dtype, errors='ignore')
                            else:
                                new_synthetic_data[col_name] = pd.Series([pd.NA] * num_to_generate, name=col_name).astype(original_column_series.dtype, errors='ignore')

                    newly_generated_df = pd.DataFrame(new_synthetic_data)
                    
                    st.session_state.newly_generated_df_tab3 = newly_generated_df # Store for drift detection
                    # Concatenate original df with the newly generated_df
                    combined_df = pd.concat([df, newly_generated_df], ignore_index=True)

                    # Clear data from other generation paths
                    st.session_state.prompt_generated_df = None
                    st.session_state.generated_data_frames = {} # From Smart Schema
                    st.session_state.active_display_table_name = None
                    st.session_state.playground_generated_df = None
                    st.session_state.synthetic_df_from_file_tab3 = combined_df # Store the combined df

                if 'synthetic_df_from_file_tab3' in st.session_state and st.session_state.synthetic_df_from_file_tab3 is not None:
                    synthetic_df_from_file = st.session_state.synthetic_df_from_file_tab3
                    st.success(f"Successfully generated {len(synthetic_df_from_file) - len(df)} additional synthetic rows and appended to original data!")
                    st.dataframe(synthetic_df_from_file) # Show all generated data
                    st.markdown(f"**Rows:** {synthetic_df_from_file.shape[0]} | **Columns:** {synthetic_df_from_file.shape[1]}")

                    pii_cols_synthetic_file = [col for col in synthetic_df_from_file.columns if col in PII_FIELDS or is_dpdp_pii(col)]
                    dpdp_cols_synthetic_file = [col for col in synthetic_df_from_file.columns if is_dpdp_pii(col)]

                    if pii_cols_synthetic_file:
                        st.warning(f"⚠️ Generated synthetic data (from file) contains PII-like columns: {', '.join(pii_cols_synthetic_file)}")
                    if dpdp_cols_synthetic_file:
                        for field in dpdp_cols_synthetic_file:
                            st.markdown(f""" <div class="dpdp-warning"> 🔐 <strong>DPDP Compliance Note</strong>: "{field}" (in synthetic data) contains PII under India's DPDP Act </div> """, unsafe_allow_html=True)

                    csv_syn_file = synthetic_df_from_file.to_csv(index=False)
                    excel_buffer_syn_file = io.BytesIO()
                    with pd.ExcelWriter(excel_buffer_syn_file, engine='openpyxl') as writer_excel_syn_file:
                        synthetic_df_from_file.to_excel(writer_excel_syn_file, index=False, sheet_name='Sheet1')
                    excel_data_syn_file = excel_buffer_syn_file.getvalue()

                    dl_col1, dl_col2 = st.columns(2)
                    with dl_col1:
                        st.download_button(label="Download CSV (Synthetic)", data=csv_syn_file, file_name="synthetic_data_from_file.csv", mime="text/csv", key="download_csv_synthetic_file")
                    with dl_col2:
                        st.download_button(label="Download Excel (Synthetic)", data=excel_data_syn_file, file_name="synthetic_data_from_file.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="download_excel_synthetic_file")
                    
                    # --- Data Drift Analysis ---
                    if 'newly_generated_df_tab3' in st.session_state and st.session_state.newly_generated_df_tab3 is not None:
                        original_df_for_drift = df # Uploaded original
                        synthetic_part_df_for_drift = st.session_state.newly_generated_df_tab3

                        st.markdown("---")
                        st.subheader("🔬 Data Drift Analysis (Original vs. Synthetic Part)")
                        drift_messages = []
                        drift_count = 0
                        for col in original_df_for_drift.columns:
                            if col not in synthetic_part_df_for_drift.columns:
                                drift_messages.append(f"Column '{col}' missing in synthetic part, cannot compare.")
                                continue

                            if pd.api.types.is_numeric_dtype(original_df_for_drift[col]):
                                drifted, msg = detect_numerical_drift(original_df_for_drift[col], synthetic_part_df_for_drift[col], col)
                            else: # Assume categorical for others
                                drifted, msg = detect_categorical_drift(original_df_for_drift[col], synthetic_part_df_for_drift[col], col)
                            
                            if drifted:
                                st.warning(msg)
                                drift_count += 1
                            drift_messages.append(msg) # Store all messages for a summary or detailed report later if needed
                        st.info(f"Drift analysis complete. {drift_count} column(s) showed significant drift out of {len(original_df_for_drift.columns)}.")

            elif st.session_state.file_action_tab3 == "compliance":
                st.markdown("---")
                st.subheader("Compliance Analysis of Original Uploaded File")
                # Bias Checker for original uploaded data
                st.subheader("📈 Bias Checker (Original Uploaded Data)")
                if not df.empty:
                    categorical_cols_upload = df.select_dtypes(include='object').columns.tolist()
                    if not categorical_cols_upload:
                        categorical_cols_upload = [col for col in df.columns if df[col].nunique() < 20 and df[col].nunique() > 1]
                    
                    if categorical_cols_upload:
                        selected_col_upload = st.selectbox("Select a categorical column to check bias", categorical_cols_upload, key="bias_checker_upload_tab3_compliance")
                        if selected_col_upload:
                            value_counts_upload = df[selected_col_upload].value_counts(normalize=True) * 100
                            fig_upload = px.bar(value_counts_upload, x=value_counts_upload.index, y=value_counts_upload.values, labels={"x": selected_col_upload, "y": "Percentage"}, title=f"Distribution in '{selected_col_upload}' (Original Data)")
                            st.plotly_chart(fig_upload)
                    else:
                        st.info("No suitable categorical columns found in the uploaded data for bias checking.")
                else:
                    st.info("Error: Original data seems to be empty for bias checking.")

                # Ethical Scorecard for original uploaded data
                st.subheader("✅ Ethical Scorecard (Original Uploaded Data Analysis)")
                score_col1, score_col2, score_col3 = st.columns(3)
                pii_cols_uploaded_orig = [col for col in df.columns if col in PII_FIELDS or is_dpdp_pii(col)]
                dpdp_cols_uploaded_specific_orig = [col for col in df.columns if is_dpdp_pii(col)]
                pii_risk_level_uploaded_orig = "High" if pii_cols_uploaded_orig else "Low"
                dpdp_risk_level_uploaded_orig = "High" if dpdp_cols_uploaded_specific_orig else "Low"
                bias_score_uploaded_orig = calculate_bias_score(df)

                score_col1.metric("Bias Score (Original)", f"{bias_score_uploaded_orig:.0f} / 100")
                score_col2.metric("PII Risk (Original)", pii_risk_level_uploaded_orig)
                score_col3.metric("DPDP Risk (Original)", dpdp_risk_level_uploaded_orig)

                # Download PDF Report for original uploaded data
                st.subheader("📄 Download Summary Report (Original Uploaded Data)")
                if st.button("Download PDF Report", key="pdf_report_upload_tab3_compliance", use_container_width=True):
                    pdf_upload = FPDF()
                    pdf_upload.add_page()
                    pdf_upload.set_font("Arial", size=12)
                    pdf_upload.cell(200, 10, txt=f"{APP_NAME} - Original Data Report", ln=True, align="C")
                    pdf_upload.ln(10)
                    pdf_upload.cell(200, 10, txt=f"Dataset: {uploaded_file.name}", ln=True)
                    pdf_upload.cell(200, 10, txt=f"Bias Score: {bias_score_uploaded_orig:.0f} / 100", ln=True)
                    pdf_upload.cell(200, 10, txt=f"PII Risk: {pii_risk_level_uploaded_orig}", ln=True)
                    if pii_cols_uploaded_orig: pdf_upload.cell(200, 10, txt=f"Detected PII Fields: {', '.join(pii_cols_uploaded_orig)}", ln=True)
                    pdf_upload.cell(200, 10, txt=f"DPDP Risk: {dpdp_risk_level_uploaded_orig}", ln=True)
                    if dpdp_cols_uploaded_specific_orig: pdf_upload.cell(200, 10, txt=f"Detected DPDP-Specific Fields: {', '.join(dpdp_cols_uploaded_specific_orig)}", ln=True)
                    pdf_upload.ln(10)
                    pdf_upload.cell(200, 10, txt="Note: This report reflects analysis of the original uploaded data.", ln=True)
                    buffer_upload_pdf = io.BytesIO()
                    pdf_upload_bytes = pdf_upload.output(dest='S').encode('latin-1') # Output to string and encode
                    st.download_button(label="Download PDF Now", data=pdf_upload_bytes, file_name="original_data_report.pdf", mime="application/pdf", key="download_pdf_upload_final_tab3_compliance")

with tab4:
    st.header("🎲 Scenario & Edge Case Playground")
    st.markdown("Design and test specific data scenarios or edge cases quickly. Select a base schema template, define your edge cases, and generate a small dataset to see the results.")

    # 1. Select Schema Template
    st.subheader("1. Base Schema for Scenario")
    playground_template_options = list(SCHEMA_TEMPLATES.keys())
    
    current_pg_template_idx = 0
    if st.session_state.playground_selected_template_name in playground_template_options:
        current_pg_template_idx = playground_template_options.index(st.session_state.playground_selected_template_name)
    
    selected_pg_template = st.selectbox(
        "Select a base schema template:",
        options=playground_template_options,
        index=current_pg_template_idx,
        key="playground_template_selector"
    )

    if selected_pg_template != st.session_state.playground_selected_template_name:
        st.session_state.playground_selected_template_name = selected_pg_template
        if selected_pg_template != "None (Custom Schema)":
            template_data_pg = SCHEMA_TEMPLATES[selected_pg_template]
            st.session_state.playground_schema_fields = [] # Clear previous
            default_global_pii_strategy = st.session_state.get(DEFAULT_PII_STRATEGY_KEY, "realistic_fake")
            for field_template in template_data_pg.get("fields", []): # Access the 'fields' key
                new_field = field_template.copy()
                canonical_suggestion = new_field.pop("_canonical_suggestion", None)
                if canonical_suggestion and canonical_suggestion in CANONICAL_FIELD_TO_SCHEMA_DETAILS_MAP:
                    canonical_details = CANONICAL_FIELD_TO_SCHEMA_DETAILS_MAP[canonical_suggestion]
                    new_field["type"] = canonical_details.get("type", new_field["type"])
                    new_field["constraint"] = canonical_details.get("constraint", new_field["constraint"])
                is_sensitive = new_field["type"] in ["email", "phone", "aadhaar", "pan", "passport", "voterid", "ifsc", "upi", "name", "address"]
                new_field["pii_handling"] = new_field.get("pii_handling", default_global_pii_strategy if is_sensitive else "realistic_fake")
                st.session_state.playground_schema_fields.append(new_field)
        else:
            st.session_state.playground_schema_fields = [] # Clear if "None" is selected
        st.session_state.playground_generated_df = None # Clear previous results
        st.session_state.playground_edge_cases = [] # Clear edge cases when template changes
        st.rerun()

    if st.session_state.playground_schema_fields:
        st.markdown("**Selected Schema Fields:**")
        for field in st.session_state.playground_schema_fields:
            st.caption(f"- {field['name']} ({FIELD_TYPES.get(field['type'], field['type'])})")
    elif st.session_state.playground_selected_template_name != "None (Custom Schema)":
        st.info("Selected template has no fields or an issue occurred. Please select another.")
    else:
        st.info("Please select a schema template to define fields for your scenario.")

    # 2. Define Edge Cases for this Scenario
    playground_schema_for_edge_cases = {st.session_state.playground_table_name_for_conditions: st.session_state.playground_schema_fields}
    render_edge_case_ui(
        ui_key_prefix="playground",
        edge_cases_list_key='playground_edge_cases',
        table_schemas_for_fields=playground_schema_for_edge_cases,
        available_table_names_for_conditions=[st.session_state.playground_table_name_for_conditions] if st.session_state.playground_schema_fields else []
    )

    # 3. Generate Scenario Data
    st.subheader("3. Generate Scenario Data")
    st.session_state.playground_num_rows = st.number_input("Number of rows for scenario:", min_value=1, max_value=1000, value=st.session_state.playground_num_rows, step=5, key="pg_num_rows")

    if st.button("🔄 Generate Scenario Data", key="generate_playground_data", disabled=not st.session_state.playground_schema_fields):
        if st.session_state.playground_schema_fields:
            st.session_state.playground_generated_df = generate_single_table_data_with_edge_cases(
                schema_fields=st.session_state.playground_schema_fields,
                num_rows=st.session_state.playground_num_rows,
                edge_cases_list=st.session_state.playground_edge_cases,
                pii_strategy_global=st.session_state.get(DEFAULT_PII_STRATEGY_KEY, "realistic_fake"),
                table_name_for_conditions=st.session_state.playground_table_name_for_conditions
            )
            if st.session_state.playground_generated_df is not None:
                # Clear data from other main generation paths
                st.session_state.prompt_generated_df = None
                st.session_state.uploaded_df_for_schema = None
                st.session_state.generated_data_frames = {}
                st.session_state.active_display_table_name = None
                st.session_state.newly_generated_df_tab3 = None
                st.session_state.synthetic_df_from_file_tab3 = None
                st.success("Scenario data generated!")
            else:
                st.error("Failed to generate scenario data. Check if schema is valid.")
        else:
            st.warning("Please select a schema template with fields before generating data.")

    # 4. Display Generated Scenario Data
    if st.session_state.playground_generated_df is not None:
        st.subheader("📊 Generated Scenario Data")
        st.dataframe(st.session_state.playground_generated_df)
        
        csv_playground = st.session_state.playground_generated_df.to_csv(index=False).encode('utf-8')
        st.download_button(
            label="Download Scenario CSV",
            data=csv_playground,
            file_name="scenario_playground_data.csv",
            mime="text/csv",
            key="download_csv_playground"
        )

with tab5:
    st.header("🏛️ Community Schema Gallery")
    st.markdown("Explore and load predefined schema templates to kickstart your data generation in the **Smart Schema Editor**.")
    st.markdown("---")

    for template_name, template_data in SCHEMA_TEMPLATES.items():
        if template_name == "None (Custom Schema)": # Skip the placeholder
            continue

        with st.expander(f"**{template_name}** - {template_data.get('description', 'No description available.')}"):
            st.markdown("**Fields in this template:**")
            if template_data.get("fields"):
                for field in template_data["fields"]:
                    st.caption(f"- {field['name']} ({FIELD_TYPES.get(field['type'], field['type'])})")
            else:
                st.caption("No fields defined for this template.")
            
            if st.button(f"Load '{template_name}' into Smart Schema Editor", key=f"load_gallery_{template_name}"):
                # Determine target table in Smart Schema Editor
                target_table_name_for_gallery_load = st.session_state.active_table_name
                if not target_table_name_for_gallery_load or target_table_name_for_gallery_load not in st.session_state.table_schemas:
                    target_table_name_for_gallery_load = "Table1" # Default if no active or valid table
                    if target_table_name_for_gallery_load not in st.session_state.table_schemas:
                        st.session_state.table_schemas[target_table_name_for_gallery_load] = [] # Create if not exists
                
                st.session_state.table_schemas[target_table_name_for_gallery_load] = [] # Clear target table's schema
                default_pii_strategy = st.session_state.get(DEFAULT_PII_STRATEGY_KEY, "realistic_fake")
                
                for field_template in template_data.get("fields", []):
                    new_field = field_template.copy()
                    canonical_suggestion = new_field.pop("_canonical_suggestion", None)
                    if canonical_suggestion and canonical_suggestion in CANONICAL_FIELD_TO_SCHEMA_DETAILS_MAP:
                        details = CANONICAL_FIELD_TO_SCHEMA_DETAILS_MAP[canonical_suggestion]
                        new_field["type"] = details.get("type", new_field.get("type", "string"))
                        new_field["constraint"] = details.get("constraint", new_field.get("constraint", ""))
                    is_sensitive = new_field["type"] in ["email", "phone", "aadhaar", "pan", "passport", "voterid", "ifsc", "upi", "name", "address"]
                    new_field["pii_handling"] = new_field.get("pii_handling", default_pii_strategy if is_sensitive else "realistic_fake")
                    st.session_state.table_schemas[target_table_name_for_gallery_load].append(new_field)
                
                st.session_state.active_table_name = target_table_name_for_gallery_load # Ensure this is the active table
                st.session_state.selected_template_name = template_name # Update Tab 2's dropdown state
                st.session_state.initial_schema_populated = True
                st.session_state.generated_data_frames = {} # Clear any old generated data
                # Clear data from other generation paths
                st.session_state.prompt_generated_df = None
                st.session_state.uploaded_df_for_schema = None
                st.session_state.playground_generated_df = None
                st.session_state.newly_generated_df_tab3 = None
                st.session_state.synthetic_df_from_file_tab3 = None
                st.success(f"Template '{template_name}' loaded into table '{target_table_name_for_gallery_load}' in the Smart Schema Editor. Please navigate there to continue.")
                st.rerun() # Rerun to switch context to Tab 2 if user is on Tab 5

with tab_advanced_lab:
    st.header("🔬 Advanced Lab")
    st.markdown("Explore cutting-edge techniques for synthetic data generation, privacy enhancement, and quality assessment. These features are currently conceptual placeholders for future development.")

    advanced_lab_options = ["🤖 AI-Powered Generation", "🛡️ Differential Privacy", "🏆 Quality Benchmarking", "🤝 Federated Generation", "🌐 Community Marketplace", "☁️ Cloud Deploy (Sandbox)"]
    current_adv_lab_selection_idx = advanced_lab_options.index(st.session_state.advanced_lab_selection)

    st.session_state.advanced_lab_selection = st.radio(
        "Select an Advanced Tool:",
        options=advanced_lab_options,
        index=current_adv_lab_selection_idx,
        key="adv_lab_radio",
        horizontal=True
    )
    st.markdown("---")

    if st.session_state.advanced_lab_selection == "🤖 AI-Powered Generation":
        # Content from old tab_deep_gen
        st.subheader("🚀 AI-Powered Generation (Deep Generative Models)")
        st.markdown("""
        This section is for integrating advanced deep learning models like Generative Adversarial Networks (GANs),
        Variational Autoencoders (VAEs), Transformer-based models (e.g., GPT for text), and diffusion models (for images)
        for synthetic data generation. These models can produce highly realistic and complex data,
        capturing intricate patterns from original datasets, generating creative text, or even synthesizing images.
        or generating creative text based on prompts.

        **Current Status:** This is a placeholder for future development. Full integration of these models
        requires significant backend infrastructure and model training capabilities.
        """)
        st.info("✨ **Coming Soon:** Full integration of deep generative models for cutting-edge synthetic data.")

        st.session_state.deep_model_type = st.selectbox(
            "Select Deep Generative Model Type:",
            options=["GAN (Tabular)", "VAE (Tabular)", "Transformer/GPT (Text/Structured from Prompt)", "NLP Model (Advanced Text Generation)", "Image Generation Model (e.g., Diffusion)"],
            key="deep_model_selector",
            help="Choose the type of deep learning model to use."
        )

        if st.session_state.deep_model_type in ["GAN (Tabular)", "VAE (Tabular)"]:
            st.markdown(f"""
            {st.session_state.deep_model_type} models learn the underlying patterns from an existing tabular dataset to generate new, synthetic samples.
            You would typically upload your real dataset here for the model to train on.
            """)
            uploaded_deep_model_file = st.file_uploader(
                f"Upload Training Data (CSV/XLSX) for {st.session_state.deep_model_type}",
                type=["csv", "xlsx"],
                key="deep_model_file_uploader"
            )
            if uploaded_deep_model_file:
                st.success(f"File '{uploaded_deep_model_file.name}' uploaded. In a full implementation, this would be used for training.")
                st.session_state.deep_model_data_uploaded = True
            else:
                st.session_state.deep_model_data_uploaded = False

            st.number_input("Number of Synthetic Rows to Generate:", min_value=10, value=100, key="deep_model_num_rows")
            st.text_input("Model Parameters (e.g., epochs, batch_size - conceptual):", value="epochs=100, batch_size=32", key="deep_model_params_tabular", disabled=True)

            if st.button(f"Train & Generate with {st.session_state.deep_model_type}", key="deep_model_generate_tabular_btn", disabled=not st.session_state.deep_model_data_uploaded):
                st.info(f"Placeholder: Simulating training of {st.session_state.deep_model_type} and generation of data. This would involve significant computation in a real scenario.")
                st.success("Conceptual generation complete! (No actual data generated in this placeholder).")

        elif st.session_state.deep_model_type == "Transformer/GPT (Text/Structured from Prompt)":
            st.markdown("""
            Transformer-based models like GPT can generate coherent text or even structured data (like JSON or code snippets)
            based on a detailed natural language prompt.
            """)
            st.session_state.deep_model_prompt = st.text_area("Enter your detailed prompt for generation:", height=150, value=st.session_state.deep_model_prompt, key="deep_model_prompt_input")
            st.text_input("Model Parameters (e.g., temperature, max_tokens - conceptual):", value="temperature=0.7, max_tokens=500", key="deep_model_params_gpt", disabled=True)

            if st.button("Generate with Transformer/GPT Model", key="deep_model_generate_gpt_btn", disabled=not st.session_state.deep_model_prompt):
                st.info("Placeholder: Simulating generation with a Transformer/GPT model. This would involve API calls or local model inference in a real scenario.")
                st.success(f"Conceptual generation based on prompt: '{st.session_state.deep_model_prompt[:100]}...' complete! (No actual data generated in this placeholder).")

        elif st.session_state.deep_model_type == "NLP Model (Advanced Text Generation)":
            st.markdown("""
            Advanced NLP models can perform a variety of text generation tasks beyond simple prompting, such as summarization,
            translation, question answering based on context, or more elaborate creative writing.
            """)
            st.session_state.nlp_model_task = st.selectbox("Select NLP Task (Conceptual):", ["Creative Writing", "Summarization", "Question Answering", "Code Generation"], key="nlp_task_selector")
            st.session_state.nlp_model_prompt = st.text_area("Enter your detailed prompt or context for the NLP model:", height=150, value=st.session_state.nlp_model_prompt, key="nlp_model_prompt_input")
            st.text_input("NLP Model Parameters (e.g., model_name, max_length - conceptual):", value="model_name=gpt-3.5-turbo, max_length=1024", key="nlp_model_params", disabled=True)

            if st.button("Generate Text with Advanced NLP Model", key="nlp_model_generate_btn", disabled=not st.session_state.nlp_model_prompt):
                st.info(f"Placeholder: Simulating '{st.session_state.nlp_model_task}' with an advanced NLP model. This would involve significant computation or API calls.")
                st.success(f"Conceptual NLP generation for task '{st.session_state.nlp_model_task}' complete! (No actual text generated in this placeholder).")

        elif st.session_state.deep_model_type == "Image Generation Model (e.g., Diffusion)":
            st.markdown("""
            Image generation models (like DALL-E, Stable Diffusion, Midjourney) can create novel images from textual descriptions.
            This can be useful for generating synthetic image datasets for computer vision tasks.
            """)
            st.session_state.image_model_prompt = st.text_input("Enter your image prompt (e.g., 'a photo of an astronaut riding a horse on the moon'):", value=st.session_state.image_model_prompt, key="image_model_prompt_input")
            st.text_input("Image Model Parameters (e.g., resolution, style - conceptual):", value="resolution=512x512, style=photorealistic", key="image_model_params", disabled=True)

            if st.button("Generate Image with Model", key="image_model_generate_btn", disabled=not st.session_state.image_model_prompt):
                st.info("Placeholder: Simulating image generation. This would involve API calls or powerful local GPU inference.")
                st.image("https://via.placeholder.com/256x256.png?text=Conceptual+Image", caption=f"Conceptual image for: '{st.session_state.image_model_prompt}' (Placeholder)")
                st.success("Conceptual image generation complete!")

    elif st.session_state.advanced_lab_selection == "🛡️ Differential Privacy":
        # Content from old tab_adv_privacy
        st.subheader("🛡️ Advanced Privacy: Differential Privacy (DP)")
        st.markdown("""
        Differential Privacy (DP) offers a strong, mathematical guarantee of privacy. It ensures that the output of a data analysis or synthetic data generation process does not significantly change whether any single individual's data is included or excluded from the original dataset. This is the gold standard for privacy and a major requirement for many enterprise and government applications.

        **Key Concepts:**
        - **Epsilon (ε):** The privacy loss parameter. A smaller epsilon means stronger privacy (less information leakage about individuals).
        - **Delta (δ):** The probability that the privacy guarantee might be broken. Often set to a very small number. (For simplicity, we are focusing on ε-DP here).

        **Current Status:** This section is a placeholder for future development. Implementing true Differential Privacy is complex and requires specialized algorithms and careful calibration.
        """)
        st.info("✨ **Coming Soon:** Full integration of Differential Privacy mechanisms for provably private synthetic data.")

        st.markdown("Below are conceptual controls for how DP might be applied. Actual implementation would involve adding calibrated noise during the data generation or aggregation process.")
        st.session_state.dp_epsilon = st.number_input(
            "Privacy Budget - Epsilon (ε):",
            min_value=0.01, max_value=10.0, value=st.session_state.dp_epsilon, step=0.1,
            help="Lower epsilon = stronger privacy. Common values range from 0.1 to 1.0.",
            key="dp_epsilon_input"
        )
        st.session_state.dp_mechanism_numeric = st.selectbox(
            "Conceptual DP Mechanism for Numerical Data:",
            options=["Laplace Mechanism", "Gaussian Mechanism"],
            index=["Laplace Mechanism", "Gaussian Mechanism"].index(st.session_state.dp_mechanism_numeric),
            key="dp_mech_numeric_selector",
            help="Technique to add noise to numerical values to achieve DP."
        )
        st.session_state.dp_mechanism_categorical = st.selectbox(
            "Conceptual DP Mechanism for Categorical Data:",
            options=["Randomized Response", "Exponential Mechanism (for counts/histograms)"],
            index=["Randomized Response", "Exponential Mechanism (for counts/histograms)"].index(st.session_state.dp_mechanism_categorical),
            key="dp_mech_categorical_selector",
            help="Technique to perturb categorical values or their distributions for DP."
        )
        st.markdown("**Note on Data Utility:** Applying Differential Privacy involves adding noise, which can reduce the utility (accuracy or realism) of the synthetic data. There's always a trade-off between privacy (ε) and utility.")
        if st.button("Generate Data with Conceptual DP", key="dp_generate_btn"):
            st.info(f"Placeholder: Simulating data generation with Differential Privacy (ε={st.session_state.dp_epsilon}). This would involve applying noise based on the selected mechanisms and epsilon to the generation process.")
            st.success("Conceptual DP-aware generation complete! (No actual data generated in this placeholder).")
        st.markdown("A real DP implementation would typically provide a report detailing: Epsilon (ε) and Delta (δ) values, mechanisms used, assumptions, and interpretation guidance.")

    elif st.session_state.advanced_lab_selection == "🏆 Quality Benchmarking":
        # Content from old tab_quality
        st.subheader("🏆 Synthetic Data Quality Benchmarking")
        st.markdown("""
        This section is dedicated to evaluating the quality of the generated synthetic data against original (real) data.
        Benchmarking helps build trust and allows users to understand the trade-offs of different generation methods.
        It typically involves assessing:
        - **Fidelity:** How closely the statistical properties of synthetic data match the real data.
        - **Utility:** How well the synthetic data performs for a specific task (e.g., training a machine learning model) compared to real data.
        - **Privacy:** How well the synthetic data protects individual privacy from the original dataset.

        **Current Status:** This is a placeholder for future development. Full integration requires a reference (real) dataset and implementation of various metrics, potentially using libraries like `sdmetrics`.
        """)
        st.info("✨ **Coming Soon:** Automated benchmarking with detailed reports on fidelity, utility, and privacy.")

        original_data_available = 'uploaded_file_name_tab3' in st.session_state and st.session_state.uploaded_file_name_tab3 is not None and \
                                  'uploaded_df_for_schema' in st.session_state and st.session_state.uploaded_df_for_schema is not None
        synthetic_data_available_tab3 = 'newly_generated_df_tab3' in st.session_state and st.session_state.newly_generated_df_tab3 is not None

        if original_data_available and synthetic_data_available_tab3:
            st.success(f"Original data ('{st.session_state.uploaded_file_name_tab3}') and its synthetic counterpart are available from 'File-based Generation'.")
            st.session_state.benchmark_data_source = "File-based Generation Output"
        else:
            st.warning("To perform benchmarking, you typically need an original (real) dataset and a synthetic dataset generated from it. "
                       "The most direct way to get this is by using the 'File-based Generation' tab to upload your data and then generate a synthetic version.")
            st.session_state.benchmark_data_source = None

        if st.button("Run Conceptual Benchmark Analysis", key="run_benchmark_btn", disabled=(not original_data_available or not synthetic_data_available_tab3)):
            st.info("Placeholder: Simulating benchmark analysis. This would involve complex calculations and comparisons.")
            st.session_state.benchmark_fidelity_score = random.uniform(60, 95)
            st.session_state.benchmark_utility_score = random.uniform(50, 90)
            st.session_state.benchmark_privacy_score = random.uniform(70, 98)
            st.success("Conceptual benchmark analysis complete!")

        if st.session_state.benchmark_fidelity_score is not None:
            st.markdown(f"**Data Source for Benchmark:** {st.session_state.get('benchmark_data_source', 'N/A')}")
            q_col1, q_col2, q_col3 = st.columns(3)
            q_col1.metric("Fidelity Score", f"{st.session_state.benchmark_fidelity_score:.1f}/100", help="Statistical similarity. Higher is better.")
            q_col2.metric("ML Utility Score", f"{st.session_state.benchmark_utility_score:.1f}/100", help="ML model performance. Higher is better.")
            q_col3.metric("Privacy Protection Score", f"{st.session_state.benchmark_privacy_score:.1f}/100", help="Resistance to re-identification. Higher is better.")
            st.markdown("Conceptual Metrics: Column Shape Similarity, Correlation Similarity, ML Task Performance, Membership Inference Resistance.")
            st.caption("Actual implementation would use libraries like `sdmetrics`.")

    elif st.session_state.advanced_lab_selection == "🤝 Federated Generation":
        st.subheader("🤝 Federated Synthetic Data Generation")
        st.markdown("""
        Federated Synthetic Data Generation applies principles from federated learning to enable multiple organizations
        to collaboratively generate a rich synthetic dataset **without sharing their raw, sensitive data directly**. Each organization
        would train a local model on its data, and these models (or their parameters/updates) are securely aggregated
        to create a global synthetic data generator.

        **Why it's important (especially for India):**
        - **Healthcare:** Hospitals can collaborate to create diverse synthetic patient datasets for research without exposing actual patient records.
        - **Fintech:** Banks and financial institutions can generate synthetic transaction data to improve fraud detection models or understand market trends collectively.
        - **Government:** Different government departments can contribute to synthetic datasets for policy making, urban planning, or public service improvement while respecting data silos and privacy mandates.

        **Current Status:** This is a highly advanced feature and a placeholder for future exploration.
        Implementing federated generation requires robust infrastructure for:
        - Secure multi-party computation.
        - Distributed model training and aggregation.
        - Strong privacy-preserving techniques (like differential privacy) applied at each step.
        """)
        st.info("✨ **Vision for the Future:** Enabling collaborative, privacy-preserving synthetic data ecosystems.")
        st.markdown("---")

        st.subheader("1. Define Participants (Conceptual)")
        participant_cols = st.columns([3,1])
        st.session_state.federated_new_participant_name = participant_cols[0].text_input(
            "New Participant Name (e.g., Hospital A, Bank X)", 
            value=st.session_state.federated_new_participant_name,
            key="fed_new_participant_name_input"
        )
        if participant_cols[1].button("Add Participant", key="fed_add_participant_btn", use_container_width=True):
            if st.session_state.federated_new_participant_name and not any(p['name'] == st.session_state.federated_new_participant_name for p in st.session_state.federated_participants_list):
                st.session_state.federated_participants_list.append({"name": st.session_state.federated_new_participant_name, "trained": False})
                st.session_state.federated_new_participant_name = "" # Clear input
                st.session_state.federated_aggregation_done = False # Reset aggregation if participants change
                st.session_state.federated_generated_df_output = None
                st.rerun()
            elif not st.session_state.federated_new_participant_name:
                st.warning("Participant name cannot be empty.")
            else:
                st.warning(f"Participant '{st.session_state.federated_new_participant_name}' already exists.")

        if st.session_state.federated_participants_list:
            st.markdown("**Current Participants:**")
            for i, p_info in enumerate(st.session_state.federated_participants_list):
                p_cols = st.columns([3,2,1])
                p_cols[0].write(f"- **{p_info['name']}** (Status: {'Model Trained' if p_info['trained'] else 'Awaiting Training'})")
                if not p_info['trained']:
                    if p_cols[1].button(f"Simulate Local Training for {p_info['name']}", key=f"fed_train_{i}"):
                        with st.spinner(f"Simulating local model training for {p_info['name']}..."):
                            # In a real scenario, this would be a long, complex process.
                            random_sleep_time = random.uniform(1, 3) # Simulate some work
                            time.sleep(random_sleep_time)
                        st.session_state.federated_participants_list[i]['trained'] = True
                        st.session_state.federated_aggregation_done = False # Reset aggregation
                        st.session_state.federated_generated_df_output = None
                        st.rerun()
                else:
                    p_cols[1].success("Local Model Trained 👍")
                
                if p_cols[2].button("Remove", key=f"fed_remove_p_{i}"):
                    st.session_state.federated_participants_list.pop(i)
                    st.session_state.federated_aggregation_done = False
                    st.session_state.federated_generated_df_output = None
                    st.rerun()
            st.markdown("---")

        st.subheader("2. Simulate Model Aggregation")
        all_trained = all(p['trained'] for p in st.session_state.federated_participants_list)
        if not st.session_state.federated_participants_list:
            st.info("Add participants to enable aggregation.")
        elif not all_trained:
            st.info("All participants must complete 'Simulate Local Training' before aggregation.")
        
        if st.button("Simulate Secure Model Aggregation", key="fed_aggregate_btn", disabled=(not st.session_state.federated_participants_list or not all_trained)):
            with st.spinner("Simulating secure aggregation of local models into a global model..."):
                # Real aggregation is complex (FedAvg, secure MPC, etc.)
                time.sleep(random.uniform(1,3)) # Simulate work
            st.session_state.federated_aggregation_done = True
            st.session_state.federated_generated_df_output = None
            st.success("Global model aggregation simulated successfully! Ready to generate data.")
            st.rerun()

        if st.session_state.federated_aggregation_done:
            st.success("✅ Global Model Ready (Simulated)")
            st.markdown("---")
            st.subheader("3. Generate Synthetic Data from Simulated Global Model")

            fed_schema_options = list(SCHEMA_TEMPLATES.keys())
            # Ensure the stored template name is valid, otherwise default
            if st.session_state.federated_global_schema_template_output not in fed_schema_options:
                st.session_state.federated_global_schema_template_output = "None (Custom Schema)"
            
            current_fed_schema_idx = fed_schema_options.index(st.session_state.federated_global_schema_template_output)

            st.session_state.federated_global_schema_template_output = st.selectbox(
                "Select Schema Template for Federated Output:",
                options=fed_schema_options,
                index=current_fed_schema_idx,
                key="fed_output_schema_selector",
                help="This schema will define the structure of the synthetic data generated by the simulated global model."
            )
            st.session_state.federated_num_rows_output = st.number_input("Number of rows for federated output:", min_value=10, value=st.session_state.federated_num_rows_output, key="fed_num_rows_output")

            if st.button("Generate Federated Data (Simulated)", key="fed_generate_output_btn", disabled=(st.session_state.federated_global_schema_template_output == "None (Custom Schema)")):
                if st.session_state.federated_global_schema_template_output != "None (Custom Schema)":
                    template_fields = SCHEMA_TEMPLATES[st.session_state.federated_global_schema_template_output].get("fields", [])
                    if template_fields:
                        # Use the single_table_data_with_edge_cases for simplicity, without edge cases for this simulation
                        st.session_state.federated_generated_df_output = generate_single_table_data_with_edge_cases(
                            schema_fields=template_fields,
                            num_rows=st.session_state.federated_num_rows_output,
                            edge_cases_list=[], # No edge cases for this simulation
                            pii_strategy_global=st.session_state.get(DEFAULT_PII_STRATEGY_KEY, "realistic_fake"),
                            table_name_for_conditions="FederatedOutputTable" # Dummy name
                        )
                        st.success("Federated synthetic data generated (simulated)!")
                    else:
                        st.error(f"Selected template '{st.session_state.federated_global_schema_template_output}' has no fields defined.")
                else:
                    st.warning("Please select a valid schema template for the output.")

            if st.session_state.federated_generated_df_output is not None:
                st.dataframe(st.session_state.federated_generated_df_output)
                csv_fed_output = st.session_state.federated_generated_df_output.to_csv(index=False).encode('utf-8')
                st.download_button(
                    label="Download Federated Output CSV",
                    data=csv_fed_output,
                    file_name="federated_synthetic_data.csv",
                    mime="text/csv",
                    key="download_csv_federated_output"
                )